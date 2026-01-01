"""
Parse Document AI batch output JSON and convert to Excel.

Uses VoterID positions as anchors to group entities into cards.
"""

import json
import re
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox
import subprocess

# Install packages
def install_packages():
    packages = ['openpyxl']
    for pkg in packages:
        try:
            __import__(pkg)
        except ImportError:
            print(f"Installing {pkg}...")
            subprocess.check_call(['uv', 'pip', 'install', pkg])

install_packages()

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


def extract_part_number(filename):
    """Extract part number from filename."""
    if not filename:
        return ''
    match = re.search(r'-TAM-(\d+)-WI', filename, re.IGNORECASE)
    return match.group(1) if match else ''


def get_entity_position(entity):
    """Get page, x, y position from entity."""
    page_anchor = entity.get('pageAnchor', {})
    page_refs = page_anchor.get('pageRefs', [{}])
    if not page_refs:
        return 0, 0, 0
    page_ref = page_refs[0]
    page = int(page_ref.get('page', 0))
    bounding_poly = page_ref.get('boundingPoly', {})
    vertices = bounding_poly.get('normalizedVertices', [{}])
    if not vertices:
        return page, 0, 0
    x = vertices[0].get('x', 0)
    y = vertices[0].get('y', 0)
    return page, x, y


def group_entities_to_cards(entities):
    """
    Group flat entities into voter cards.
    Uses VoterID positions as anchors for each card.
    """
    # Add position info to all entities
    for entity in entities:
        page, x, y = get_entity_position(entity)
        entity['_page'] = page
        entity['_x'] = x
        entity['_y'] = y
        # Determine column (0, 1, or 2)
        if x < 0.33:
            entity['_col'] = 0
        elif x < 0.66:
            entity['_col'] = 1
        else:
            entity['_col'] = 2

    # Find all voterID positions - these anchor each card
    voter_ids = [e for e in entities if e.get('type', '').lower() == 'voterid']

    # Create card slots based on voterID positions
    cards = []
    for vid in voter_ids:
        card = {
            'serial_no': '',
            'voter_id': vid.get('mentionText', '').strip().upper(),
            'name': '',
            'relation_name': '',
            'house_no': '',
            'age': '',
            'gender': '',
            'page': vid['_page'],
            'col': vid['_col'],
            '_y_start': vid['_y'],
            '_y_end': vid['_y'] + 0.10  # Card spans ~10% of page height
        }
        cards.append(card)

    # Sort cards by page, col, y
    cards.sort(key=lambda c: (c['page'], c['col'], c['_y_start']))

    # Update y_end for each card (ends where next card in same page/col starts)
    for i, card in enumerate(cards):
        for j in range(i + 1, len(cards)):
            if cards[j]['page'] == card['page'] and cards[j]['col'] == card['col']:
                card['_y_end'] = cards[j]['_y_start']
                break

    # Assign all other entities to their respective cards
    for entity in entities:
        if entity.get('type', '').lower() == 'voterid':
            continue  # Already processed

        page = entity['_page']
        col = entity['_col']
        y = entity['_y']
        entity_type = entity.get('type', '').lower()
        mention_text = entity.get('mentionText', '').strip()

        # Clean mention text
        mention_text = re.sub(r'\s*[-–]\s*$', '', mention_text)
        mention_text = re.sub(r'^\s*[-–]\s*', '', mention_text)

        # Find the card this entity belongs to
        best_card = None
        for card in cards:
            if card['page'] == page and card['col'] == col:
                if card['_y_start'] <= y < card['_y_end']:
                    best_card = card
                    break
                elif abs(y - card['_y_start']) < 0.12:
                    best_card = card

        if not best_card:
            # Fallback: find closest card in same page/col
            candidates = [c for c in cards if c['page'] == page and c['col'] == col]
            if candidates:
                best_card = min(candidates, key=lambda c: abs(c['_y_start'] - y))

        if best_card:
            if entity_type == 'sno':
                best_card['serial_no'] = mention_text
            elif entity_type == 'name':
                best_card['name'] = mention_text
            elif entity_type == 'relativename':
                best_card['relation_name'] = mention_text
            elif entity_type == 'houseno':
                best_card['house_no'] = mention_text
            elif entity_type == 'age':
                age_match = re.search(r'(\d+)', mention_text)
                if age_match:
                    age = int(age_match.group(1))
                    if 18 <= age <= 120:
                        best_card['age'] = str(age)
            elif entity_type == 'sex':
                if 'பெண்' in mention_text or 'பெண' in mention_text:
                    best_card['gender'] = 'Female'
                elif 'ஆண்' in mention_text or 'ஆண' in mention_text:
                    best_card['gender'] = 'Male'

    return cards


def save_to_excel(cards, output_path, source_name):
    """Save cards to Excel file."""
    part_no = extract_part_number(source_name)

    wb = Workbook()
    ws = wb.active
    ws.title = "Voter Data"

    headers = ['S.No', 'Part No.', 'Voter S.No', 'Voter ID', 'Name',
               'Relation Name', 'House No', 'Age', 'Gender', 'Page']

    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    missing_stats = {'name': 0, 'age': 0, 'gender': 0, 'voter_id': 0}

    for row_num, card in enumerate(cards, 2):
        ws.cell(row=row_num, column=1, value=row_num - 1)
        ws.cell(row=row_num, column=2, value=part_no)
        ws.cell(row=row_num, column=3, value=card.get('serial_no', ''))

        vid = card.get('voter_id', '')
        vc = ws.cell(row=row_num, column=4, value=vid)
        if not vid:
            vc.fill = yellow_fill
            missing_stats['voter_id'] += 1

        name = card.get('name', '')
        nc = ws.cell(row=row_num, column=5, value=name)
        if not name:
            nc.fill = yellow_fill
            missing_stats['name'] += 1

        ws.cell(row=row_num, column=6, value=card.get('relation_name', ''))
        ws.cell(row=row_num, column=7, value=card.get('house_no', ''))

        age = card.get('age', '')
        ac = ws.cell(row=row_num, column=8, value=age)
        if not age:
            ac.fill = yellow_fill
            missing_stats['age'] += 1

        gender = card.get('gender', '')
        gc = ws.cell(row=row_num, column=9, value=gender)
        if not gender:
            gc.fill = yellow_fill
            missing_stats['gender'] += 1

        ws.cell(row=row_num, column=10, value=card.get('page', ''))

    # Column widths
    widths = [8, 10, 10, 15, 25, 25, 12, 8, 10, 8]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width

    wb.save(output_path)
    return missing_stats


def process_json_file(json_path):
    """Process a single JSON file."""
    print(f"\nProcessing: {json_path.name}")

    # Load JSON
    print("Loading JSON...")
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    entities = data.get('entities', [])
    print(f"Found {len(entities)} entities")

    if not entities:
        print("No entities found!")
        return None, None

    # Count types
    type_counts = {}
    for e in entities:
        t = e.get('type', 'unknown')
        type_counts[t] = type_counts.get(t, 0) + 1

    print("\nEntity types:")
    for t, count in sorted(type_counts.items()):
        print(f"  {t}: {count}")

    # Group into cards
    print("\nGrouping entities into voter cards...")
    cards = group_entities_to_cards(entities)
    print(f"Created {len(cards)} voter cards")

    # Save to Excel
    output_path = json_path.parent / f"{json_path.stem}_excel.xlsx"
    missing_stats = save_to_excel(cards, output_path, json_path.stem)

    return cards, missing_stats, output_path


def main():
    # Select JSON file
    root = tk.Tk()
    root.withdraw()

    json_file = filedialog.askopenfilename(
        title="Select Document AI JSON Output",
        filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
    )

    if not json_file:
        print("No file selected")
        return

    json_path = Path(json_file)
    result = process_json_file(json_path)

    if result[0] is None:
        messagebox.showerror("Error", "No entities found in the JSON file")
        return

    cards, missing_stats, output_path = result

    # Summary
    print(f"\n{'='*50}")
    print(f"SUMMARY")
    print(f"{'='*50}")
    print(f"Total cards: {len(cards)}")
    print(f"Missing Voter ID: {missing_stats['voter_id']}")
    print(f"Missing Name: {missing_stats['name']}")
    print(f"Missing Age: {missing_stats['age']}")
    print(f"Missing Gender: {missing_stats['gender']}")
    print(f"\nExcel saved to: {output_path}")

    # Calculate accuracy
    total = len(cards)
    if total > 0:
        name_acc = ((total - missing_stats['name']) / total) * 100
        age_acc = ((total - missing_stats['age']) / total) * 100
        gender_acc = ((total - missing_stats['gender']) / total) * 100
        voter_acc = ((total - missing_stats['voter_id']) / total) * 100

        messagebox.showinfo("Complete",
            f"Processing complete!\n\n"
            f"Cards: {len(cards)}\n\n"
            f"Accuracy:\n"
            f"  Voter ID: {voter_acc:.1f}%\n"
            f"  Name: {name_acc:.1f}%\n"
            f"  Age: {age_acc:.1f}%\n"
            f"  Gender: {gender_acc:.1f}%\n\n"
            f"Missing:\n"
            f"  Voter ID: {missing_stats['voter_id']}\n"
            f"  Name: {missing_stats['name']}\n"
            f"  Age: {missing_stats['age']}\n"
            f"  Gender: {missing_stats['gender']}\n\n"
            f"Saved to: {output_path.name}")


if __name__ == "__main__":
    main()
