"""
Electoral Roll Voter Counter - Direct PDF Processing via Document AI

Loads a PDF file directly and sends it to Document AI for processing.
Uses custom processor to extract voter data.

Setup:
1. Enable Document AI API in Google Cloud Console
2. Create a Custom Document Extractor processor and train it
3. Add credentials to .env file:
   - GOOGLE_CLIENT_ID
   - GOOGLE_CLIENT_SECRET
   - GOOGLE_PROJECT_ID
   - DOCAI_PROCESSOR_ID
   - DOCAI_LOCATION (us or eu)
"""

import asyncio
import aiohttp
import base64
import json
import re
import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from pathlib import Path
import threading
import subprocess
import time
from concurrent.futures import ThreadPoolExecutor
import webbrowser
from http.server import HTTPServer, BaseHTTPRequestHandler
from urllib.parse import urlencode, parse_qs, urlparse
import secrets

# Install packages
def install_packages():
    packages = ['pillow', 'openpyxl', 'aiohttp', 'python-dotenv']
    for pkg in packages:
        try:
            if pkg == 'python-dotenv':
                __import__('dotenv')
            else:
                __import__(pkg.replace('-', '_'))
        except ImportError:
            print(f"Installing {pkg}...")
            subprocess.check_call(['uv', 'pip', 'install', pkg])

install_packages()

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from dotenv import load_dotenv


def extract_part_number(filename):
    """Extract part number from filename."""
    if not filename:
        return ''
    match = re.search(r'-TAM-(\d+)-WI', filename, re.IGNORECASE)
    if match:
        return match.group(1)
    match = re.search(r'-(\d+)-WI', filename, re.IGNORECASE)
    if match:
        return match.group(1)
    return ''


class OAuthManager:
    """Handles Google OAuth 2.0 authentication flow."""

    SCOPES = ['https://www.googleapis.com/auth/cloud-platform']
    AUTH_URL = 'https://accounts.google.com/o/oauth2/v2/auth'
    TOKEN_URL = 'https://oauth2.googleapis.com/token'
    REDIRECT_PORT = 8089
    REDIRECT_URI = f'http://localhost:{REDIRECT_PORT}/callback'

    def __init__(self, client_id, client_secret):
        self.client_id = client_id
        self.client_secret = client_secret
        self.access_token = None
        self.refresh_token = None
        self.token_expiry = 0
        self.token_file = Path.home() / '.google_docai_oauth_token.json'
        self.load_saved_token()

    def load_saved_token(self):
        if self.token_file.exists():
            try:
                data = json.loads(self.token_file.read_text())
                self.refresh_token = data.get('refresh_token')
                self.access_token = data.get('access_token')
                self.token_expiry = data.get('expiry', 0)
                print("Loaded saved OAuth token")
            except Exception as e:
                print(f"Error loading token: {e}")

    def save_token(self):
        try:
            data = {
                'refresh_token': self.refresh_token,
                'access_token': self.access_token,
                'expiry': self.token_expiry
            }
            self.token_file.write_text(json.dumps(data))
            print("Saved OAuth token")
        except Exception as e:
            print(f"Error saving token: {e}")

    def get_valid_token(self):
        if self.access_token and time.time() < self.token_expiry - 60:
            return self.access_token
        if self.refresh_token:
            if self.refresh_access_token():
                return self.access_token
        if self.do_oauth_flow():
            return self.access_token
        return None

    def refresh_access_token(self):
        try:
            import urllib.request
            data = urlencode({
                'client_id': self.client_id,
                'client_secret': self.client_secret,
                'refresh_token': self.refresh_token,
                'grant_type': 'refresh_token'
            }).encode()
            req = urllib.request.Request(self.TOKEN_URL, data=data)
            with urllib.request.urlopen(req, timeout=30) as response:
                result = json.loads(response.read().decode())
                self.access_token = result['access_token']
                self.token_expiry = time.time() + result.get('expires_in', 3600)
                self.save_token()
                print("Access token refreshed")
                return True
        except Exception as e:
            print(f"Error refreshing token: {e}")
            return False

    def do_oauth_flow(self):
        auth_code = None
        state = secrets.token_urlsafe(16)

        class CallbackHandler(BaseHTTPRequestHandler):
            def do_GET(self):
                nonlocal auth_code
                parsed = urlparse(self.path)
                if parsed.path == '/callback':
                    params = parse_qs(parsed.query)
                    if params.get('state', [None])[0] == state:
                        auth_code = params.get('code', [None])[0]
                        self.send_response(200)
                        self.send_header('Content-type', 'text/html')
                        self.end_headers()
                        self.wfile.write(b'''
                            <html><body style="font-family: Arial; text-align: center; padding: 50px;">
                            <h1 style="color: #4CAF50;">Authorization Successful!</h1>
                            <p>You can close this window and return to the application.</p>
                            </body></html>
                        ''')
                    else:
                        self.send_response(400)
                        self.end_headers()
                else:
                    self.send_response(404)
                    self.end_headers()

            def log_message(self, format, *args):
                pass

        auth_params = {
            'client_id': self.client_id,
            'redirect_uri': self.REDIRECT_URI,
            'response_type': 'code',
            'scope': ' '.join(self.SCOPES),
            'state': state,
            'access_type': 'offline',
            'prompt': 'consent'
        }
        auth_url = f"{self.AUTH_URL}?{urlencode(auth_params)}"

        server = HTTPServer(('localhost', self.REDIRECT_PORT), CallbackHandler)
        server.timeout = 120

        print(f"Opening browser for authorization...")
        webbrowser.open(auth_url)

        while auth_code is None:
            server.handle_request()

        server.server_close()

        if not auth_code:
            print("Authorization failed - no code received")
            return False

        try:
            import urllib.request
            data = urlencode({
                'client_id': self.client_id,
                'client_secret': self.client_secret,
                'code': auth_code,
                'redirect_uri': self.REDIRECT_URI,
                'grant_type': 'authorization_code'
            }).encode()
            req = urllib.request.Request(self.TOKEN_URL, data=data)
            with urllib.request.urlopen(req, timeout=30) as response:
                result = json.loads(response.read().decode())
                self.access_token = result['access_token']
                self.refresh_token = result.get('refresh_token', self.refresh_token)
                self.token_expiry = time.time() + result.get('expires_in', 3600)
                self.save_token()
                print("Authorization successful!")
                return True
        except Exception as e:
            print(f"Error exchanging code for token: {e}")
            return False


def process_pdf_with_docai(pdf_content_b64, oauth_manager, project_id, location, processor_id, log_func=None):
    """Send PDF to Document AI and get results."""

    token = oauth_manager.get_valid_token()
    if not token:
        if log_func:
            log_func("Failed to get valid OAuth token")
        return None

    url = f"https://{location}-documentai.googleapis.com/v1/projects/{project_id}/locations/{location}/processors/{processor_id}:process"

    headers = {
        'Authorization': f'Bearer {token}',
        'Content-Type': 'application/json'
    }

    payload = {
        "rawDocument": {
            "content": pdf_content_b64,
            "mimeType": "application/pdf"
        }
    }

    try:
        import urllib.request
        req = urllib.request.Request(
            url,
            data=json.dumps(payload).encode('utf-8'),
            headers=headers,
            method='POST'
        )

        if log_func:
            log_func("Sending PDF to Document AI...")

        with urllib.request.urlopen(req, timeout=300) as response:
            result = json.loads(response.read().decode())
            return result.get('document', {})

    except urllib.error.HTTPError as e:
        error_body = e.read().decode() if e.fp else ''
        if log_func:
            log_func(f"API Error {e.code}: {error_body[:500]}")
        return None
    except Exception as e:
        if log_func:
            log_func(f"Request error: {e}")
        return None


def parse_document_response(document, pdf_name):
    """Parse Document AI response to extract voter cards.

    Handles both:
    1. Custom processor with entities (trained fields)
    2. Fallback text parsing
    """
    cards = []

    if not document:
        return cards

    part_no = extract_part_number(pdf_name)

    # Get full text
    full_text = document.get('text', '')

    # Try entities first (custom processor)
    entities = document.get('entities', [])

    if entities:
        cards = parse_custom_entities(entities, document, pdf_name, part_no)

    # If no entities or no cards, try page-by-page parsing
    if not cards:
        pages = document.get('pages', [])
        for page_num, page in enumerate(pages):
            page_cards = parse_page(page, full_text, pdf_name, part_no, page_num + 1)
            cards.extend(page_cards)

    # Last fallback: parse full text
    if not cards:
        cards = parse_full_text(full_text, pdf_name, part_no)

    return cards


def parse_custom_entities(entities, document, pdf_name, part_no):
    """Parse entities from a custom trained processor."""
    cards = []
    full_text = document.get('text', '')

    # Custom processors typically return entities with types matching your training labels
    # Group entities by their page location to identify individual cards

    # First, check if entities are grouped by card (nested structure)
    for entity in entities:
        entity_type = entity.get('type', '').lower()

        # If the entity type suggests a voter card (adjust based on your training)
        if 'voter' in entity_type or 'card' in entity_type or entity_type == 'voter_card':
            # This is a card-level entity with nested properties
            card = extract_card_from_entity(entity, full_text, pdf_name, part_no)
            if card.get('name') or card.get('voter_id'):
                cards.append(card)
        else:
            # Flat entity structure - need to group by position
            pass

    # If no card-level entities, try to group flat entities by Y position
    if not cards and entities:
        cards = group_entities_to_cards(entities, document, pdf_name, part_no)

    return cards


def extract_card_from_entity(entity, full_text, pdf_name, part_no):
    """Extract card data from a card-level entity."""
    card = {
        'serial_no': '',
        'voter_id': '',
        'name': '',
        'relation_type': '',
        'relation_name': '',
        'house_no': '',
        'age': '',
        'gender': '',
        'folder_name': pdf_name,
        'part_no': part_no
    }

    # Get the text for this entity
    mention_text = entity.get('mentionText', '')

    # Check for nested properties
    properties = entity.get('properties', [])

    if properties:
        for prop in properties:
            prop_type = prop.get('type', '').lower()
            prop_text = prop.get('mentionText', '').strip()

            if 'name' in prop_type and 'relation' not in prop_type and 'father' not in prop_type and 'husband' not in prop_type and 'mother' not in prop_type:
                card['name'] = clean_name(prop_text)
            elif 'voter_id' in prop_type or 'epic' in prop_type or prop_type == 'id':
                card['voter_id'] = prop_text.upper()
            elif 'serial' in prop_type:
                card['serial_no'] = prop_text
            elif 'father' in prop_type:
                card['relation_type'] = 'Father'
                card['relation_name'] = clean_name(prop_text)
            elif 'husband' in prop_type:
                card['relation_type'] = 'Husband'
                card['relation_name'] = clean_name(prop_text)
            elif 'mother' in prop_type:
                card['relation_type'] = 'Mother'
                card['relation_name'] = clean_name(prop_text)
            elif 'relation' in prop_type:
                card['relation_name'] = clean_name(prop_text)
            elif 'house' in prop_type or 'address' in prop_type:
                card['house_no'] = extract_house_no(prop_text)
            elif 'age' in prop_type:
                card['age'] = extract_age(prop_text)
            elif 'gender' in prop_type or 'sex' in prop_type:
                card['gender'] = extract_gender(prop_text)
    else:
        # No properties, parse from mention text
        card = parse_card_text(mention_text, pdf_name, part_no)

    return card


def group_entities_to_cards(entities, document, pdf_name, part_no):
    """Group flat entities into cards based on page position."""
    cards = []
    full_text = document.get('text', '')

    # Sort entities by page and Y position
    sorted_entities = sorted(entities, key=lambda e: (
        get_entity_page(e),
        get_entity_y(e)
    ))

    # Group entities that are close together (same card)
    current_card = new_card(pdf_name, part_no)
    current_y = -1
    current_page = -1
    y_threshold = 0.08  # ~8% of page height

    for entity in sorted_entities:
        entity_y = get_entity_y(entity)
        entity_page = get_entity_page(entity)

        # New page or significant Y jump = new card
        if entity_page != current_page or (current_y >= 0 and abs(entity_y - current_y) > y_threshold):
            if current_card.get('name') or current_card.get('voter_id'):
                cards.append(current_card)
            current_card = new_card(pdf_name, part_no)
            current_y = entity_y
            current_page = entity_page

        # Add entity to current card
        entity_type = entity.get('type', '').lower()
        mention_text = entity.get('mentionText', '').strip()

        apply_entity_to_card(current_card, entity_type, mention_text)

        if current_y < 0:
            current_y = entity_y

    # Don't forget the last card
    if current_card.get('name') or current_card.get('voter_id'):
        cards.append(current_card)

    return cards


def get_entity_page(entity):
    """Get page number from entity."""
    page_anchor = entity.get('pageAnchor', {})
    page_refs = page_anchor.get('pageRefs', [{}])
    return int(page_refs[0].get('page', 0)) if page_refs else 0


def get_entity_y(entity):
    """Get Y position from entity bounding box."""
    page_anchor = entity.get('pageAnchor', {})
    page_refs = page_anchor.get('pageRefs', [{}])
    if page_refs:
        bounding_poly = page_refs[0].get('boundingPoly', {})
        vertices = bounding_poly.get('normalizedVertices', [{}])
        if vertices:
            return vertices[0].get('y', 0)
    return 0


def new_card(pdf_name, part_no):
    """Create a new empty card."""
    return {
        'serial_no': '',
        'voter_id': '',
        'name': '',
        'relation_type': '',
        'relation_name': '',
        'house_no': '',
        'age': '',
        'gender': '',
        'folder_name': pdf_name,
        'part_no': part_no
    }


def apply_entity_to_card(card, entity_type, text):
    """Apply an entity to the appropriate card field."""
    if not text:
        return

    if 'name' in entity_type and 'relation' not in entity_type and 'father' not in entity_type:
        if not card['name']:
            card['name'] = clean_name(text)
    elif 'voter' in entity_type or 'epic' in entity_type or entity_type == 'id':
        card['voter_id'] = text.upper()
    elif 'serial' in entity_type or entity_type == 'number':
        if not card['serial_no']:
            card['serial_no'] = text
    elif 'father' in entity_type:
        card['relation_type'] = 'Father'
        card['relation_name'] = clean_name(text)
    elif 'husband' in entity_type:
        card['relation_type'] = 'Husband'
        card['relation_name'] = clean_name(text)
    elif 'mother' in entity_type:
        card['relation_type'] = 'Mother'
        card['relation_name'] = clean_name(text)
    elif 'house' in entity_type or 'address' in entity_type:
        card['house_no'] = extract_house_no(text)
    elif 'age' in entity_type:
        card['age'] = extract_age(text)
    elif 'gender' in entity_type or 'sex' in entity_type:
        card['gender'] = extract_gender(text)


def parse_page(page, full_text, pdf_name, part_no, page_num):
    """Parse a single page to extract voter cards."""
    cards = []

    # Try tables first
    tables = page.get('tables', [])
    if tables:
        for table in tables:
            table_cards = parse_table(table, full_text, pdf_name, part_no)
            cards.extend(table_cards)

    # Try blocks
    if not cards:
        blocks = page.get('blocks', [])
        if blocks:
            cards = parse_blocks(blocks, full_text, pdf_name, part_no)

    return cards


def parse_table(table, full_text, pdf_name, part_no):
    """Parse a detected table."""
    cards = []
    body_rows = table.get('bodyRows', [])

    for row in body_rows:
        cells = row.get('cells', [])
        for cell in cells:
            cell_text = get_text_from_layout(cell.get('layout', {}), full_text)
            if cell_text and len(cell_text) > 20:  # Filter out small text
                card = parse_card_text(cell_text, pdf_name, part_no)
                if card.get('name') or card.get('voter_id'):
                    cards.append(card)

    return cards


def parse_blocks(blocks, full_text, pdf_name, part_no):
    """Parse text blocks to extract cards."""
    cards = []

    # Group blocks by Y position
    sorted_blocks = sorted(blocks, key=lambda b: (
        b.get('layout', {}).get('boundingPoly', {}).get('normalizedVertices', [{}])[0].get('y', 0),
        b.get('layout', {}).get('boundingPoly', {}).get('normalizedVertices', [{}])[0].get('x', 0)
    ))

    current_text = []
    current_y = -1
    y_threshold = 0.08

    for block in sorted_blocks:
        layout = block.get('layout', {})
        vertices = layout.get('boundingPoly', {}).get('normalizedVertices', [])
        if not vertices:
            continue

        block_y = vertices[0].get('y', 0)
        block_text = get_text_from_layout(layout, full_text)

        if current_y < 0:
            current_y = block_y

        if abs(block_y - current_y) > y_threshold:
            # Process accumulated text
            if current_text:
                combined = ' '.join(current_text)
                # Split by voter ID pattern
                parts = re.split(r'(?=[A-Z]{2,3}\d{6,10})', combined)
                for part in parts:
                    if part.strip():
                        card = parse_card_text(part, pdf_name, part_no)
                        if card.get('name') or card.get('voter_id'):
                            cards.append(card)
            current_text = []
            current_y = block_y

        current_text.append(block_text)

    # Process remaining
    if current_text:
        combined = ' '.join(current_text)
        parts = re.split(r'(?=[A-Z]{2,3}\d{6,10})', combined)
        for part in parts:
            if part.strip():
                card = parse_card_text(part, pdf_name, part_no)
                if card.get('name') or card.get('voter_id'):
                    cards.append(card)

    return cards


def get_text_from_layout(layout, full_text):
    """Extract text using text anchors."""
    text_anchor = layout.get('textAnchor', {})
    text_segments = text_anchor.get('textSegments', [])

    result = []
    for segment in text_segments:
        start_idx = int(segment.get('startIndex', 0))
        end_idx = int(segment.get('endIndex', 0))
        if end_idx > start_idx:
            result.append(full_text[start_idx:end_idx])

    return ' '.join(result)


def parse_full_text(full_text, pdf_name, part_no):
    """Fallback: parse full text to extract voter cards."""
    cards = []

    if not full_text:
        return cards

    # Split by voter ID pattern
    voter_id_pattern = r'([A-Z]{2,3}\d{6,10})'
    parts = re.split(voter_id_pattern, full_text)

    i = 0
    while i < len(parts):
        card_text = parts[i]
        voter_id = ''

        if i + 1 < len(parts) and re.match(voter_id_pattern, parts[i + 1]):
            voter_id = parts[i + 1]
            i += 2
        else:
            i += 1

        if card_text.strip() or voter_id:
            card = parse_card_text(card_text + ' ' + voter_id, pdf_name, part_no)
            card['voter_id'] = voter_id
            if card.get('name') or voter_id:
                cards.append(card)

    return cards


def parse_card_text(text, pdf_name, part_no):
    """Parse text from a single voter card using Tamil patterns."""
    card = {
        'serial_no': '',
        'voter_id': '',
        'name': '',
        'relation_type': '',
        'relation_name': '',
        'house_no': '',
        'age': '',
        'gender': '',
        'folder_name': pdf_name,
        'part_no': part_no
    }

    if not text:
        return card

    # Clean text
    text = re.sub(r'Photo\s*is\s*available', '', text, flags=re.IGNORECASE)
    text = re.sub(r'\s+', ' ', text).strip()

    # Voter ID
    voter_match = re.search(r'\b([A-Z]{2,3}\d{6,10})\b', text)
    if voter_match:
        card['voter_id'] = voter_match.group(1)

    # Serial Number
    serial_match = re.search(r'^\s*(\d{1,4})\s', text)
    if serial_match:
        card['serial_no'] = serial_match.group(1)

    # Name (பெயர்)
    name_match = re.search(r'பெயர்\s*[:\-–]?\s*([^\-\n]+?)(?=\s*[-–]|\s*கணவர்|\s*தந்தை|\s*தாய்|\s*வீட்|$)', text)
    if name_match:
        card['name'] = clean_name(name_match.group(1))

    # Husband (கணவர் பெயர்)
    husband_match = re.search(r'கணவர்\s*பெயர்\s*[:\-–]?\s*([^\-\n]+?)(?=\s*[-–]|\s*வீட்|$)', text)
    if husband_match:
        card['relation_type'] = 'Husband'
        card['relation_name'] = clean_name(husband_match.group(1))

    # Father (தந்தையின் பெயர்)
    if not card['relation_name']:
        father_match = re.search(r'தந்தையின்\s*பெயர்\s*[:\-–]?\s*([^\-\n]+?)(?=\s*[-–]|\s*வீட்|$)', text)
        if father_match:
            card['relation_type'] = 'Father'
            card['relation_name'] = clean_name(father_match.group(1))

    # Mother (தாயின் பெயர்)
    if not card['relation_name']:
        mother_match = re.search(r'தாயின்\s*பெயர்\s*[:\-–]?\s*([^\-\n]+?)(?=\s*[-–]|\s*வீட்|$)', text)
        if mother_match:
            card['relation_type'] = 'Mother'
            card['relation_name'] = clean_name(mother_match.group(1))

    # House Number (வீட்டு எண்)
    house_match = re.search(r'வீட்டு\s*எண்\s*[:\-–]?\s*(\d+[-/]?\d*[A-Za-z]?)', text)
    if house_match:
        card['house_no'] = house_match.group(1).strip()

    # Age (வயது)
    age_match = re.search(r'வயது\s*[:\-–]?\s*(\d{1,3})', text)
    if age_match:
        age = int(age_match.group(1))
        if 18 <= age <= 120:
            card['age'] = str(age)

    # Gender (பாலினம்)
    card['gender'] = extract_gender(text)

    return card


def clean_name(name):
    """Clean extracted name."""
    if not name:
        return ''
    name = re.sub(r'\s*பெயர்.*$', '', name)
    name = re.sub(r'\s*[:\-–]\s*$', '', name)
    name = re.sub(r'\s+', ' ', name)
    return name.strip()


def extract_house_no(text):
    """Extract house number from text."""
    if not text:
        return ''
    match = re.search(r'(\d+[-/]?\d*[A-Za-z]?)', text)
    return match.group(1) if match else text.strip()


def extract_age(text):
    """Extract age from text."""
    if not text:
        return ''
    match = re.search(r'(\d{1,3})', text)
    if match:
        age = int(match.group(1))
        if 18 <= age <= 120:
            return str(age)
    return ''


def extract_gender(text):
    """Extract gender from text."""
    if not text:
        return ''
    if 'பெண்' in text or 'பெண' in text or 'female' in text.lower():
        return 'Female'
    elif 'ஆண்' in text or 'ஆண' in text or 'male' in text.lower():
        return 'Male'
    return ''


class PDFVoterCounter:
    def __init__(self, root):
        self.root = root
        self.root.title("Electoral Roll - PDF Processor (Document AI)")
        self.root.geometry("800x700")
        self.root.resizable(True, True)

        self.oauth_manager = None

        # Load environment variables
        env_file = Path(__file__).parent / '.env'
        if env_file.exists():
            load_dotenv(env_file)
            self.client_id = os.getenv('GOOGLE_CLIENT_ID', '')
            self.client_secret = os.getenv('GOOGLE_CLIENT_SECRET', '')
            self.project_id = os.getenv('GOOGLE_PROJECT_ID', '')
            self.processor_id = os.getenv('DOCAI_PROCESSOR_ID', '')
            self.location = os.getenv('DOCAI_LOCATION', 'us')
        else:
            self.client_id = ''
            self.client_secret = ''
            self.project_id = ''
            self.processor_id = ''
            self.location = 'us'

        self.style = ttk.Style()
        self.style.configure('Title.TLabel', font=('Helvetica', 16, 'bold'))
        self.style.configure('Header.TLabel', font=('Helvetica', 12, 'bold'))
        self.style.configure('Big.TLabel', font=('Helvetica', 24, 'bold'))

        self.create_widgets()
        self.check_credentials()

    def create_widgets(self):
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Title
        title_label = ttk.Label(
            main_frame,
            text="Electoral Roll PDF Processor\nGoogle Document AI (Custom)",
            style='Title.TLabel',
            justify=tk.CENTER
        )
        title_label.pack(pady=(0, 10))

        # Cost info
        cost_info = ttk.Label(
            main_frame,
            text="Cost: $0.10/page (online) | $0.01/page (batch) | First 1000 pages/month FREE",
            foreground='#E91E63',
            font=('Helvetica', 9, 'bold')
        )
        cost_info.pack(pady=(0, 10))

        # Credentials frame
        cred_frame = ttk.LabelFrame(main_frame, text="Google Cloud Credentials", padding="10")
        cred_frame.pack(fill=tk.X, pady=(0, 10))

        self.cred_status_var = tk.StringVar(value="Checking...")
        ttk.Label(cred_frame, textvariable=self.cred_status_var).pack(side=tk.LEFT)

        ttk.Button(cred_frame, text="Authorize", command=self.authorize).pack(side=tk.RIGHT, padx=5)
        ttk.Button(cred_frame, text="Edit .env", command=self.open_env_file).pack(side=tk.RIGHT, padx=5)

        # Document AI Settings
        docai_frame = ttk.LabelFrame(main_frame, text="Document AI Settings", padding="10")
        docai_frame.pack(fill=tk.X, pady=(0, 10))

        # Project ID
        proj_row = ttk.Frame(docai_frame)
        proj_row.pack(fill=tk.X, pady=2)
        ttk.Label(proj_row, text="Project ID:", width=15).pack(side=tk.LEFT)
        self.project_id_var = tk.StringVar(value=self.project_id)
        ttk.Entry(proj_row, textvariable=self.project_id_var, width=40).pack(side=tk.LEFT, fill=tk.X, expand=True)

        # Processor ID
        proc_row = ttk.Frame(docai_frame)
        proc_row.pack(fill=tk.X, pady=2)
        ttk.Label(proc_row, text="Processor ID:", width=15).pack(side=tk.LEFT)
        self.processor_id_var = tk.StringVar(value=self.processor_id)
        ttk.Entry(proc_row, textvariable=self.processor_id_var, width=40).pack(side=tk.LEFT, fill=tk.X, expand=True)

        # Location
        loc_row = ttk.Frame(docai_frame)
        loc_row.pack(fill=tk.X, pady=2)
        ttk.Label(loc_row, text="Location:", width=15).pack(side=tk.LEFT)
        self.location_var = tk.StringVar(value=self.location)
        ttk.Combobox(loc_row, textvariable=self.location_var, values=['us', 'eu'], width=10).pack(side=tk.LEFT)

        # PDF selection
        pdf_frame = ttk.LabelFrame(main_frame, text="Select PDF File", padding="10")
        pdf_frame.pack(fill=tk.X, pady=(0, 10))

        self.pdf_path_var = tk.StringVar()
        ttk.Entry(pdf_frame, textvariable=self.pdf_path_var, width=60).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        ttk.Button(pdf_frame, text="Browse...", command=self.browse_pdf).pack(side=tk.LEFT)

        # Info
        info_frame = ttk.Frame(main_frame)
        info_frame.pack(fill=tk.X, pady=(0, 10))

        self.file_info_var = tk.StringVar(value="No file selected")
        ttk.Label(info_frame, textvariable=self.file_info_var).pack(side=tk.LEFT)

        self.cost_estimate_var = tk.StringVar(value="Est. Cost: --")
        ttk.Label(info_frame, textvariable=self.cost_estimate_var, foreground='#E91E63',
                  font=('Helvetica', 10, 'bold')).pack(side=tk.RIGHT)

        # Process button
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(pady=(0, 10))

        self.process_btn = ttk.Button(btn_frame, text="Process PDF", command=self.process_pdf)
        self.process_btn.pack()

        # Progress
        progress_frame = ttk.LabelFrame(main_frame, text="Progress", padding="10")
        progress_frame.pack(fill=tk.X, pady=(0, 10))

        self.status_var = tk.StringVar(value="Ready")
        ttk.Label(progress_frame, textvariable=self.status_var).pack(anchor=tk.W)

        self.progress = ttk.Progressbar(progress_frame, mode='indeterminate')
        self.progress.pack(fill=tk.X, pady=5)

        # Results
        results_frame = ttk.LabelFrame(main_frame, text="Results", padding="10")
        results_frame.pack(fill=tk.X, pady=(0, 10))

        results_grid = ttk.Frame(results_frame)
        results_grid.pack()

        ttk.Label(results_grid, text="Cards Extracted:", style='Header.TLabel').grid(row=0, column=0, padx=10)
        self.cards_var = tk.StringVar(value="--")
        ttk.Label(results_grid, textvariable=self.cards_var, style='Big.TLabel',
                  foreground='#E91E63').grid(row=0, column=1, padx=10)

        ttk.Label(results_grid, text="Time:", style='Header.TLabel').grid(row=0, column=2, padx=10)
        self.time_var = tk.StringVar(value="--")
        ttk.Label(results_grid, textvariable=self.time_var, style='Big.TLabel',
                  foreground='#4CAF50').grid(row=0, column=3, padx=10)

        # Log
        log_frame = ttk.LabelFrame(main_frame, text="Log", padding="5")
        log_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))

        self.log_text = tk.Text(log_frame, height=10, width=80, font=('Courier', 9))
        log_scroll = ttk.Scrollbar(log_frame, orient=tk.VERTICAL, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=log_scroll.set)
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        log_scroll.pack(side=tk.RIGHT, fill=tk.Y)

    def log(self, message):
        self.log_text.insert(tk.END, f"{time.strftime('%H:%M:%S')} - {message}\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()

    def check_credentials(self):
        if self.client_id and self.client_secret:
            self.cred_status_var.set(f"Client ID: {self.client_id[:20]}...")
            self.oauth_manager = OAuthManager(self.client_id, self.client_secret)
            if self.oauth_manager.refresh_token:
                self.cred_status_var.set(f"Client ID: {self.client_id[:20]}... (authorized)")
        else:
            self.cred_status_var.set("Not configured - Edit .env file")

    def open_env_file(self):
        env_file = Path(__file__).parent / '.env'
        if not env_file.exists():
            env_file.write_text(
                "# Google Cloud Credentials\n"
                "GOOGLE_CLIENT_ID=your_client_id.apps.googleusercontent.com\n"
                "GOOGLE_CLIENT_SECRET=your_client_secret\n"
                "\n"
                "# Document AI Settings\n"
                "GOOGLE_PROJECT_ID=your_project_id\n"
                "DOCAI_PROCESSOR_ID=your_processor_id\n"
                "DOCAI_LOCATION=us\n"
            )
        os.startfile(str(env_file)) if os.name == 'nt' else subprocess.run(['xdg-open', str(env_file)])

    def authorize(self):
        env_file = Path(__file__).parent / '.env'
        if env_file.exists():
            load_dotenv(env_file, override=True)
            self.client_id = os.getenv('GOOGLE_CLIENT_ID', '')
            self.client_secret = os.getenv('GOOGLE_CLIENT_SECRET', '')

        if not self.client_id or not self.client_secret:
            messagebox.showerror("Error", "Please configure credentials in .env file")
            return

        self.oauth_manager = OAuthManager(self.client_id, self.client_secret)
        self.log("Starting OAuth authorization...")

        def do_auth():
            token = self.oauth_manager.get_valid_token()
            if token:
                self.root.after(0, lambda: self.cred_status_var.set(f"Client ID: {self.client_id[:20]}... (authorized)"))
                self.root.after(0, lambda: self.log("Authorization successful!"))
            else:
                self.root.after(0, lambda: self.log("Authorization failed"))

        thread = threading.Thread(target=do_auth)
        thread.daemon = True
        thread.start()

    def browse_pdf(self):
        pdf_file = filedialog.askopenfilename(
            title="Select PDF File",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
        )
        if pdf_file:
            self.pdf_path_var.set(pdf_file)
            self.estimate_cost(pdf_file)

    def estimate_cost(self, pdf_path):
        """Estimate cost based on PDF file."""
        try:
            # Try to get page count using PyMuPDF or similar
            # For now, estimate based on file size
            file_size = Path(pdf_path).stat().st_size
            estimated_pages = max(1, file_size // 100000)  # ~100KB per page rough estimate

            # Try to get actual page count
            try:
                import fitz  # PyMuPDF
                doc = fitz.open(pdf_path)
                estimated_pages = len(doc)
                doc.close()
            except:
                pass

            cost = estimated_pages * 0.10  # Online processing cost

            self.file_info_var.set(f"~{estimated_pages} pages")
            self.cost_estimate_var.set(f"Est. Cost: ${cost:.2f}")
            self.log(f"Selected: {Path(pdf_path).name}")
            self.log(f"Estimated pages: ~{estimated_pages}, Cost: ~${cost:.2f}")
        except Exception as e:
            self.file_info_var.set("Error reading file")
            self.log(f"Error: {e}")

    def process_pdf(self):
        pdf_path = self.pdf_path_var.get()
        if not pdf_path or not Path(pdf_path).exists():
            messagebox.showerror("Error", "Please select a valid PDF file")
            return

        if not self.oauth_manager:
            messagebox.showerror("Error", "Please authorize first")
            return

        project_id = self.project_id_var.get()
        processor_id = self.processor_id_var.get()
        location = self.location_var.get()

        if not project_id or not processor_id:
            messagebox.showerror("Error", "Please enter Project ID and Processor ID")
            return

        token = self.oauth_manager.get_valid_token()
        if not token:
            messagebox.showerror("Error", "Failed to get valid OAuth token")
            return

        self.process_btn.config(state=tk.DISABLED)
        self.progress.start()

        thread = threading.Thread(target=self.do_process, args=(pdf_path, project_id, processor_id, location))
        thread.daemon = True
        thread.start()

    def do_process(self, pdf_path, project_id, processor_id, location):
        """Process the PDF file."""
        start_time = time.time()
        pdf_name = Path(pdf_path).stem

        self.root.after(0, lambda: self.status_var.set("Loading PDF..."))
        self.root.after(0, lambda: self.log(f"Loading PDF: {pdf_name}"))

        try:
            # Read PDF content
            with open(pdf_path, 'rb') as f:
                pdf_content = f.read()

            pdf_b64 = base64.b64encode(pdf_content).decode('utf-8')

            # Check file size (Document AI has 20MB limit for online processing)
            file_size_mb = len(pdf_content) / (1024 * 1024)
            self.root.after(0, lambda s=file_size_mb: self.log(f"PDF size: {s:.2f} MB"))

            if file_size_mb > 20:
                self.root.after(0, lambda: self.log("WARNING: File > 20MB. May need batch processing."))

            self.root.after(0, lambda: self.status_var.set("Processing with Document AI..."))

            # Process with Document AI
            document = process_pdf_with_docai(
                pdf_b64,
                self.oauth_manager,
                project_id,
                location,
                processor_id,
                lambda msg: self.root.after(0, lambda m=msg: self.log(m))
            )

            if not document:
                self.root.after(0, lambda: self.status_var.set("Error: No response from API"))
                self.root.after(0, lambda: self.progress.stop())
                self.root.after(0, lambda: self.process_btn.config(state=tk.NORMAL))
                return

            self.root.after(0, lambda: self.status_var.set("Parsing results..."))
            self.root.after(0, lambda: self.log("Parsing Document AI response..."))

            # Parse the response
            cards = parse_document_response(document, pdf_name)

            self.root.after(0, lambda c=len(cards): self.log(f"Extracted {c} voter cards"))

            # Save to Excel
            self.root.after(0, lambda: self.status_var.set("Saving to Excel..."))

            wb = Workbook()
            ws = wb.active
            ws.title = "Voter Data"

            headers = ['S.No', 'Part No.', 'Voter S.No', 'Voter ID', 'Name', 'Relation Type',
                       'Relation Name', 'House No', 'Age', 'Gender', 'Source']

            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            missing_stats = {'name': 0, 'age': 0, 'gender': 0}

            for row_num, card in enumerate(cards, 2):
                ws.cell(row=row_num, column=1, value=row_num - 1)
                ws.cell(row=row_num, column=2, value=card.get('part_no', ''))
                ws.cell(row=row_num, column=3, value=card.get('serial_no', ''))
                ws.cell(row=row_num, column=4, value=card.get('voter_id', ''))

                name_val = card.get('name', '')
                name_cell = ws.cell(row=row_num, column=5, value=name_val)
                if not name_val:
                    name_cell.fill = yellow_fill
                    missing_stats['name'] += 1

                ws.cell(row=row_num, column=6, value=card.get('relation_type', ''))
                ws.cell(row=row_num, column=7, value=card.get('relation_name', ''))
                ws.cell(row=row_num, column=8, value=card.get('house_no', ''))

                age_val = card.get('age', '')
                age_cell = ws.cell(row=row_num, column=9, value=age_val)
                if not age_val:
                    age_cell.fill = yellow_fill
                    missing_stats['age'] += 1

                gender_val = card.get('gender', '')
                gender_cell = ws.cell(row=row_num, column=10, value=gender_val)
                if not gender_val:
                    gender_cell.fill = yellow_fill
                    missing_stats['gender'] += 1

                ws.cell(row=row_num, column=11, value=card.get('folder_name', ''))

            # Column widths
            widths = [8, 12, 10, 15, 25, 12, 25, 15, 8, 10, 40]
            for col, width in enumerate(widths, 1):
                ws.column_dimensions[chr(64 + col)].width = width

            excel_path = Path(pdf_path).parent / f"{pdf_name}_docai_excel.xlsx"
            wb.save(excel_path)

            elapsed = time.time() - start_time
            elapsed_str = f"{int(elapsed//60)}m {int(elapsed%60)}s"

            self.root.after(0, lambda: self.progress.stop())
            self.root.after(0, lambda: self.status_var.set("Complete!"))
            self.root.after(0, lambda c=len(cards): self.cards_var.set(f"{c:,}"))
            self.root.after(0, lambda e=elapsed_str: self.time_var.set(e))
            self.root.after(0, lambda: self.process_btn.config(state=tk.NORMAL))

            self.root.after(0, lambda p=str(excel_path): self.log(f"Saved: {p}"))
            self.root.after(0, lambda: self.log(f"Missing - Name: {missing_stats['name']}, Age: {missing_stats['age']}, Gender: {missing_stats['gender']}"))

            self.root.after(0, lambda: messagebox.showinfo("Complete",
                f"Processing complete!\n\n"
                f"Cards: {len(cards):,}\n"
                f"Time: {elapsed_str}\n\n"
                f"Missing Name: {missing_stats['name']}\n"
                f"Missing Age: {missing_stats['age']}\n"
                f"Missing Gender: {missing_stats['gender']}\n\n"
                f"Excel: {excel_path.name}"))

        except Exception as e:
            self.root.after(0, lambda: self.progress.stop())
            self.root.after(0, lambda err=str(e): self.status_var.set(f"Error: {err}"))
            self.root.after(0, lambda err=str(e): self.log(f"ERROR: {err}"))
            self.root.after(0, lambda: self.process_btn.config(state=tk.NORMAL))
            import traceback
            self.root.after(0, lambda: self.log(traceback.format_exc()))


def main():
    root = tk.Tk()
    app = PDFVoterCounter(root)
    root.mainloop()


if __name__ == "__main__":
    main()
