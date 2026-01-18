"""
Age Extraction Viewer - Multi-Preprocessing Comparison
Test different OCR approaches to find the best one for Age extraction.

Similar to house_number_simple_5spaces.py but for Age field.

Usage:
1. Load Excel file with voter data
2. Load Image folder with card images
3. Navigate through cards and compare Age extraction results
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import re
from pathlib import Path
from PIL import Image, ImageTk, ImageEnhance
import pytesseract


def extract_age_from_text(text):
    """Extract age from OCR text using multiple patterns."""

    # Pattern 1: வயது : <age>
    match = re.search(r'வயது\s*[:\-;.]\s*(\d{1,3})', text)
    if match:
        age = match.group(1)
        if 18 <= int(age) <= 120:
            return (age, f"Pattern 1: வயது : {age}")

    # Pattern 2: Age : <age>
    match = re.search(r'Age\s*[:\-;.]\s*(\d{1,3})', text, re.IGNORECASE)
    if match:
        age = match.group(1)
        if 18 <= int(age) <= 120:
            return (age, f"Pattern 2: Age : {age}")

    # Pattern 3: Look for 2-digit number near பாலினம் (gender)
    if 'பாலினம்' in text:
        # Age usually appears before gender
        lines = text.split('\n')
        for i, line in enumerate(lines):
            if 'பாலினம்' in line:
                # Check this line and previous line for age
                for check_line in [line, lines[i-1] if i > 0 else '']:
                    match = re.search(r'\b(\d{2})\b', check_line)
                    if match:
                        age = match.group(1)
                        if 18 <= int(age) <= 99:
                            return (age, f"Pattern 3: Near பாலினம் - {age}")

    # Pattern 4: Standalone 2-digit number (18-99)
    matches = re.findall(r'\b(\d{2})\b', text)
    for age in matches:
        if 18 <= int(age) <= 99:
            return (age, f"Pattern 4: Standalone - {age}")

    return ('', 'NOT FOUND')


def ocr_with_all_approaches(img):
    """
    OCR the image with ALL preprocessing approaches.
    Returns dict with results from each approach.
    """
    approaches = [
        ('Original', lambda i: i),
        ('Contrast 2x', lambda i: ImageEnhance.Contrast(i).enhance(2.0)),
        ('Contrast 3x', lambda i: ImageEnhance.Contrast(i).enhance(3.0)),
        ('Grayscale', lambda i: i.convert('L')),
        ('Gray+Sharp', lambda i: ImageEnhance.Sharpness(i.convert('L')).enhance(2.0)),
        ('Binarize', lambda i: i.convert('L').point(lambda x: 0 if x < 140 else 255, '1')),
        ('Scale 2x', lambda i: i.resize((i.size[0] * 2, i.size[1] * 2), Image.LANCZOS)),
        ('Scale 3x', lambda i: i.resize((i.size[0] * 3, i.size[1] * 3), Image.LANCZOS)),
    ]

    results = {}

    for name, transform in approaches:
        try:
            processed = transform(img)
            text = pytesseract.image_to_string(processed, lang='tam+eng')
            age, pattern_info = extract_age_from_text(text)
            results[name] = {
                'age': age,
                'pattern': pattern_info,
                'full_text': text[:500]  # First 500 chars for debugging
            }
        except Exception as e:
            results[name] = {
                'age': f'ERROR: {str(e)[:30]}',
                'pattern': '',
                'full_text': ''
            }

    return results


class AgeViewerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Age Viewer - Multi-Preprocessing Comparison")
        self.root.geometry("1400x800")

        self.excel_data = []
        self.image_folder = None
        self.current_index = 0
        self.card_images = []

        self.setup_ui()

    def setup_ui(self):
        # Main frame
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Top: File selection
        file_frame = ttk.LabelFrame(main_frame, text="Load Files", padding="5")
        file_frame.pack(fill=tk.X, pady=(0, 10))

        ttk.Label(file_frame, text="Excel File:").grid(row=0, column=0, sticky=tk.W)
        self.excel_path_var = tk.StringVar()
        ttk.Entry(file_frame, textvariable=self.excel_path_var, width=60).grid(row=0, column=1, padx=5)
        ttk.Button(file_frame, text="Browse", command=self.browse_excel).grid(row=0, column=2)
        ttk.Button(file_frame, text="Load", command=self.load_data).grid(row=0, column=3, padx=5)

        ttk.Label(file_frame, text="Image Folder:").grid(row=1, column=0, sticky=tk.W)
        self.folder_path_var = tk.StringVar()
        ttk.Entry(file_frame, textvariable=self.folder_path_var, width=60).grid(row=1, column=1, padx=5)
        ttk.Button(file_frame, text="Browse", command=self.browse_folder).grid(row=1, column=2)

        # Navigation
        nav_frame = ttk.Frame(main_frame)
        nav_frame.pack(fill=tk.X, pady=(0, 10))

        self.card_info_var = tk.StringVar(value="Card 0 / 0 | Row: 0 | Voter ID: ")
        ttk.Label(nav_frame, textvariable=self.card_info_var, font=('Helvetica', 12, 'bold'),
                 foreground='blue').pack(side=tk.LEFT)

        ttk.Button(nav_frame, text="Next >>>>", command=self.next_card).pack(side=tk.RIGHT, padx=5)
        ttk.Button(nav_frame, text="<<<< Previous", command=self.prev_card).pack(side=tk.RIGHT)

        # Age in Excel
        excel_frame = ttk.LabelFrame(main_frame, text="Age in Excel (Current Value)", padding="5")
        excel_frame.pack(fill=tk.X, pady=(0, 10))

        self.excel_age_var = tk.StringVar(value="")
        ttk.Label(excel_frame, textvariable=self.excel_age_var, font=('Helvetica', 14, 'bold'),
                 foreground='green').pack(side=tk.LEFT)

        # Results frame
        results_frame = ttk.LabelFrame(main_frame, text="OCR Results - ALL Preprocessing Approaches", padding="10")
        results_frame.pack(fill=tk.X, pady=(0, 10))

        # Create labels for each approach
        self.approach_labels = {}
        approaches = ['Original', 'Contrast 2x', 'Contrast 3x', 'Grayscale', 'Gray+Sharp', 'Binarize', 'Scale 2x', 'Scale 3x']

        for i, approach in enumerate(approaches):
            row_frame = ttk.Frame(results_frame)
            row_frame.pack(fill=tk.X, pady=2)

            # Approach name
            ttk.Label(row_frame, text=f"{approach}:", width=12,
                     font=('Helvetica', 10, 'bold')).pack(side=tk.LEFT)

            # Age result
            age_var = tk.StringVar(value="")
            age_label = ttk.Label(row_frame, textvariable=age_var, width=8,
                                 font=('Helvetica', 10, 'bold'))
            age_label.pack(side=tk.LEFT, padx=5)

            # Pattern info
            pattern_var = tk.StringVar(value="")
            ttk.Label(row_frame, textvariable=pattern_var, width=50,
                     font=('Helvetica', 9)).pack(side=tk.LEFT, padx=5)

            self.approach_labels[approach] = {
                'age_var': age_var,
                'age_label': age_label,
                'pattern_var': pattern_var
            }

        # Bottom: Image display
        image_frame = ttk.LabelFrame(main_frame, text="FULL VOTER CARD", padding="5")
        image_frame.pack(fill=tk.BOTH, expand=True)

        self.image_label = ttk.Label(image_frame)
        self.image_label.pack(fill=tk.BOTH, expand=True)

    def browse_excel(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if path:
            self.excel_path_var.set(path)
            # Auto-pick images folder based on Excel filename
            # e.g., edapadi-lingesh-9.0_excel.xlsx → .edapadi-lingesh-9.0_temp_cards
            excel_name = Path(path).stem  # e.g., "edapadi-lingesh-9.0_excel"
            excel_dir = Path(path).parent

            # Try to find matching temp_cards folder
            base_name = excel_name.replace('_excel', '').replace('-excel', '')
            possible_folders = [
                excel_dir / f".{base_name}_temp_cards",
                excel_dir / f"{base_name}_temp_cards",
                excel_dir / f".{base_name.replace('-', '_')}_temp_cards",
            ]

            for folder in possible_folders:
                if folder.exists():
                    self.folder_path_var.set(str(folder))
                    self.image_folder = folder
                    print(f"Auto-selected folder: {folder}")
                    break

    def browse_folder(self):
        path = filedialog.askdirectory()
        if path:
            self.folder_path_var.set(path)
            self.image_folder = Path(path)

    def load_data(self):
        excel_path = self.excel_path_var.get()
        folder_path = self.folder_path_var.get()

        if not excel_path or not folder_path:
            messagebox.showerror("Error", "Please select both Excel file and Image folder")
            return

        try:
            from openpyxl import load_workbook
            wb = load_workbook(excel_path)
            ws = wb.active

            self.excel_data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Has serial number
                    self.excel_data.append({
                        'serial_no': row[0],
                        'voter_id': row[1] if len(row) > 1 else '',
                        'name': row[2] if len(row) > 2 else '',
                        'age': row[7] if len(row) > 7 else '',  # Age is column 8 (index 7)
                        'gender': row[8] if len(row) > 8 else '',
                    })

            # Load images - try multiple patterns
            self.image_folder = Path(folder_path)
            self.card_images = []

            # Try direct folder
            self.card_images = sorted(list(self.image_folder.glob("*.png")) + list(self.image_folder.glob("*.jpg")))

            # If no images, try recursive search
            if not self.card_images:
                self.card_images = sorted(list(self.image_folder.rglob("*.png")) + list(self.image_folder.rglob("*.jpg")))

            # Sort by numeric filename if possible
            def sort_key(p):
                try:
                    return int(p.stem)
                except:
                    return p.stem
            self.card_images = sorted(self.card_images, key=sort_key)

            # Debug info
            print(f"Excel rows: {len(self.excel_data)}")
            print(f"Images found: {len(self.card_images)}")
            if self.card_images:
                print(f"First image: {self.card_images[0]}")

            if self.excel_data and self.card_images:
                self.current_index = 0
                self.show_card(0)
                messagebox.showinfo("Loaded", f"Loaded {len(self.excel_data)} rows and {len(self.card_images)} images")
            else:
                error_msg = f"Excel rows: {len(self.excel_data)}, Images: {len(self.card_images)}\nFolder: {folder_path}"
                messagebox.showerror("Error", f"No data or images found\n{error_msg}")

        except Exception as e:
            import traceback
            traceback.print_exc()
            messagebox.showerror("Error", f"Failed to load: {str(e)}")

    def show_card(self, index):
        if not self.card_images or index >= len(self.card_images):
            return

        self.current_index = index
        img_path = self.card_images[index]

        # Get excel data for this card
        excel_row = self.excel_data[index] if index < len(self.excel_data) else {}

        # Update info
        voter_id = excel_row.get('voter_id', '') or ''
        name = excel_row.get('name', '') or ''
        self.card_info_var.set(
            f"Card {index + 1} / {len(self.card_images)} | "
            f"Row: {index + 2} | "
            f"Voter ID: {voter_id} | "
            f"Name: {name[:20] if name else ''}"
        )

        # Show age from Excel
        excel_age = excel_row.get('age', '') or ''
        if excel_age:
            self.excel_age_var.set(f"Age in Excel: {excel_age}")
        else:
            self.excel_age_var.set("Age in Excel: MISSING")

        # Load image
        img = Image.open(img_path)

        # Run OCR with all approaches
        results = ocr_with_all_approaches(img)

        # Update results display
        for approach, data in results.items():
            if approach in self.approach_labels:
                labels = self.approach_labels[approach]
                age = data['age']
                pattern = data['pattern']

                labels['age_var'].set(age if age else 'NOT FOUND')
                labels['pattern_var'].set(f"({pattern})")

                # Color coding
                if age and age.isdigit() and 18 <= int(age) <= 99:
                    labels['age_label'].configure(foreground='green')
                else:
                    labels['age_label'].configure(foreground='red')

        # Display image
        display_img = img.copy()
        display_img.thumbnail((800, 400), Image.LANCZOS)
        photo = ImageTk.PhotoImage(display_img)
        self.image_label.configure(image=photo)
        self.image_label.image = photo

    def next_card(self):
        if self.current_index < len(self.card_images) - 1:
            self.show_card(self.current_index + 1)

    def prev_card(self):
        if self.current_index > 0:
            self.show_card(self.current_index - 1)


if __name__ == "__main__":
    root = tk.Tk()
    app = AgeViewerApp(root)
    root.mainloop()
