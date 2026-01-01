"""
Missing Age/Gender Finder Tool v3.0
Loads Excel file, finds missing Age/Gender fields, displays card images, and allows fixing.

Updated for v4.0 Excel format with Source Folder and Card File columns:
- Source Folder (column 11) = folder name containing the image
- Card File (column 12) = image filename (1.png, 2.png, etc.)

Features:
- Import Excel file with voter data
- Auto-detect image folder from Excel location
- Uses Source Folder + Card File to locate exact image
- Show list of rows with missing Age/Gender
- Display voter card image for visual verification
- Manual entry or OCR retry for missing fields
- Save updates back to Excel
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from pathlib import Path
import subprocess
import re

# Install dependencies
def install_packages():
    packages = ['openpyxl', 'pillow', 'pytesseract']
    for pkg in packages:
        try:
            __import__(pkg.replace('-', '_'))
        except ImportError:
            print(f"Installing {pkg}...")
            subprocess.check_call(['uv', 'pip', 'install', pkg])

install_packages()

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from PIL import Image, ImageTk, ImageEnhance

try:
    import pytesseract
except ImportError:
    subprocess.check_call(['uv', 'pip', 'install', 'pytesseract'])
    import pytesseract

TESSERACT_CONFIG = '--psm 6 --oem 1'


class MissingDataFinder:
    def __init__(self, root):
        self.root = root
        self.root.title("Missing Age/Gender Finder v3.0")
        self.root.geometry("1200x800")
        self.root.resizable(True, True)

        # Data
        self.excel_path = None
        self.workbook = None
        self.worksheet = None
        self.image_folder = None
        self.missing_rows = []  # List of row data dictionaries
        self.current_index = 0
        self.current_image = None
        self.changes_made = False

        self.style = ttk.Style()
        self.style.configure('Title.TLabel', font=('Helvetica', 14, 'bold'))
        self.style.configure('Header.TLabel', font=('Helvetica', 11, 'bold'))
        self.style.configure('Big.TButton', font=('Helvetica', 12))

        self.create_widgets()

    def create_widgets(self):
        # Main container
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Top section - File selection
        file_frame = ttk.LabelFrame(main_frame, text="File Selection", padding="10")
        file_frame.pack(fill=tk.X, pady=(0, 10))

        # Excel file row
        excel_row = ttk.Frame(file_frame)
        excel_row.pack(fill=tk.X, pady=2)
        ttk.Label(excel_row, text="Excel File:", width=12).pack(side=tk.LEFT)
        self.excel_path_var = tk.StringVar()
        ttk.Entry(excel_row, textvariable=self.excel_path_var, width=70).pack(side=tk.LEFT, padx=5)
        ttk.Button(excel_row, text="Browse...", command=self.browse_excel).pack(side=tk.LEFT, padx=5)
        ttk.Button(excel_row, text="Load", command=self.load_excel).pack(side=tk.LEFT, padx=5)

        # Image folder row
        img_row = ttk.Frame(file_frame)
        img_row.pack(fill=tk.X, pady=2)
        ttk.Label(img_row, text="Image Folder:", width=12).pack(side=tk.LEFT)
        self.image_folder_var = tk.StringVar()
        ttk.Entry(img_row, textvariable=self.image_folder_var, width=70).pack(side=tk.LEFT, padx=5)
        ttk.Button(img_row, text="Browse...", command=self.browse_image_folder).pack(side=tk.LEFT, padx=5)

        # Stats row
        stats_row = ttk.Frame(file_frame)
        stats_row.pack(fill=tk.X, pady=(5, 0))
        self.stats_var = tk.StringVar(value="No file loaded")
        ttk.Label(stats_row, textvariable=self.stats_var, foreground='#2196F3').pack(side=tk.LEFT)

        # Middle section - Split view
        content_frame = ttk.Frame(main_frame)
        content_frame.pack(fill=tk.BOTH, expand=True, pady=10)

        # Left panel - Missing rows list
        left_frame = ttk.LabelFrame(content_frame, text="Missing Age/Gender Rows", padding="5")
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=False, padx=(0, 5))

        # List with scrollbar
        list_container = ttk.Frame(left_frame)
        list_container.pack(fill=tk.BOTH, expand=True)

        self.row_listbox = tk.Listbox(list_container, width=50, height=25, font=('Courier', 9))
        scrollbar = ttk.Scrollbar(list_container, orient=tk.VERTICAL, command=self.row_listbox.yview)
        self.row_listbox.configure(yscrollcommand=scrollbar.set)

        self.row_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.row_listbox.bind('<<ListboxSelect>>', self.on_row_select)

        # Filter options
        filter_frame = ttk.Frame(left_frame)
        filter_frame.pack(fill=tk.X, pady=(5, 0))

        self.filter_var = tk.StringVar(value="all")
        ttk.Radiobutton(filter_frame, text="All Missing", variable=self.filter_var,
                       value="all", command=self.apply_filter).pack(side=tk.LEFT)
        ttk.Radiobutton(filter_frame, text="Missing Age", variable=self.filter_var,
                       value="age", command=self.apply_filter).pack(side=tk.LEFT)
        ttk.Radiobutton(filter_frame, text="Missing Gender", variable=self.filter_var,
                       value="gender", command=self.apply_filter).pack(side=tk.LEFT)

        # Right panel - Image and editor
        right_frame = ttk.Frame(content_frame)
        right_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(5, 0))

        # Image display
        image_frame = ttk.LabelFrame(right_frame, text="Voter Card Image", padding="5")
        image_frame.pack(fill=tk.BOTH, expand=True)

        self.image_label = ttk.Label(image_frame, text="No image loaded", anchor=tk.CENTER)
        self.image_label.pack(fill=tk.BOTH, expand=True)

        # Editor panel
        editor_frame = ttk.LabelFrame(right_frame, text="Edit Missing Data", padding="10")
        editor_frame.pack(fill=tk.X, pady=(10, 0))

        # Current row info
        info_row = ttk.Frame(editor_frame)
        info_row.pack(fill=tk.X, pady=(0, 10))

        ttk.Label(info_row, text="Row:", style='Header.TLabel').pack(side=tk.LEFT)
        self.current_row_var = tk.StringVar(value="--")
        ttk.Label(info_row, textvariable=self.current_row_var, foreground='#E91E63',
                 font=('Helvetica', 11, 'bold')).pack(side=tk.LEFT, padx=(5, 20))

        ttk.Label(info_row, text="Part No.:", style='Header.TLabel').pack(side=tk.LEFT)
        self.current_part_var = tk.StringVar(value="--")
        ttk.Label(info_row, textvariable=self.current_part_var, foreground='#2196F3',
                 font=('Helvetica', 11, 'bold')).pack(side=tk.LEFT, padx=(5, 20))

        ttk.Label(info_row, text="Image:", style='Header.TLabel').pack(side=tk.LEFT)
        self.current_img_var = tk.StringVar(value="--")
        ttk.Label(info_row, textvariable=self.current_img_var, foreground='#4CAF50',
                 font=('Helvetica', 11)).pack(side=tk.LEFT, padx=5)

        # Voter info display
        voter_info_row = ttk.Frame(editor_frame)
        voter_info_row.pack(fill=tk.X, pady=(0, 10))

        ttk.Label(voter_info_row, text="Voter ID:", width=10).pack(side=tk.LEFT)
        self.voter_id_var = tk.StringVar(value="--")
        ttk.Label(voter_info_row, textvariable=self.voter_id_var, width=15).pack(side=tk.LEFT, padx=(0, 10))

        ttk.Label(voter_info_row, text="Name:", width=8).pack(side=tk.LEFT)
        self.voter_name_var = tk.StringVar(value="--")
        ttk.Label(voter_info_row, textvariable=self.voter_name_var, width=30).pack(side=tk.LEFT)

        # Age and Gender inputs
        input_row = ttk.Frame(editor_frame)
        input_row.pack(fill=tk.X, pady=5)

        # Age input
        age_frame = ttk.Frame(input_row)
        age_frame.pack(side=tk.LEFT, padx=(0, 30))

        ttk.Label(age_frame, text="Age:", style='Header.TLabel').pack(side=tk.LEFT)
        self.age_var = tk.StringVar()
        self.age_entry = ttk.Entry(age_frame, textvariable=self.age_var, width=10, font=('Helvetica', 14))
        self.age_entry.pack(side=tk.LEFT, padx=5)

        self.age_status_var = tk.StringVar(value="")
        ttk.Label(age_frame, textvariable=self.age_status_var, foreground='#F44336').pack(side=tk.LEFT)

        # Gender input
        gender_frame = ttk.Frame(input_row)
        gender_frame.pack(side=tk.LEFT)

        ttk.Label(gender_frame, text="Gender:", style='Header.TLabel').pack(side=tk.LEFT)
        self.gender_var = tk.StringVar()
        self.gender_combo = ttk.Combobox(gender_frame, textvariable=self.gender_var,
                                         values=['', 'Male', 'Female', 'Third Gender'], width=14, font=('Helvetica', 14))
        self.gender_combo.pack(side=tk.LEFT, padx=5)

        self.gender_status_var = tk.StringVar(value="")
        ttk.Label(gender_frame, textvariable=self.gender_status_var, foreground='#F44336').pack(side=tk.LEFT)

        # Buttons row
        btn_row = ttk.Frame(editor_frame)
        btn_row.pack(fill=tk.X, pady=(10, 5))

        ttk.Button(btn_row, text="< Previous", command=self.prev_row, width=12).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_row, text="Next >", command=self.next_row, width=12).pack(side=tk.LEFT, padx=2)

        ttk.Button(btn_row, text="Search Image", command=self.search_image, width=14).pack(side=tk.LEFT, padx=10)
        ttk.Button(btn_row, text="OCR Retry", command=self.ocr_retry, width=12).pack(side=tk.LEFT, padx=2)

        ttk.Button(btn_row, text="Apply", command=self.apply_changes, width=12).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_row, text="Apply & Next", command=self.apply_and_next, width=14).pack(side=tk.LEFT, padx=2)

        # Bottom section - Save
        bottom_frame = ttk.Frame(main_frame)
        bottom_frame.pack(fill=tk.X, pady=(10, 0))

        self.progress_var = tk.StringVar(value="0/0")
        ttk.Label(bottom_frame, textvariable=self.progress_var, font=('Helvetica', 12, 'bold')).pack(side=tk.LEFT)

        ttk.Button(bottom_frame, text="Save Excel", command=self.save_excel,
                  style='Big.TButton').pack(side=tk.RIGHT, padx=5)

        self.save_status_var = tk.StringVar(value="")
        ttk.Label(bottom_frame, textvariable=self.save_status_var, foreground='#4CAF50').pack(side=tk.RIGHT, padx=10)

        # Keyboard shortcuts
        self.root.bind('<Left>', lambda e: self.prev_row())
        self.root.bind('<Right>', lambda e: self.next_row())
        self.root.bind('<Return>', lambda e: self.apply_and_next())
        self.root.bind('<Control-s>', lambda e: self.save_excel())

    def browse_excel(self):
        file_path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if file_path:
            self.excel_path_var.set(file_path)

            # Auto-detect image folder
            excel_path = Path(file_path)
            parent = excel_path.parent

            # Try to find temp_cards folder
            constituency_name = excel_path.stem.replace('_excel', '').replace('_gpu_excel', '')
            possible_folders = [
                parent / f".{constituency_name}_temp_cards",
                parent / f".{constituency_name}_cards",
                parent / constituency_name / "cards",
            ]

            for folder in possible_folders:
                if folder.exists():
                    self.image_folder_var.set(str(folder))
                    break

    def browse_image_folder(self):
        folder = filedialog.askdirectory(title="Select Image Folder")
        if folder:
            self.image_folder_var.set(folder)

    def load_excel(self):
        excel_path = self.excel_path_var.get()
        if not excel_path or not Path(excel_path).exists():
            messagebox.showerror("Error", "Please select a valid Excel file.")
            return

        image_folder = self.image_folder_var.get()
        if not image_folder or not Path(image_folder).exists():
            messagebox.showwarning("Warning", "Image folder not found. Images won't be displayed.")
            self.image_folder = None
        else:
            self.image_folder = Path(image_folder)

        try:
            self.excel_path = Path(excel_path)
            self.workbook = load_workbook(self.excel_path)
            self.worksheet = self.workbook.active

            # Find missing rows
            self.find_missing_rows()

            # Update stats
            total_rows = self.worksheet.max_row - 1  # Exclude header
            missing_count = len(self.missing_rows)
            self.stats_var.set(f"Loaded: {total_rows:,} rows | Missing: {missing_count:,}")

            # Populate list
            self.apply_filter()

            if missing_count > 0:
                self.row_listbox.selection_set(0)
                self.on_row_select(None)

        except Exception as e:
            messagebox.showerror("Error", f"Failed to load Excel: {e}")
            import traceback
            traceback.print_exc()

    def find_missing_rows(self):
        """Find all rows with missing Age or Gender."""
        self.missing_rows = []

        # Column indices for v4.0 format (1-based)
        # Headers: S.No, Part No., Voter ID, Name, Relation Type, Relation Name, House No, Age, Gender, Constituency, Source Folder, Card File
        sno_col = 1           # S.No
        part_col = 2          # Part No.
        voter_id_col = 3
        name_col = 4
        age_col = 8
        gender_col = 9
        source_folder_col = 11  # Source Folder (folder name)
        card_file_col = 12      # Card File (image filename like 1.png)

        for row_num in range(2, self.worksheet.max_row + 1):
            sno_val = self.worksheet.cell(row=row_num, column=sno_col).value
            part_no = self.worksheet.cell(row=row_num, column=part_col).value
            voter_id = self.worksheet.cell(row=row_num, column=voter_id_col).value
            name = self.worksheet.cell(row=row_num, column=name_col).value
            age_val = self.worksheet.cell(row=row_num, column=age_col).value
            gender_val = self.worksheet.cell(row=row_num, column=gender_col).value
            source_folder = self.worksheet.cell(row=row_num, column=source_folder_col).value
            card_file = self.worksheet.cell(row=row_num, column=card_file_col).value

            age_missing = not age_val or str(age_val).strip() == ''
            gender_missing = not gender_val or str(gender_val).strip() == ''

            if age_missing or gender_missing:
                self.missing_rows.append({
                    'row_num': row_num,
                    'sno': sno_val,                           # S.No (1, 2, 3...)
                    'part_no': str(part_no) if part_no else '',
                    'source_folder': source_folder or '',     # Folder name from Excel
                    'card_file': card_file or '',             # Image filename from Excel (1.png)
                    'age': age_val or '',
                    'gender': gender_val or '',
                    'age_missing': age_missing,
                    'gender_missing': gender_missing,
                    'voter_id': voter_id or '',
                    'name': name or ''
                })

    def apply_filter(self):
        """Apply filter and update listbox."""
        self.row_listbox.delete(0, tk.END)

        filter_type = self.filter_var.get()

        for i, row_data in enumerate(self.missing_rows):
            if filter_type == 'age' and not row_data['age_missing']:
                continue
            if filter_type == 'gender' and not row_data['gender_missing']:
                continue

            # Format: Row# | Part# | S.No | Age | Gender
            age_status = '?' if row_data['age_missing'] else str(row_data['age'])[:3]
            gender_status = '?' if row_data['gender_missing'] else row_data['gender'][0] if row_data['gender'] else '?'

            display = f"R{row_data['row_num']:4d} | P:{row_data['part_no']:>2s} | #{row_data['sno']:>3s} | A:{age_status:3s} G:{gender_status}"
            self.row_listbox.insert(tk.END, display)

            # Color code
            if row_data['age_missing'] and row_data['gender_missing']:
                self.row_listbox.itemconfig(tk.END, fg='#F44336')  # Red - both missing
            elif row_data['age_missing']:
                self.row_listbox.itemconfig(tk.END, fg='#FF9800')  # Orange - age missing
            else:
                self.row_listbox.itemconfig(tk.END, fg='#9C27B0')  # Purple - gender missing

        # Update progress
        self.update_progress()

    def update_progress(self):
        """Update progress display."""
        total = len(self.missing_rows)
        fixed = sum(1 for r in self.missing_rows if not r['age_missing'] and not r['gender_missing'])
        self.progress_var.set(f"Fixed: {fixed}/{total}")

    def on_row_select(self, event):
        """Handle row selection in listbox."""
        selection = self.row_listbox.curselection()
        if not selection:
            return

        index = selection[0]

        # Map listbox index to missing_rows index based on filter
        filter_type = self.filter_var.get()
        filtered_indices = []

        for i, row_data in enumerate(self.missing_rows):
            if filter_type == 'age' and not row_data['age_missing']:
                continue
            if filter_type == 'gender' and not row_data['gender_missing']:
                continue
            filtered_indices.append(i)

        if index < len(filtered_indices):
            self.current_index = filtered_indices[index]
            self.display_current_row()

    def display_current_row(self):
        """Display the current row's image and data."""
        if not self.missing_rows or self.current_index >= len(self.missing_rows):
            return

        row_data = self.missing_rows[self.current_index]

        # Update info displays
        self.current_row_var.set(f"{row_data['row_num']}")
        self.current_part_var.set(f"{row_data['part_no']}")

        # Show image path: source_folder/card_file
        source_folder = row_data['source_folder'] or f"Part-{row_data['part_no']}"
        card_file = row_data['card_file'] or f"{row_data['sno']}.png"
        self.current_img_var.set(f"{source_folder[:25]}.../{card_file}" if len(source_folder) > 25 else f"{source_folder}/{card_file}")

        self.voter_id_var.set(row_data['voter_id'])
        self.voter_name_var.set(row_data['name'])

        # Update input fields
        self.age_var.set(row_data['age'])
        self.gender_var.set(row_data['gender'])

        # Update status indicators
        self.age_status_var.set("MISSING" if row_data['age_missing'] else "")
        self.gender_status_var.set("MISSING" if row_data['gender_missing'] else "")

        # Load and display image using Source Folder and Card File from Excel
        self.load_image(row_data['source_folder'], row_data['card_file'])

    def load_image(self, source_folder, card_file):
        """Load and display the voter card image.

        Args:
            source_folder: The folder name containing the image (from Source Folder column)
            card_file: The image filename (from Card File column, e.g., 1.png)
        """
        if not self.image_folder:
            self.image_label.configure(image='', text="Image folder not set")
            return

        if not source_folder:
            self.image_label.configure(image='', text=f"No Source Folder in Excel\n\nCard File: {card_file}\n\nClick 'Search Image' to find by Voter ID")
            return

        if not card_file:
            self.image_label.configure(image='', text=f"No Card File in Excel\n\nSource Folder: {source_folder}\n\nClick 'Search Image' to find by Voter ID")
            return

        # Construct image path: {image_folder}/{source_folder}/{card_file}
        image_path = self.image_folder / source_folder / card_file

        if not image_path.exists():
            # Try alternate extensions
            card_stem = Path(card_file).stem
            alt_paths = [
                self.image_folder / source_folder / f"{card_stem}.png",
                self.image_folder / source_folder / f"{card_stem}.jpg",
                self.image_folder / source_folder / f"{card_stem}.jpeg",
            ]
            found = False
            for alt in alt_paths:
                if alt.exists():
                    image_path = alt
                    found = True
                    break

            if not found:
                # Check if the folder exists
                folder_path = self.image_folder / source_folder
                if not folder_path.exists():
                    self.image_label.configure(image='',
                        text=f"Folder not found:\n{folder_path}\n\nClick 'Search Image' to find by Voter ID")
                else:
                    # List available images in folder
                    try:
                        available = sorted(list(folder_path.glob("*.png")) + list(folder_path.glob("*.jpg")),
                                         key=lambda x: int(x.stem) if x.stem.isdigit() else 0)[:5]
                        available_str = ", ".join([p.name for p in available])
                        self.image_label.configure(image='',
                            text=f"Image not found:\n{card_file}\n\nFolder: {source_folder}\nAvailable: {available_str}...\n\nClick 'Search Image' to find by Voter ID")
                    except:
                        self.image_label.configure(image='', text=f"Image not found:\n{image_path}\n\nClick 'Search Image' to find by Voter ID")
                return

        try:
            # Load and resize image to fit
            img = Image.open(image_path)

            # Calculate resize to fit in label (max 700x500)
            max_width, max_height = 700, 500
            ratio = min(max_width / img.width, max_height / img.height)
            new_size = (int(img.width * ratio), int(img.height * ratio))
            img = img.resize(new_size, Image.LANCZOS)

            # Convert to PhotoImage
            self.current_image = ImageTk.PhotoImage(img)
            self.image_label.configure(image=self.current_image, text='')

            # Store PIL image for OCR
            self.current_pil_image = Image.open(image_path)

        except Exception as e:
            self.image_label.configure(image='', text=f"Error loading image:\n{e}")

    def search_image(self):
        """Search for the correct image by matching Voter ID and Name using OCR."""
        if not self.missing_rows or self.current_index >= len(self.missing_rows):
            messagebox.showwarning("Warning", "No row selected")
            return

        if not self.image_folder:
            messagebox.showwarning("Warning", "Image folder not set")
            return

        row_data = self.missing_rows[self.current_index]
        voter_id = row_data.get('voter_id', '')
        name = row_data.get('name', '')
        source_folder = row_data.get('source_folder', '')

        if not voter_id:
            messagebox.showwarning("Warning", "No Voter ID found for this row")
            return

        # Update UI to show searching
        self.image_label.configure(image='', text=f"Searching for Voter ID: {voter_id}...\n\nThis may take a while as we OCR each image.")
        self.root.update()

        # Perform search
        found_path = self.search_image_by_voter_id(voter_id, name, source_folder)

        if found_path:
            try:
                # Load and display the found image
                img = Image.open(found_path)

                # Calculate resize to fit in label (max 700x500)
                max_width, max_height = 700, 500
                ratio = min(max_width / img.width, max_height / img.height)
                new_size = (int(img.width * ratio), int(img.height * ratio))
                img_resized = img.resize(new_size, Image.LANCZOS)

                # Convert to PhotoImage
                self.current_image = ImageTk.PhotoImage(img_resized)
                self.image_label.configure(image=self.current_image, text='')

                # Store PIL image for OCR
                self.current_pil_image = Image.open(found_path)

                # Update the image info
                found_path_obj = Path(found_path)
                self.current_img_var.set(f"Found: {found_path_obj.parent.name}/{found_path_obj.name}")

                messagebox.showinfo("Found", f"Image found!\n\nPath: {found_path}")

            except Exception as e:
                messagebox.showerror("Error", f"Found image but failed to load:\n{found_path}\n\n{e}")
        else:
            self.image_label.configure(image='', text=f"Image NOT found for:\nVoter ID: {voter_id}\nName: {name}\n\nSearched in: {source_folder or 'all folders'}")
            messagebox.showwarning("Not Found", f"Could not find image matching:\nVoter ID: {voter_id}\nName: {name}")

    def search_image_by_voter_id(self, voter_id, name, source_folder):
        """Search for image by matching Voter ID and Name using OCR."""
        if not self.image_folder or not voter_id:
            return None

        # First, try the specific folder from Source Folder column
        search_folders = []
        if source_folder:
            specific_folder = self.image_folder / source_folder
            if specific_folder.exists():
                search_folders.append(specific_folder)

        # If not found in specific folder, search all folders
        if not search_folders:
            search_folders = [f for f in self.image_folder.iterdir() if f.is_dir()]

        voter_id_clean = str(voter_id).strip().upper()

        for folder in search_folders:
            # Look for both .png and .jpg files
            image_files = sorted(
                list(folder.glob("*.png")) + list(folder.glob("*.jpg")),
                key=lambda x: int(x.stem) if x.stem.isdigit() else 0
            )

            # Update status during search
            self.image_label.configure(text=f"Searching in: {folder.name}\n({len(image_files)} images)")
            self.root.update()

            for img_path in image_files:
                try:
                    img = Image.open(img_path)
                    text = pytesseract.image_to_string(img, lang='tam+eng', config=TESSERACT_CONFIG)

                    # Check if Voter ID matches
                    if voter_id_clean in text.upper():
                        return str(img_path)

                    # Also try matching name if Voter ID not found
                    if name and len(name) > 3:
                        if name in text:
                            return str(img_path)

                except Exception as e:
                    continue

        return None

    def ocr_retry(self):
        """Retry OCR on the current image to extract Age and Gender."""
        if not hasattr(self, 'current_pil_image'):
            messagebox.showwarning("Warning", "No image loaded")
            return

        try:
            img = self.current_pil_image
            width, height = img.size

            # Crop bottom portion where Age/Gender typically appears
            bottom_crop = img.crop((0, int(height * 0.6), width, height))

            # Try different preprocessing approaches
            approaches = [
                ('original', lambda i: i),
                ('contrast', lambda i: ImageEnhance.Contrast(i).enhance(2.0)),
                ('grayscale', lambda i: i.convert('L')),
                ('binarize', lambda i: i.convert('L').point(lambda x: 0 if x < 140 else 255, '1')),
                ('scale', lambda i: i.resize((i.size[0] * 2, i.size[1] * 2), Image.LANCZOS)),
            ]

            found_age = ''
            found_gender = ''

            for name, transform in approaches:
                try:
                    processed = transform(bottom_crop)
                    text = pytesseract.image_to_string(processed, lang='tam+eng', config=TESSERACT_CONFIG)

                    # Extract age
                    if not found_age:
                        age_match = re.search(r'வயது\s*:\s*(\d+)', text)
                        if age_match:
                            found_age = age_match.group(1)

                    # Extract gender
                    if not found_gender:
                        if 'பாலினம்' in text:
                            if 'ஆண்' in text:
                                found_gender = 'Male'
                            elif 'பெண்' in text:
                                found_gender = 'Female'
                            elif 'திருநங்கை' in text or 'மூன்றாம்' in text or 'Third' in text:
                                found_gender = 'Third Gender'

                    if found_age and found_gender:
                        break

                except Exception:
                    continue

            # If not found in bottom, try full image
            if not found_age or not found_gender:
                for name, transform in approaches:
                    try:
                        processed = transform(img)
                        text = pytesseract.image_to_string(processed, lang='tam+eng', config=TESSERACT_CONFIG)

                        if not found_age:
                            age_match = re.search(r'வயது\s*:\s*(\d+)', text)
                            if age_match:
                                found_age = age_match.group(1)

                        if not found_gender:
                            if 'பாலினம்' in text:
                                if 'ஆண்' in text:
                                    found_gender = 'Male'
                                elif 'பெண்' in text:
                                    found_gender = 'Female'
                                elif 'திருநங்கை' in text or 'மூன்றாம்' in text or 'Third' in text:
                                    found_gender = 'Third Gender'

                        if found_age and found_gender:
                            break

                    except Exception:
                        continue

            # Update fields if found
            if found_age:
                self.age_var.set(found_age)
                self.age_status_var.set("FOUND!")
            else:
                self.age_status_var.set("NOT FOUND")

            if found_gender:
                self.gender_var.set(found_gender)
                self.gender_status_var.set("FOUND!")
            else:
                self.gender_status_var.set("NOT FOUND")

            result_msg = f"Age: {'Found - ' + found_age if found_age else 'Not found'}\nGender: {'Found - ' + found_gender if found_gender else 'Not found'}"
            messagebox.showinfo("OCR Result", result_msg)

        except Exception as e:
            messagebox.showerror("OCR Error", f"Failed to run OCR:\n{e}")

    def apply_changes(self):
        """Apply changes to the current row."""
        if not self.missing_rows or self.current_index >= len(self.missing_rows):
            return

        row_data = self.missing_rows[self.current_index]
        row_num = row_data['row_num']

        age_val = self.age_var.get().strip()
        gender_val = self.gender_var.get().strip()

        # Validate age
        if age_val:
            try:
                age_int = int(age_val)
                if age_int < 18 or age_int > 120:
                    messagebox.showwarning("Warning", "Age should be between 18 and 120")
                    return
            except ValueError:
                messagebox.showerror("Error", "Age must be a number")
                return

        # Update worksheet (v4.0 format: Age is column 8, Gender is column 9)
        self.worksheet.cell(row=row_num, column=8, value=age_val if age_val else None)
        self.worksheet.cell(row=row_num, column=9, value=gender_val if gender_val else None)

        # Remove yellow fill if data is now complete
        if age_val:
            self.worksheet.cell(row=row_num, column=8).fill = PatternFill()
        if gender_val:
            self.worksheet.cell(row=row_num, column=9).fill = PatternFill()

        # Update tracking
        row_data['age'] = age_val
        row_data['gender'] = gender_val
        row_data['age_missing'] = not age_val
        row_data['gender_missing'] = not gender_val

        self.changes_made = True
        self.save_status_var.set("Unsaved changes")

        # Update display
        self.age_status_var.set("" if age_val else "MISSING")
        self.gender_status_var.set("" if gender_val else "MISSING")

        # Refresh list
        self.apply_filter()

        # Re-select current item
        self.highlight_current_in_list()

    def apply_and_next(self):
        """Apply changes and move to next row."""
        self.apply_changes()
        self.next_row()

    def prev_row(self):
        """Go to previous row."""
        if self.current_index > 0:
            self.current_index -= 1
            self.display_current_row()
            self.highlight_current_in_list()

    def next_row(self):
        """Go to next row."""
        if self.current_index < len(self.missing_rows) - 1:
            self.current_index += 1
            self.display_current_row()
            self.highlight_current_in_list()

    def highlight_current_in_list(self):
        """Highlight current row in listbox."""
        # Find position in filtered list
        filter_type = self.filter_var.get()
        list_index = 0

        for i, row_data in enumerate(self.missing_rows):
            if filter_type == 'age' and not row_data['age_missing']:
                continue
            if filter_type == 'gender' and not row_data['gender_missing']:
                continue

            if i == self.current_index:
                self.row_listbox.selection_clear(0, tk.END)
                self.row_listbox.selection_set(list_index)
                self.row_listbox.see(list_index)
                break

            list_index += 1

    def save_excel(self):
        """Save changes to Excel file."""
        if not self.workbook:
            messagebox.showwarning("Warning", "No Excel file loaded")
            return

        try:
            # Create backup
            backup_path = self.excel_path.with_suffix('.xlsx.bak')
            if self.excel_path.exists():
                import shutil
                shutil.copy(self.excel_path, backup_path)

            # Save
            self.workbook.save(self.excel_path)

            self.changes_made = False
            self.save_status_var.set(f"Saved! Backup: {backup_path.name}")

            # Count remaining missing
            remaining = sum(1 for r in self.missing_rows if r['age_missing'] or r['gender_missing'])
            messagebox.showinfo("Saved", f"Excel saved successfully!\n\nRemaining missing: {remaining}")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to save:\n{e}")

    def on_closing(self):
        """Handle window close."""
        if self.changes_made:
            result = messagebox.askyesnocancel(
                "Unsaved Changes",
                "You have unsaved changes. Save before closing?"
            )
            if result is None:  # Cancel
                return
            elif result:  # Yes
                self.save_excel()

        self.root.destroy()


def main():
    root = tk.Tk()
    app = MissingDataFinder(root)
    root.protocol("WM_DELETE_WINDOW", app.on_closing)
    root.mainloop()


if __name__ == "__main__":
    main()
