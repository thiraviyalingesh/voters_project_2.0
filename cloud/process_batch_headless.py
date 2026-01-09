#!/usr/bin/env python3
"""
Voter Analytics - Headless Batch Processor v8.3 (Ubuntu VM Optimized)
Based on v4.0 Windows version with Linux multiprocessing fixes.
v8.3: Added Name to Phase 3 enhanced OCR (now fixes Name/Age/Gender)
v8.2: Fixed semaphore leak + checkpoint saves every 50 cards in Phase 3

Features:
- Phase 1: PDF extraction
- Phase 2: OCR processing
- Phase 3: Enhanced OCR for missing Name/Age/Gender
- Phase 4: Excel generation with full formatting
- Ntfy push notifications
- Checkpoint/resume capability
- OMP_THREAD_LIMIT fix for Linux
"""

import os
# --- CRITICAL: MUST BE BEFORE ANY OTHER IMPORTS ---
# Prevents Tesseract/OpenMP thread explosion (8 cores * 8 threads = 64 threads = HANG)
os.environ["OMP_THREAD_LIMIT"] = "1"
os.environ["MKL_NUM_THREADS"] = "1"
os.environ["OPENBLAS_NUM_THREADS"] = "1"

import re
import sys
import argparse
from pathlib import Path
import time
import json
import requests
import multiprocessing
from concurrent.futures import ProcessPoolExecutor, as_completed
import gc

# Import image processing libraries
import fitz  # PyMuPDF
from PIL import Image, ImageEnhance
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import pytesseract

# Configuration
NTFY_TOPIC = None
NUM_WORKERS = max(1, (os.cpu_count() or 8) - 1)


def send_notification(title, message, topic=None):
    """Send push notification via Ntfy."""
    topic = topic or NTFY_TOPIC
    if not topic:
        return
    try:
        clean_title = title.encode('ascii', 'ignore').decode('ascii').strip() or "Notification"
        requests.post(f"https://ntfy.sh/{topic}", headers={"Title": clean_title},
                      data=message.encode('utf-8'), timeout=10)
        print(f"[NOTIFY] {clean_title}")
    except Exception as e:
        print(f"[NOTIFY ERROR] {e}")


def log(message):
    """Print log message with timestamp."""
    timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{timestamp}] {message}", flush=True)


def format_time(seconds):
    """Format seconds to human readable string."""
    hours = int(seconds // 3600)
    minutes = int((seconds % 3600) // 60)
    secs = int(seconds % 60)
    if hours > 0:
        return f"{hours}h {minutes}m {secs}s"
    elif minutes > 0:
        return f"{minutes}m {secs}s"
    return f"{secs}s"


def extract_part_number(pdf_name):
    """Extract part number from PDF filename."""
    if not pdf_name:
        return ''
    match = re.search(r'-TAM-(\d+)-WI', pdf_name, re.IGNORECASE)
    if match:
        return match.group(1)
    match = re.search(r'-(\d+)-WI$', pdf_name, re.IGNORECASE)
    if match:
        return match.group(1)
    match = re.search(r'(\d+)[^0-9]*WI', pdf_name, re.IGNORECASE)
    if match:
        return match.group(1)
    match = re.search(r'-(\d+)$', pdf_name)
    if match:
        return match.group(1)
    return ''


def clean_ocr_text(text):
    """Clean common OCR artifacts from text."""
    if not text:
        return ''
    text = re.sub(r'\s*Photo\s*is\s*', ' ', text, flags=re.IGNORECASE)
    text = re.sub(r'\s*available\s*', ' ', text, flags=re.IGNORECASE)
    text = re.sub(r'^[\s\-–.,:]+|[\s\-–.,:]+$', '', text)
    text = re.sub(r'\s+', ' ', text)
    return text.strip()


def parse_voter_card(text):
    """Parse OCR text from voter card to extract structured data."""
    data = {
        'serial_no': '', 'voter_id': '', 'name': '', 'relation_name': '',
        'relation_type': '', 'house_no': '', 'age': '', 'gender': ''
    }

    # Extract Voter ID
    for pattern in [r'\b([A-Z]{2,3}\d{6,10})\b', r'\b([A-Z0-9]{2,3}\d{6,10})\b']:
        match = re.search(pattern, text)
        if match and len(match.group(1)) >= 9:
            data['voter_id'] = match.group(1)
            break

    for line in text.split('\n'):
        line = line.strip()
        if not line:
            continue

        # Serial number
        if not data['serial_no']:
            serial_match = re.match(r'^(\d{1,4})\s*$', line)
            if serial_match:
                data['serial_no'] = serial_match.group(1)
            else:
                serial_match = re.match(r'^(\d{1,4})\s+\S', line)
                if serial_match and int(serial_match.group(1)) < 2000:
                    data['serial_no'] = serial_match.group(1)

        # Name
        if 'பெயர்' in line and ':' in line and 'தந்தை' not in line and 'கணவர்' not in line:
            name_part = clean_ocr_text(line.split(':', 1)[-1])
            if name_part and not data['name']:
                data['name'] = name_part

        # Father's name
        if ('தந்தை' in line or 'தந்தையின்' in line) and 'பெயர்' in line and ':' in line:
            rel_part = clean_ocr_text(line.split(':', 1)[-1])
            if rel_part and not data['relation_name']:
                data['relation_name'] = rel_part
                data['relation_type'] = 'Father'

        # Husband's name
        if ('கணவர்' in line or 'கணவரின்' in line) and ':' in line:
            rel_part = clean_ocr_text(line.split(':', 1)[-1])
            if rel_part and not data['relation_name']:
                data['relation_name'] = rel_part
                data['relation_type'] = 'Husband'

        # Mother's name
        if ('தாய்' in line or 'தாயின்' in line) and 'பெயர்' in line and ':' in line:
            rel_part = clean_ocr_text(line.split(':', 1)[-1])
            if rel_part and not data['relation_name']:
                data['relation_name'] = rel_part
                data['relation_type'] = 'Mother'

        # House number
        if ('வீட்டு' in line or 'ட்டு' in line) and 'எண்' in line and ':' in line:
            house_part = clean_ocr_text(line.split(':', 1)[-1])
            if house_part and not data['house_no']:
                data['house_no'] = house_part

        # Age
        if 'வயது' in line and ':' in line:
            age_match = re.search(r'வயது\s*:\s*(\d+)', line)
            if age_match:
                data['age'] = age_match.group(1)

        # Gender
        if 'பாலினம்' in line:
            if 'ஆண்' in line:
                data['gender'] = 'Male'
            elif 'பெண்' in line:
                data['gender'] = 'Female'
            elif 'திருநங்கை' in line or 'மூன்றாம்' in line:
                data['gender'] = 'Third Gender'

    return data


# --- WORKER FUNCTIONS (Must be at top level for Spawn) ---

def ocr_single_card(args):
    """Worker function to OCR a single voter card image."""
    jpg_path, global_idx, pdf_name = args
    try:
        img = Image.open(jpg_path)
        custom_config = r'-c omp_thread_limit=1'
        text = pytesseract.image_to_string(img, lang='tam', config=custom_config)
        data = parse_voter_card(text)
        stem = Path(jpg_path).stem
        return global_idx, stem, data, pdf_name
    except Exception as e:
        return global_idx, str(global_idx), None, pdf_name


def enhanced_ocr_name_age_gender(args):
    """Worker function for enhanced OCR focused on Name, Age and Gender."""
    jpg_path, global_idx, need_name, need_age, need_gender = args
    try:
        img = Image.open(str(jpg_path))
        width, height = img.size

        # Crop regions: top for name, bottom for age/gender
        top_crop = img.crop((0, 0, width, int(height * 0.45)))
        bottom_crop = img.crop((0, int(height * 0.65), width, height))

        approaches = [
            ('original', lambda i: i),
            ('contrast', lambda i: ImageEnhance.Contrast(i).enhance(2.0)),
            ('grayscale_sharp', lambda i: ImageEnhance.Sharpness(i.convert('L')).enhance(2.0)),
            ('binarize', lambda i: i.convert('L').point(lambda x: 0 if x < 140 else 255, '1')),
            ('scale_2x', lambda i: i.resize((i.size[0] * 2, i.size[1] * 2), Image.LANCZOS)),
        ]

        result = {'name': '', 'age': '', 'gender': ''}
        custom_config = r'-c omp_thread_limit=1'

        # Try top crop for Name
        if need_name:
            for name, transform in approaches:
                try:
                    processed_img = transform(top_crop)
                    text = pytesseract.image_to_string(processed_img, lang='tam+eng', config=custom_config)

                    if not result['name']:
                        # Look for name pattern: பெயர் : <name>
                        for line in text.split('\n'):
                            if 'பெயர்' in line and ':' in line and 'தந்தை' not in line and 'கணவர்' not in line and 'தாய்' not in line:
                                name_part = line.split(':', 1)[-1].strip()
                                # Clean OCR artifacts
                                name_part = re.sub(r'\s*Photo\s*is\s*', ' ', name_part, flags=re.IGNORECASE)
                                name_part = re.sub(r'\s*available\s*', ' ', name_part, flags=re.IGNORECASE)
                                name_part = re.sub(r'^[\s\-–.,:]+|[\s\-–.,:]+$', '', name_part)
                                name_part = re.sub(r'\s+', ' ', name_part).strip()
                                if name_part and len(name_part) > 1:
                                    result['name'] = name_part
                                    break

                    if result['name']:
                        break
                except:
                    continue

        # Try bottom crop for Age/Gender
        if need_age or need_gender:
            for name, transform in approaches:
                try:
                    processed_img = transform(bottom_crop)
                    text = pytesseract.image_to_string(processed_img, lang='tam+eng', config=custom_config)

                    if need_age and not result['age']:
                        age_match = re.search(r'வயது\s*:\s*(\d+)', text)
                        if age_match:
                            result['age'] = age_match.group(1)

                    if need_gender and not result['gender']:
                        if 'பாலினம்' in text:
                            if 'ஆண்' in text:
                                result['gender'] = 'Male'
                            elif 'பெண்' in text:
                                result['gender'] = 'Female'
                            elif 'திருநங்கை' in text or 'மூன்றாம்' in text:
                                result['gender'] = 'Third Gender'

                    if (not need_age or result['age']) and (not need_gender or result['gender']):
                        break
                except:
                    continue

        # Try full image if still missing anything
        still_need_name = need_name and not result['name']
        still_need_age = need_age and not result['age']
        still_need_gender = need_gender and not result['gender']

        if still_need_name or still_need_age or still_need_gender:
            for name, transform in approaches:
                try:
                    processed_img = transform(img)
                    text = pytesseract.image_to_string(processed_img, lang='tam+eng', config=custom_config)

                    if still_need_name and not result['name']:
                        for line in text.split('\n'):
                            if 'பெயர்' in line and ':' in line and 'தந்தை' not in line and 'கணவர்' not in line and 'தாய்' not in line:
                                name_part = line.split(':', 1)[-1].strip()
                                name_part = re.sub(r'\s*Photo\s*is\s*', ' ', name_part, flags=re.IGNORECASE)
                                name_part = re.sub(r'\s*available\s*', ' ', name_part, flags=re.IGNORECASE)
                                name_part = re.sub(r'^[\s\-–.,:]+|[\s\-–.,:]+$', '', name_part)
                                name_part = re.sub(r'\s+', ' ', name_part).strip()
                                if name_part and len(name_part) > 1:
                                    result['name'] = name_part
                                    break

                    if still_need_age and not result['age']:
                        age_match = re.search(r'வயது\s*:\s*(\d+)', text)
                        if age_match:
                            result['age'] = age_match.group(1)

                    if still_need_gender and not result['gender']:
                        if 'பாலினம்' in text:
                            if 'ஆண்' in text:
                                result['gender'] = 'Male'
                            elif 'பெண்' in text:
                                result['gender'] = 'Female'

                    # Check if we got everything we needed
                    got_name = not need_name or result['name']
                    got_age = not need_age or result['age']
                    got_gender = not need_gender or result['gender']
                    if got_name and got_age and got_gender:
                        break
                except:
                    continue

        return global_idx, result
    except Exception:
        return global_idx, None


def extract_cards_from_pdf_sequential(pdf_path, temp_base_dir, pdf_index):
    """Extract cards from a single PDF."""
    try:
        pdf_path = Path(pdf_path)
        pdf_name = pdf_path.stem
        output_path = Path(temp_base_dir) / pdf_name
        output_path.mkdir(parents=True, exist_ok=True)

        doc = fitz.open(str(pdf_path))
        num_pages = len(doc)
        card_count = 0

        for page_num in range(3, num_pages - 1):
            page = doc[page_num]
            pix = page.get_pixmap(matrix=fitz.Matrix(1.5, 1.5))
            page_img = Image.open(io.BytesIO(pix.tobytes("png")))

            page_width, page_height = page_img.size
            header_height = int(page_height * 0.035)
            footer_height = int(page_height * 0.025)
            content_height = page_height - header_height - footer_height

            card_width = page_width // 3
            row_height = content_height // 10

            for row in range(10):
                for col in range(3):
                    x1 = col * card_width + 1
                    y1 = header_height + row * row_height + 1
                    x2 = x1 + card_width - 2
                    y2 = y1 + row_height - 2

                    card_img = page_img.crop((x1, y1, x2, y2))

                    # Fast brightness check
                    try:
                        pixels = list(card_img.getdata())
                        sample_size = min(100, len(pixels))
                        if sample_size > 0:
                            step = max(1, len(pixels) // sample_size)
                            sampled = [pixels[i * step] for i in range(sample_size)]
                            avg_brightness = sum(sum(p[:3]) / 3 for p in sampled) / sample_size
                            if avg_brightness > 252:
                                continue
                    except:
                        pass

                    card_count += 1
                    card_img.save(output_path / f"{card_count}.png", "PNG", compress_level=1)

        doc.close()
        return pdf_index, pdf_name, card_count, str(output_path)
    except Exception as e:
        log(f"Error extracting from {pdf_path}: {e}")
        return pdf_index, Path(pdf_path).stem, 0, None


class CheckpointManager:
    """Manages saving and loading checkpoint data for resume capability."""

    def __init__(self, checkpoint_path):
        self.checkpoint_path = Path(checkpoint_path)
        self.data = {
            'phase': 0, 'constituency_name': '', 'folder_path': '', 'total_pdfs': 0,
            'extracted_pdfs': {}, 'ocr_results': {}, 'enhanced_ocr_done': [],
            'all_cards': [], 'start_time': 0, 'elapsed_before_resume': 0
        }

    def exists(self):
        return self.checkpoint_path.exists()

    def load(self):
        if self.exists():
            try:
                with open(self.checkpoint_path, 'r', encoding='utf-8') as f:
                    self.data = json.load(f)
                log(f"Loaded checkpoint: Phase {self.data['phase']}")
                return True
            except Exception as e:
                log(f"Checkpoint load error: {e}")
                return False
        return False

    def save(self):
        temp_path = self.checkpoint_path.with_suffix('.tmp')
        try:
            with open(temp_path, 'w', encoding='utf-8') as f:
                json.dump(self.data, f, ensure_ascii=False)
            temp_path.replace(self.checkpoint_path)
        except Exception as e:
            log(f"Error saving checkpoint: {e}")

    def delete(self):
        if self.exists():
            self.checkpoint_path.unlink()
            log("Checkpoint deleted")


def process_constituency(folder_path, ntfy_topic=None, num_workers=None, cleanup=True):
    """Process a single constituency folder."""
    global NTFY_TOPIC
    NTFY_TOPIC = ntfy_topic

    folder_path = Path(folder_path)
    if not folder_path.exists():
        log(f"ERROR: Folder not found: {folder_path}")
        return False

    constituency_name = folder_path.name
    pdf_files = sorted(folder_path.glob("*.pdf"))
    total_pdfs = len(pdf_files)

    if total_pdfs == 0:
        log(f"ERROR: No PDF files found in {folder_path}")
        return False

    log(f"Processing: {constituency_name}")
    log(f"PDFs found: {total_pdfs}")

    if num_workers is None:
        num_workers = NUM_WORKERS
    log(f"Using {num_workers} workers (spawn context)")

    checkpoint_path = folder_path.parent / f".{constituency_name}_checkpoint.json"
    checkpoint = CheckpointManager(checkpoint_path)

    image_dir = folder_path.parent / f".{constituency_name}_temp_cards"
    image_dir.mkdir(parents=True, exist_ok=True)

    resume_mode = checkpoint.load()
    if resume_mode:
        current_phase = checkpoint.data['phase']
        pdf_card_info = checkpoint.data.get('extracted_pdfs', {})
        elapsed_before = checkpoint.data.get('elapsed_before_resume', 0)
    else:
        current_phase = 0
        pdf_card_info = {}
        elapsed_before = 0
        checkpoint.data['constituency_name'] = constituency_name
        checkpoint.data['folder_path'] = str(folder_path)
        checkpoint.data['total_pdfs'] = total_pdfs

    start_time = time.time()

    def get_elapsed():
        return time.time() - start_time + elapsed_before

    send_notification("Processing Started", f"Constituency: {constituency_name}\nPDFs: {total_pdfs}\nWorkers: {num_workers}")

    # ===== PHASE 1: Extract cards from PDFs =====
    if current_phase < 1:
        log("Phase 1/4: Extracting cards from PDFs...")
        completed_pdfs = 0
        total_cards = 0

        for idx, pdf in enumerate(pdf_files):
            if pdf.stem in pdf_card_info:
                continue
            pdf_idx, pdf_name, card_count, output_path = extract_cards_from_pdf_sequential(str(pdf), str(image_dir), idx)
            pdf_card_info[pdf_name] = (card_count, output_path)
            completed_pdfs += 1
            total_cards += card_count
            log(f"  Extracted: {completed_pdfs}/{total_pdfs} PDFs ({total_cards:,} cards) [{format_time(get_elapsed())}]")
            checkpoint.data['extracted_pdfs'] = pdf_card_info
            checkpoint.save()

        # Build card list
        all_cards = []
        global_idx = 0
        for pdf_name, (card_count, output_path) in pdf_card_info.items():
            if output_path:
                pdf_card_dir = Path(output_path)
                try:
                    png_files = sorted(pdf_card_dir.glob("*.png"), key=lambda x: int(x.stem) if x.stem.isdigit() else 0)
                    for png_file in png_files:
                        all_cards.append((str(png_file), global_idx, pdf_name))
                        global_idx += 1
                except:
                    pass

        checkpoint.data['all_cards'] = all_cards
        checkpoint.data['phase'] = 1
        checkpoint.save()
        current_phase = 1
        log(f"  Phase 1 complete: {len(all_cards):,} cards extracted")

    # ===== PHASE 2: OCR all cards =====
    if current_phase < 2:
        log("Phase 2/4: OCR processing all cards...")
        all_cards = checkpoint.data.get('all_cards', [])
        total_cards_to_ocr = len(all_cards)

        ocr_data = checkpoint.data.get('ocr_results', {})
        completed_indices = set(int(k) for k in ocr_data.keys())
        log(f"  Already completed: {len(completed_indices):,} cards")

        cards_to_ocr = [(p, idx, name) for p, idx, name in all_cards if idx not in completed_indices]
        ocr_results = {int(k): tuple(v) for k, v in ocr_data.items()}
        completed_ocr = len(completed_indices)

        if cards_to_ocr:
            log(f"  OCR remaining: {len(cards_to_ocr):,} cards")
            ctx = multiprocessing.get_context('spawn')

            with ProcessPoolExecutor(max_workers=num_workers, mp_context=ctx) as executor:
                futures = {executor.submit(ocr_single_card, card): card[1] for card in cards_to_ocr}

                for future in as_completed(futures):
                    try:
                        global_idx, s_no, data, pdf_name = future.result()
                        ocr_results[global_idx] = (s_no, data, pdf_name)
                        checkpoint.data['ocr_results'][str(global_idx)] = (s_no, data, pdf_name)
                        completed_ocr += 1

                        if completed_ocr % 100 == 0 or completed_ocr == total_cards_to_ocr:
                            rate = (completed_ocr - len(completed_indices)) / max(1, get_elapsed() - elapsed_before)
                            remaining = (total_cards_to_ocr - completed_ocr) / max(1, rate) if rate > 0 else 0
                            log(f"  OCR: {completed_ocr:,}/{total_cards_to_ocr:,} [{format_time(get_elapsed())}, ~{format_time(remaining)} left]")

                        if completed_ocr % 500 == 0:
                            checkpoint.data['elapsed_before_resume'] = get_elapsed()
                            checkpoint.save()
                    except Exception as e:
                        log(f"  OCR error: {e}")

        checkpoint.data['phase'] = 2
        checkpoint.save()
        current_phase = 2
        gc.collect()  # Clean up multiprocessing resources
        log(f"  Phase 2 complete: {len(ocr_results):,} cards OCR'd")
    else:
        all_cards = checkpoint.data.get('all_cards', [])
        ocr_results = {int(k): tuple(v) for k, v in checkpoint.data.get('ocr_results', {}).items()}

    # ===== PHASE 3: Enhanced OCR for missing Name/Age/Gender =====
    if current_phase < 3:
        log("Phase 3/4: Fixing missing Name/Age/Gender...")
        enhanced_done = set(checkpoint.data.get('enhanced_ocr_done', []))

        cards_to_fix = []
        for global_idx, (s_no, data, pdf_name) in ocr_results.items():
            if global_idx in enhanced_done:
                continue
            if data:
                need_name = not data.get('name')
                need_age = not data.get('age')
                need_gender = not data.get('gender')
                if need_name or need_age or need_gender:
                    if global_idx < len(all_cards):
                        jpg_path = all_cards[global_idx][0]
                        cards_to_fix.append((jpg_path, global_idx, need_name, need_age, need_gender))

        if cards_to_fix:
            log(f"  Fixing {len(cards_to_fix):,} cards with missing data...")
            fixed_count = 0
            ctx = multiprocessing.get_context('spawn')

            with ProcessPoolExecutor(max_workers=num_workers, mp_context=ctx) as executor:
                futures = {executor.submit(enhanced_ocr_name_age_gender, card): card[1] for card in cards_to_fix}

                for future in as_completed(futures):
                    try:
                        global_idx, enhanced_data = future.result()
                        checkpoint.data['enhanced_ocr_done'].append(global_idx)

                        if enhanced_data and global_idx in ocr_results:
                            s_no, old_data, pdf_name = ocr_results[global_idx]
                            if old_data:
                                if not old_data.get('name') and enhanced_data.get('name'):
                                    old_data['name'] = enhanced_data['name']
                                if not old_data.get('age') and enhanced_data.get('age'):
                                    old_data['age'] = enhanced_data['age']
                                if not old_data.get('gender') and enhanced_data.get('gender'):
                                    old_data['gender'] = enhanced_data['gender']
                                ocr_results[global_idx] = (s_no, old_data, pdf_name)
                                checkpoint.data['ocr_results'][str(global_idx)] = (s_no, old_data, pdf_name)

                        fixed_count += 1
                        if fixed_count % 50 == 0 or fixed_count == len(cards_to_fix):
                            log(f"  Fixed: {fixed_count}/{len(cards_to_fix)}")
                            checkpoint.save()  # Save progress every 50 cards
                    except Exception as e:
                        log(f"  Enhanced OCR error: {e}")
                        fixed_count += 1

            checkpoint.save()
            gc.collect()  # Clean up multiprocessing resources
        else:
            log("  No cards need fixing")

        checkpoint.data['phase'] = 3
        checkpoint.save()
        log("  Phase 3 complete")

    # ===== PHASE 4: Create Excel =====
    log("Phase 4/4: Creating Excel file...")
    ocr_results = {int(k): tuple(v) for k, v in checkpoint.data.get('ocr_results', {}).items()}

    wb = Workbook()
    ws = wb.active
    ws.title = "Voter Data"

    headers = ['S.No', 'Part No.', 'Voter ID', 'Name', 'Relation Type', 'Relation Name',
               'House No', 'Age', 'Gender', 'Constituency', 'Source Folder', 'Card File']

    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    row_num = 2
    missing_name_count = 0
    missing_age_count = 0
    missing_gender_count = 0

    for global_idx in sorted(ocr_results.keys()):
        s_no, data, pdf_name = ocr_results[global_idx]
        ws.cell(row=row_num, column=1, value=s_no)
        ws.cell(row=row_num, column=2, value=extract_part_number(pdf_name))

        if data:
            ws.cell(row=row_num, column=3, value=data.get('voter_id', ''))

            name_val = data.get('name', '')
            name_cell = ws.cell(row=row_num, column=4, value=name_val)
            if not name_val:
                name_cell.fill = yellow_fill
                missing_name_count += 1

            ws.cell(row=row_num, column=5, value=data.get('relation_type', ''))
            ws.cell(row=row_num, column=6, value=data.get('relation_name', ''))
            ws.cell(row=row_num, column=7, value=data.get('house_no', ''))

            age_val = data.get('age', '')
            gender_val = data.get('gender', '')

            age_cell = ws.cell(row=row_num, column=8, value=age_val)
            gender_cell = ws.cell(row=row_num, column=9, value=gender_val)

            if not age_val:
                age_cell.fill = yellow_fill
                missing_age_count += 1
            if not gender_val:
                gender_cell.fill = yellow_fill
                missing_gender_count += 1
        else:
            for col in range(3, 10):
                cell = ws.cell(row=row_num, column=col, value='')
                if col in [4, 8, 9]:  # Name, Age, Gender columns
                    cell.fill = yellow_fill
            missing_name_count += 1
            missing_age_count += 1
            missing_gender_count += 1

        ws.cell(row=row_num, column=10, value=constituency_name)
        ws.cell(row=row_num, column=11, value=pdf_name)
        ws.cell(row=row_num, column=12, value=f"{s_no}.png")
        row_num += 1

    # Column widths
    column_widths = [8, 10, 15, 25, 12, 25, 15, 8, 12, 30, 50, 10]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col) if col <= 26 else 'A' + chr(64 + col - 26)].width = width

    output_dir = folder_path.parent / "output"
    output_dir.mkdir(exist_ok=True)
    excel_path = output_dir / f"{constituency_name}_excel.xlsx"
    wb.save(excel_path)

    total_cards = len(ocr_results)
    elapsed_total = get_elapsed()
    max_missing = max(missing_name_count, missing_age_count, missing_gender_count)
    completeness = ((total_cards - max_missing) / total_cards * 100) if total_cards > 0 else 0

    log(f"  Excel saved: {excel_path}")
    log(f"  Total cards: {total_cards:,}")
    log(f"  Missing Name: {missing_name_count:,}")
    log(f"  Missing Age: {missing_age_count:,}")
    log(f"  Missing Gender: {missing_gender_count:,}")
    log(f"  Completeness: {completeness:.1f}%")
    log(f"  Total time: {format_time(elapsed_total)}")

    # Cleanup
    if cleanup:
        log("Cleaning up temp files...")
        import shutil
        try:
            shutil.rmtree(image_dir)
            log("  Temp images deleted")
        except Exception as e:
            log(f"  Cleanup error: {e}")

    checkpoint.delete()

    send_notification(
        "Processing Complete!",
        f"Constituency: {constituency_name}\nTotal: {total_cards:,}\n"
        f"Missing Name: {missing_name_count:,}\nMissing Age: {missing_age_count:,}\nMissing Gender: {missing_gender_count:,}\n"
        f"Completeness: {completeness:.1f}%\nTime: {format_time(elapsed_total)}"
    )

    log("=" * 50)
    log("PROCESSING COMPLETE!")
    log("=" * 50)
    return True


def main():
    parser = argparse.ArgumentParser(description='Voter Analytics - Headless Batch Processor v8.3')
    parser.add_argument('folder', help='Path to constituency folder containing PDFs')
    parser.add_argument('--ntfy-topic', help='Ntfy topic for notifications')
    parser.add_argument('--workers', type=int, default=NUM_WORKERS, help=f'Number of workers (default: {NUM_WORKERS})')
    parser.add_argument('--no-cleanup', action='store_true', help='Keep temp files after processing')

    args = parser.parse_args()

    log(f"Voter Analytics Headless Processor v8.3")
    log(f"Using spawn context with OMP_THREAD_LIMIT=1")
    log(f"Workers: {args.workers}")

    success = process_constituency(
        args.folder,
        ntfy_topic=args.ntfy_topic,
        num_workers=args.workers,
        cleanup=not args.no_cleanup
    )

    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
