"""
Batch Process Electoral Roll PDFs with Document AI

- Processes PDFs in small batches (5 at a time)
- Shows progress
- Can be stopped with Ctrl+C
- Saves results as Excel files

Setup:
1. Enable Document AI API and Cloud Storage API
2. Create a GCS bucket for input/output
3. Configure .env file with credentials
"""

import json
import re
import os
import time
import base64
import subprocess
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import secrets
import webbrowser
from http.server import HTTPServer, BaseHTTPRequestHandler
from urllib.parse import urlencode, parse_qs, urlparse
import urllib.request
import urllib.error

# Install packages
def install_packages():
    packages = ['openpyxl', 'python-dotenv', 'PyMuPDF']
    for pkg in packages:
        try:
            if pkg == 'python-dotenv':
                __import__('dotenv')
            elif pkg == 'PyMuPDF':
                __import__('fitz')
            else:
                __import__(pkg)
        except ImportError:
            print(f"Installing {pkg}...")
            subprocess.check_call(['uv', 'pip', 'install', pkg])

install_packages()

import fitz  # PyMuPDF for PDF splitting

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from dotenv import load_dotenv


# ============== OAuth ==============

class OAuthManager:
    SCOPES = ['https://www.googleapis.com/auth/cloud-platform']
    AUTH_URL = 'https://accounts.google.com/o/oauth2/v2/auth'
    TOKEN_URL = 'https://oauth2.googleapis.com/token'
    REDIRECT_PORT = 8089
    REDIRECT_URI = f'http://localhost:{REDIRECT_PORT}/callback'

    def __init__(self, client_id, client_secret):
        self.client_id = client_id
        self.client_secret = client_secret
        self.access_token = None
        self.refresh_token = None
        self.token_expiry = 0
        self.token_file = Path.home() / '.google_docai_oauth_token.json'
        self.load_saved_token()

    def load_saved_token(self):
        if self.token_file.exists():
            try:
                data = json.loads(self.token_file.read_text())
                self.refresh_token = data.get('refresh_token')
                self.access_token = data.get('access_token')
                self.token_expiry = data.get('expiry', 0)
            except:
                pass

    def save_token(self):
        try:
            data = {
                'refresh_token': self.refresh_token,
                'access_token': self.access_token,
                'expiry': self.token_expiry
            }
            self.token_file.write_text(json.dumps(data))
        except:
            pass

    def get_valid_token(self):
        if self.access_token and time.time() < self.token_expiry - 60:
            return self.access_token
        if self.refresh_token:
            if self.refresh_access_token():
                return self.access_token
        if self.do_oauth_flow():
            return self.access_token
        return None

    def refresh_access_token(self):
        try:
            data = urlencode({
                'client_id': self.client_id,
                'client_secret': self.client_secret,
                'refresh_token': self.refresh_token,
                'grant_type': 'refresh_token'
            }).encode()
            req = urllib.request.Request(self.TOKEN_URL, data=data)
            with urllib.request.urlopen(req, timeout=30) as response:
                result = json.loads(response.read().decode())
                self.access_token = result['access_token']
                self.token_expiry = time.time() + result.get('expires_in', 3600)
                self.save_token()
                return True
        except:
            return False

    def do_oauth_flow(self):
        auth_code = None
        state = secrets.token_urlsafe(16)

        class CallbackHandler(BaseHTTPRequestHandler):
            def do_GET(self):
                nonlocal auth_code
                parsed = urlparse(self.path)
                if parsed.path == '/callback':
                    params = parse_qs(parsed.query)
                    if params.get('state', [None])[0] == state:
                        auth_code = params.get('code', [None])[0]
                        self.send_response(200)
                        self.send_header('Content-type', 'text/html')
                        self.end_headers()
                        self.wfile.write(b'<html><body><h1>Success!</h1><p>You can close this window.</p></body></html>')
                    else:
                        self.send_response(400)
                        self.end_headers()
                else:
                    self.send_response(404)
                    self.end_headers()
            def log_message(self, format, *args):
                pass

        auth_params = {
            'client_id': self.client_id,
            'redirect_uri': self.REDIRECT_URI,
            'response_type': 'code',
            'scope': ' '.join(self.SCOPES),
            'state': state,
            'access_type': 'offline',
            'prompt': 'consent'
        }
        auth_url = f"{self.AUTH_URL}?{urlencode(auth_params)}"

        server = HTTPServer(('localhost', self.REDIRECT_PORT), CallbackHandler)
        server.timeout = 120

        webbrowser.open(auth_url)

        while auth_code is None:
            server.handle_request()

        server.server_close()

        if not auth_code:
            return False

        try:
            data = urlencode({
                'client_id': self.client_id,
                'client_secret': self.client_secret,
                'code': auth_code,
                'redirect_uri': self.REDIRECT_URI,
                'grant_type': 'authorization_code'
            }).encode()
            req = urllib.request.Request(self.TOKEN_URL, data=data)
            with urllib.request.urlopen(req, timeout=30) as response:
                result = json.loads(response.read().decode())
                self.access_token = result['access_token']
                self.refresh_token = result.get('refresh_token', self.refresh_token)
                self.token_expiry = time.time() + result.get('expires_in', 3600)
                self.save_token()
                return True
        except:
            return False


# ============== GCS Operations ==============

def upload_to_gcs(file_path, bucket_name, blob_name, token):
    """Upload a file to GCS."""
    url = f"https://storage.googleapis.com/upload/storage/v1/b/{bucket_name}/o?uploadType=media&name={blob_name}"

    with open(file_path, 'rb') as f:
        file_content = f.read()

    headers = {
        'Authorization': f'Bearer {token}',
        'Content-Type': 'application/pdf'
    }

    req = urllib.request.Request(url, data=file_content, headers=headers, method='POST')

    with urllib.request.urlopen(req, timeout=300) as response:
        return json.loads(response.read().decode())


def download_from_gcs(bucket_name, blob_name, token):
    """Download a file from GCS."""
    url = f"https://storage.googleapis.com/storage/v1/b/{bucket_name}/o/{urllib.parse.quote(blob_name, safe='')}?alt=media"

    headers = {'Authorization': f'Bearer {token}'}
    req = urllib.request.Request(url, headers=headers)

    with urllib.request.urlopen(req, timeout=300) as response:
        return response.read()


def list_gcs_objects(bucket_name, prefix, token):
    """List objects in GCS bucket with prefix."""
    url = f"https://storage.googleapis.com/storage/v1/b/{bucket_name}/o?prefix={urllib.parse.quote(prefix, safe='')}"

    headers = {'Authorization': f'Bearer {token}'}
    req = urllib.request.Request(url, headers=headers)

    with urllib.request.urlopen(req, timeout=60) as response:
        result = json.loads(response.read().decode())
        return result.get('items', [])


def delete_gcs_object(bucket_name, blob_name, token):
    """Delete an object from GCS."""
    url = f"https://storage.googleapis.com/storage/v1/b/{bucket_name}/o/{urllib.parse.quote(blob_name, safe='')}"

    headers = {'Authorization': f'Bearer {token}'}
    req = urllib.request.Request(url, headers=headers, method='DELETE')

    try:
        with urllib.request.urlopen(req, timeout=60) as response:
            return True
    except:
        return False


# ============== Document AI ==============

def start_batch_process(project_id, location, processor_id, input_gcs_uri, output_gcs_uri, token):
    """Start a Document AI batch processing job."""
    url = f"https://{location}-documentai.googleapis.com/v1/projects/{project_id}/locations/{location}/processors/{processor_id}:batchProcess"

    payload = {
        "inputDocuments": {
            "gcsDocuments": {
                "documents": [{"gcsUri": input_gcs_uri, "mimeType": "application/pdf"}]
            }
        },
        "documentOutputConfig": {
            "gcsOutputConfig": {
                "gcsUri": output_gcs_uri
            }
        }
    }

    headers = {
        'Authorization': f'Bearer {token}',
        'Content-Type': 'application/json'
    }

    req = urllib.request.Request(url, data=json.dumps(payload).encode(), headers=headers, method='POST')

    with urllib.request.urlopen(req, timeout=60) as response:
        result = json.loads(response.read().decode())
        return result.get('name')  # Operation name


def check_operation_status(operation_name, token):
    """Check status of a Document AI operation."""
    url = f"https://us-documentai.googleapis.com/v1/{operation_name}"

    headers = {'Authorization': f'Bearer {token}'}
    req = urllib.request.Request(url, headers=headers)

    with urllib.request.urlopen(req, timeout=60) as response:
        result = json.loads(response.read().decode())
        return result


def process_pdf_online(pdf_path, project_id, location, processor_id, token):
    """Process a single PDF using online (sync) API."""
    url = f"https://{location}-documentai.googleapis.com/v1/projects/{project_id}/locations/{location}/processors/{processor_id}:process"

    with open(pdf_path, 'rb') as f:
        pdf_content = base64.b64encode(f.read()).decode('utf-8')

    payload = {
        "rawDocument": {
            "content": pdf_content,
            "mimeType": "application/pdf"
        }
    }

    headers = {
        'Authorization': f'Bearer {token}',
        'Content-Type': 'application/json'
    }

    req = urllib.request.Request(url, data=json.dumps(payload).encode(), headers=headers, method='POST')

    with urllib.request.urlopen(req, timeout=300) as response:
        result = json.loads(response.read().decode())
        return result.get('document', {})


# ============== JSON to Excel ==============

def extract_part_number(filename):
    match = re.search(r'-TAM-(\d+)-WI', filename, re.IGNORECASE)
    return match.group(1) if match else ''


def get_entity_position(entity):
    page_anchor = entity.get('pageAnchor', {})
    page_refs = page_anchor.get('pageRefs', [{}])
    if not page_refs:
        return 0, 0, 0
    page_ref = page_refs[0]
    page = int(page_ref.get('page', 0))
    bounding_poly = page_ref.get('boundingPoly', {})
    vertices = bounding_poly.get('normalizedVertices', [{}])
    if not vertices:
        return page, 0, 0
    x = vertices[0].get('x', 0)
    y = vertices[0].get('y', 0)
    return page, x, y


def group_entities_to_cards(entities):
    """Group entities into voter cards using VoterID as anchor."""
    for entity in entities:
        page, x, y = get_entity_position(entity)
        entity['_page'] = page
        entity['_x'] = x
        entity['_y'] = y
        if x < 0.33:
            entity['_col'] = 0
        elif x < 0.66:
            entity['_col'] = 1
        else:
            entity['_col'] = 2

    voter_ids = [e for e in entities if e.get('type', '').lower() == 'voterid']

    cards = []
    for vid in voter_ids:
        card = {
            'serial_no': '',
            'voter_id': vid.get('mentionText', '').strip().upper(),
            'name': '',
            'relation_name': '',
            'house_no': '',
            'age': '',
            'gender': '',
            'page': vid['_page'],
            'col': vid['_col'],
            '_y_start': vid['_y'],
            '_y_end': vid['_y'] + 0.10
        }
        cards.append(card)

    cards.sort(key=lambda c: (c['page'], c['col'], c['_y_start']))

    for i, card in enumerate(cards):
        for j in range(i + 1, len(cards)):
            if cards[j]['page'] == card['page'] and cards[j]['col'] == card['col']:
                card['_y_end'] = cards[j]['_y_start']
                break

    for entity in entities:
        if entity.get('type', '').lower() == 'voterid':
            continue

        page = entity['_page']
        col = entity['_col']
        y = entity['_y']
        entity_type = entity.get('type', '').lower()
        mention_text = entity.get('mentionText', '').strip()
        mention_text = re.sub(r'\s*[-–]\s*$', '', mention_text)
        mention_text = re.sub(r'^\s*[-–]\s*', '', mention_text)

        best_card = None
        for card in cards:
            if card['page'] == page and card['col'] == col:
                if card['_y_start'] <= y < card['_y_end']:
                    best_card = card
                    break
                elif abs(y - card['_y_start']) < 0.12:
                    best_card = card

        if not best_card:
            candidates = [c for c in cards if c['page'] == page and c['col'] == col]
            if candidates:
                best_card = min(candidates, key=lambda c: abs(c['_y_start'] - y))

        if best_card:
            if entity_type == 'sno':
                best_card['serial_no'] = mention_text
            elif entity_type == 'name':
                best_card['name'] = mention_text
            elif entity_type == 'relativename':
                best_card['relation_name'] = mention_text
            elif entity_type == 'houseno':
                best_card['house_no'] = mention_text
            elif entity_type == 'age':
                age_match = re.search(r'(\d+)', mention_text)
                if age_match:
                    age = int(age_match.group(1))
                    if 18 <= age <= 120:
                        best_card['age'] = str(age)
            elif entity_type == 'sex':
                if 'பெண்' in mention_text or 'பெண' in mention_text:
                    best_card['gender'] = 'Female'
                elif 'ஆண்' in mention_text or 'ஆண' in mention_text:
                    best_card['gender'] = 'Male'

    return cards


def save_cards_to_excel(cards, output_path, source_name):
    """Save cards to Excel."""
    part_no = extract_part_number(source_name)

    wb = Workbook()
    ws = wb.active
    ws.title = "Voter Data"

    headers = ['S.No', 'Part No.', 'Voter S.No', 'Voter ID', 'Name',
               'Relation Name', 'House No', 'Age', 'Gender', 'Page']

    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    missing = {'name': 0, 'age': 0, 'gender': 0, 'voter_id': 0}

    for row_num, card in enumerate(cards, 2):
        ws.cell(row=row_num, column=1, value=row_num - 1)
        ws.cell(row=row_num, column=2, value=part_no)
        ws.cell(row=row_num, column=3, value=card.get('serial_no', ''))

        vid = card.get('voter_id', '')
        vc = ws.cell(row=row_num, column=4, value=vid)
        if not vid:
            vc.fill = yellow_fill
            missing['voter_id'] += 1

        name = card.get('name', '')
        nc = ws.cell(row=row_num, column=5, value=name)
        if not name:
            nc.fill = yellow_fill
            missing['name'] += 1

        ws.cell(row=row_num, column=6, value=card.get('relation_name', ''))
        ws.cell(row=row_num, column=7, value=card.get('house_no', ''))

        age = card.get('age', '')
        ac = ws.cell(row=row_num, column=8, value=age)
        if not age:
            ac.fill = yellow_fill
            missing['age'] += 1

        gender = card.get('gender', '')
        gc = ws.cell(row=row_num, column=9, value=gender)
        if not gender:
            gc.fill = yellow_fill
            missing['gender'] += 1

        ws.cell(row=row_num, column=10, value=card.get('page', ''))

    widths = [8, 10, 10, 15, 25, 25, 12, 8, 10, 8]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width

    wb.save(output_path)
    return missing


# ============== Main App ==============

class BatchProcessor:
    def __init__(self, root):
        self.root = root
        self.root.title("Document AI Batch Processor")
        self.root.geometry("900x700")

        self.stop_requested = False
        self.processing = False

        # Load .env
        env_file = Path(__file__).parent / '.env'
        if env_file.exists():
            load_dotenv(env_file)

        self.client_id = os.getenv('GOOGLE_CLIENT_ID', '')
        self.client_secret = os.getenv('GOOGLE_CLIENT_SECRET', '')
        self.project_id = os.getenv('GOOGLE_PROJECT_ID', '')
        self.processor_id = os.getenv('DOCAI_PROCESSOR_ID', '')
        self.location = os.getenv('DOCAI_LOCATION', 'us')
        self.bucket_name = os.getenv('GCS_BUCKET', '')

        self.oauth_manager = None

        self.create_widgets()
        self.check_credentials()

    def create_widgets(self):
        main_frame = ttk.Frame(self.root, padding="15")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Title
        ttk.Label(main_frame, text="Document AI Batch Processor",
                  font=('Helvetica', 16, 'bold')).pack(pady=(0, 5))
        ttk.Label(main_frame, text="Processes PDFs one by one using online API ($0.10/page)",
                  foreground='gray').pack(pady=(0, 10))

        # Credentials
        cred_frame = ttk.LabelFrame(main_frame, text="Credentials", padding="10")
        cred_frame.pack(fill=tk.X, pady=(0, 10))

        self.cred_status = tk.StringVar(value="Checking...")
        ttk.Label(cred_frame, textvariable=self.cred_status).pack(side=tk.LEFT)
        ttk.Button(cred_frame, text="Authorize", command=self.authorize).pack(side=tk.RIGHT, padx=5)
        ttk.Button(cred_frame, text="Edit .env", command=self.edit_env).pack(side=tk.RIGHT, padx=5)

        # Settings
        settings_frame = ttk.LabelFrame(main_frame, text="Document AI Settings", padding="10")
        settings_frame.pack(fill=tk.X, pady=(0, 10))

        row1 = ttk.Frame(settings_frame)
        row1.pack(fill=tk.X, pady=2)
        ttk.Label(row1, text="Project ID:", width=12).pack(side=tk.LEFT)
        self.project_var = tk.StringVar(value=self.project_id)
        ttk.Entry(row1, textvariable=self.project_var, width=40).pack(side=tk.LEFT, fill=tk.X, expand=True)

        row2 = ttk.Frame(settings_frame)
        row2.pack(fill=tk.X, pady=2)
        ttk.Label(row2, text="Processor ID:", width=12).pack(side=tk.LEFT)
        self.processor_var = tk.StringVar(value=self.processor_id)
        ttk.Entry(row2, textvariable=self.processor_var, width=40).pack(side=tk.LEFT, fill=tk.X, expand=True)

        row3 = ttk.Frame(settings_frame)
        row3.pack(fill=tk.X, pady=2)
        ttk.Label(row3, text="Location:", width=12).pack(side=tk.LEFT)
        self.location_var = tk.StringVar(value=self.location)
        ttk.Combobox(row3, textvariable=self.location_var, values=['us', 'eu'], width=10).pack(side=tk.LEFT)

        # Folder selection
        folder_frame = ttk.LabelFrame(main_frame, text="PDF Folder", padding="10")
        folder_frame.pack(fill=tk.X, pady=(0, 10))

        self.folder_var = tk.StringVar()
        ttk.Entry(folder_frame, textvariable=self.folder_var, width=60).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        ttk.Button(folder_frame, text="Browse...", command=self.browse_folder).pack(side=tk.LEFT)

        # Info
        info_frame = ttk.Frame(main_frame)
        info_frame.pack(fill=tk.X, pady=(0, 10))

        self.pdf_count_var = tk.StringVar(value="PDFs: --")
        ttk.Label(info_frame, textvariable=self.pdf_count_var).pack(side=tk.LEFT)

        self.cost_var = tk.StringVar(value="Est. Cost: --")
        ttk.Label(info_frame, textvariable=self.cost_var, foreground='#E91E63',
                  font=('Helvetica', 10, 'bold')).pack(side=tk.RIGHT)

        # Buttons
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(pady=(0, 10))

        self.start_btn = ttk.Button(btn_frame, text="Start Processing", command=self.start_processing)
        self.start_btn.pack(side=tk.LEFT, padx=5)

        self.stop_btn = ttk.Button(btn_frame, text="Stop", command=self.stop_processing, state=tk.DISABLED)
        self.stop_btn.pack(side=tk.LEFT, padx=5)

        # Progress
        progress_frame = ttk.LabelFrame(main_frame, text="Progress", padding="10")
        progress_frame.pack(fill=tk.X, pady=(0, 10))

        self.status_var = tk.StringVar(value="Ready")
        ttk.Label(progress_frame, textvariable=self.status_var).pack(anchor=tk.W)

        self.progress = ttk.Progressbar(progress_frame, mode='determinate')
        self.progress.pack(fill=tk.X, pady=5)

        self.detail_var = tk.StringVar(value="")
        ttk.Label(progress_frame, textvariable=self.detail_var).pack(anchor=tk.W)

        # Stats
        stats_frame = ttk.LabelFrame(main_frame, text="Statistics", padding="10")
        stats_frame.pack(fill=tk.X, pady=(0, 10))

        stats_grid = ttk.Frame(stats_frame)
        stats_grid.pack()

        ttk.Label(stats_grid, text="Processed:", font=('Helvetica', 10, 'bold')).grid(row=0, column=0, padx=10)
        self.processed_var = tk.StringVar(value="0 / 0")
        ttk.Label(stats_grid, textvariable=self.processed_var, font=('Helvetica', 14, 'bold'),
                  foreground='#2196F3').grid(row=0, column=1, padx=10)

        ttk.Label(stats_grid, text="Cards:", font=('Helvetica', 10, 'bold')).grid(row=0, column=2, padx=10)
        self.cards_var = tk.StringVar(value="0")
        ttk.Label(stats_grid, textvariable=self.cards_var, font=('Helvetica', 14, 'bold'),
                  foreground='#E91E63').grid(row=0, column=3, padx=10)

        ttk.Label(stats_grid, text="Time:", font=('Helvetica', 10, 'bold')).grid(row=0, column=4, padx=10)
        self.time_var = tk.StringVar(value="--")
        ttk.Label(stats_grid, textvariable=self.time_var, font=('Helvetica', 14, 'bold'),
                  foreground='#4CAF50').grid(row=0, column=5, padx=10)

        # Log
        log_frame = ttk.LabelFrame(main_frame, text="Log", padding="5")
        log_frame.pack(fill=tk.BOTH, expand=True)

        self.log_text = tk.Text(log_frame, height=10, font=('Courier', 9))
        log_scroll = ttk.Scrollbar(log_frame, orient=tk.VERTICAL, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=log_scroll.set)
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        log_scroll.pack(side=tk.RIGHT, fill=tk.Y)

    def log(self, msg):
        self.log_text.insert(tk.END, f"{time.strftime('%H:%M:%S')} - {msg}\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()

    def check_credentials(self):
        if self.client_id and self.client_secret:
            self.oauth_manager = OAuthManager(self.client_id, self.client_secret)
            if self.oauth_manager.refresh_token:
                self.cred_status.set(f"Authorized: {self.client_id[:20]}...")
            else:
                self.cred_status.set(f"Client ID: {self.client_id[:20]}... (need auth)")
        else:
            self.cred_status.set("Not configured - Edit .env file")

    def edit_env(self):
        env_file = Path(__file__).parent / '.env'
        if not env_file.exists():
            env_file.write_text(
                "GOOGLE_CLIENT_ID=your_client_id.apps.googleusercontent.com\n"
                "GOOGLE_CLIENT_SECRET=your_secret\n"
                "GOOGLE_PROJECT_ID=your_project\n"
                "DOCAI_PROCESSOR_ID=your_processor_id\n"
                "DOCAI_LOCATION=us\n"
                "GCS_BUCKET=your_bucket_name\n"
            )
        os.startfile(str(env_file)) if os.name == 'nt' else subprocess.run(['xdg-open', str(env_file)])

    def authorize(self):
        env_file = Path(__file__).parent / '.env'
        if env_file.exists():
            load_dotenv(env_file, override=True)
            self.client_id = os.getenv('GOOGLE_CLIENT_ID', '')
            self.client_secret = os.getenv('GOOGLE_CLIENT_SECRET', '')

        if not self.client_id or not self.client_secret:
            messagebox.showerror("Error", "Configure credentials in .env first")
            return

        self.oauth_manager = OAuthManager(self.client_id, self.client_secret)
        self.log("Starting authorization...")

        def do_auth():
            token = self.oauth_manager.get_valid_token()
            if token:
                self.root.after(0, lambda: self.cred_status.set(f"Authorized: {self.client_id[:20]}..."))
                self.root.after(0, lambda: self.log("Authorization successful!"))
            else:
                self.root.after(0, lambda: self.log("Authorization failed"))

        threading.Thread(target=do_auth, daemon=True).start()

    def browse_folder(self):
        folder = filedialog.askdirectory(title="Select Folder with PDFs")
        if folder:
            self.folder_var.set(folder)
            self.count_pdfs()

    def count_pdfs(self):
        folder = Path(self.folder_var.get())
        if not folder.exists():
            return

        pdfs = list(folder.glob('*.pdf')) + list(folder.glob('*.PDF'))
        count = len(pdfs)

        # Estimate pages (rough: ~25 pages per PDF)
        est_pages = count * 25
        cost = est_pages * 0.10

        self.pdf_count_var.set(f"PDFs: {count}")
        self.cost_var.set(f"Est. Cost: ${cost:.2f} (assuming ~25 pages/PDF)")
        self.log(f"Found {count} PDF files")

    def stop_processing(self):
        self.stop_requested = True
        self.stop_btn.config(state=tk.DISABLED)
        self.status_var.set("Stopping after current PDF...")
        self.log("Stop requested - will finish current PDF")

    def start_processing(self):
        folder = self.folder_var.get()
        if not folder or not Path(folder).exists():
            messagebox.showerror("Error", "Select a valid folder")
            return

        if not self.oauth_manager:
            messagebox.showerror("Error", "Please authorize first")
            return

        token = self.oauth_manager.get_valid_token()
        if not token:
            messagebox.showerror("Error", "Failed to get token - re-authorize")
            return

        project_id = self.project_var.get()
        processor_id = self.processor_var.get()

        if not project_id or not processor_id:
            messagebox.showerror("Error", "Enter Project ID and Processor ID")
            return

        self.stop_requested = False
        self.processing = True
        self.start_btn.config(state=tk.DISABLED)
        self.stop_btn.config(state=tk.NORMAL)

        threading.Thread(target=self.process_folder, args=(folder, project_id, processor_id), daemon=True).start()

    def process_folder(self, folder, project_id, processor_id):
        """Process all PDFs in folder one by one."""
        folder = Path(folder)
        location = self.location_var.get()

        pdfs = sorted(list(folder.glob('*.pdf')) + list(folder.glob('*.PDF')))
        total = len(pdfs)

        if total == 0:
            self.root.after(0, lambda: messagebox.showerror("Error", "No PDF files found"))
            self.root.after(0, self.reset_ui)
            return

        self.root.after(0, lambda: self.log(f"Starting processing of {total} PDFs"))

        # Create output folder
        output_folder = folder.parent / f"{folder.name}_excel"
        output_folder.mkdir(exist_ok=True)

        start_time = time.time()
        total_cards = 0
        processed = 0
        errors = []

        for i, pdf_path in enumerate(pdfs):
            if self.stop_requested:
                self.root.after(0, lambda: self.log("Stopped by user"))
                break

            pdf_name = pdf_path.stem
            self.root.after(0, lambda n=pdf_name, idx=i+1, t=total: (
                self.status_var.set(f"Processing: {n}"),
                self.detail_var.set(f"File {idx}/{t}"),
                self.progress.config(value=(idx/t)*100)
            ))
            self.root.after(0, lambda n=pdf_name: self.log(f"Processing: {n}"))

            try:
                # Get fresh token
                token = self.oauth_manager.get_valid_token()
                if not token:
                    raise Exception("Failed to get token")

                # Process PDF online
                document = process_pdf_online(pdf_path, project_id, location, processor_id, token)

                if not document:
                    raise Exception("No response from API")

                # Parse entities
                entities = document.get('entities', [])
                if not entities:
                    self.root.after(0, lambda n=pdf_name: self.log(f"  WARNING: No entities found in {n}"))
                    continue

                # Group into cards
                cards = group_entities_to_cards(entities)

                # Save to Excel
                excel_path = output_folder / f"{pdf_name}_excel.xlsx"
                save_cards_to_excel(cards, excel_path, pdf_name)

                total_cards += len(cards)
                processed += 1

                self.root.after(0, lambda c=len(cards), tc=total_cards: (
                    self.cards_var.set(f"{tc:,}"),
                    self.log(f"  Extracted {c} cards")
                ))
                self.root.after(0, lambda p=processed, t=total: self.processed_var.set(f"{p} / {t}"))

            except Exception as e:
                error_msg = str(e)
                errors.append((pdf_name, error_msg))
                self.root.after(0, lambda n=pdf_name, e=error_msg: self.log(f"  ERROR: {n} - {e}"))

            # Update time
            elapsed = time.time() - start_time
            self.root.after(0, lambda e=elapsed: self.time_var.set(f"{int(e//60)}m {int(e%60)}s"))

        # Complete
        elapsed = time.time() - start_time
        elapsed_str = f"{int(elapsed//60)}m {int(elapsed%60)}s"

        self.root.after(0, lambda: self.status_var.set("Complete!"))
        self.root.after(0, lambda: self.progress.config(value=100))
        self.root.after(0, self.reset_ui)

        self.root.after(0, lambda: self.log(f"\n{'='*40}"))
        self.root.after(0, lambda: self.log(f"COMPLETE!"))
        self.root.after(0, lambda p=processed, t=total: self.log(f"Processed: {p}/{t} PDFs"))
        self.root.after(0, lambda c=total_cards: self.log(f"Total cards: {c:,}"))
        self.root.after(0, lambda e=elapsed_str: self.log(f"Time: {e}"))
        self.root.after(0, lambda o=str(output_folder): self.log(f"Output: {o}"))

        if errors:
            self.root.after(0, lambda: self.log(f"\nErrors ({len(errors)}):"))
            for name, err in errors[:10]:
                self.root.after(0, lambda n=name, e=err: self.log(f"  - {n}: {e}"))

        self.root.after(0, lambda: messagebox.showinfo("Complete",
            f"Processing complete!\n\n"
            f"Processed: {processed}/{total} PDFs\n"
            f"Total cards: {total_cards:,}\n"
            f"Time: {elapsed_str}\n"
            f"Errors: {len(errors)}\n\n"
            f"Output: {output_folder}"))

    def reset_ui(self):
        self.start_btn.config(state=tk.NORMAL)
        self.stop_btn.config(state=tk.DISABLED)
        self.processing = False


def main():
    root = tk.Tk()
    app = BatchProcessor(root)
    root.mainloop()


if __name__ == "__main__":
    main()
