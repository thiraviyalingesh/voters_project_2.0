"""
Electoral Roll Voter Counter - GUI Application (FAST VERSION)
Extracts voter counts from Tamil Nadu Electoral Roll PDFs
Optimized with parallel processing for faster OCR
Supports batch processing of entire constituency folders
Features checkpoint/resume capability for interrupted sessions
"""

import re
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from pathlib import Path
import threading
import subprocess
import multiprocessing
from concurrent.futures import ProcessPoolExecutor, as_completed
import time
import json

# Install required packages
def install_packages():
    packages = ['pymupdf', 'pytesseract', 'pillow', 'openpyxl']
    for pkg in packages:
        try:
            __import__(pkg.replace('-', '_'))
        except ImportError:
            print(f"Installing {pkg}...")
            subprocess.check_call(['uv', 'pip', 'install', pkg])

install_packages()

import fitz
from PIL import Image, ImageEnhance
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

try:
    import pytesseract
except ImportError:
    subprocess.check_call(['uv', 'pip', 'install', 'pytesseract'])
    import pytesseract


# Worker functions for parallel processing (must be at module level for pickling)
def ocr_single_card(args):
    """Worker function to OCR a single voter card image."""
    jpg_path, global_idx, pdf_name = args
    try:
        img = Image.open(jpg_path)
        text = pytesseract.image_to_string(img, lang='tam+eng')
        data = parse_voter_card_standalone(text)
        return global_idx, jpg_path.stem, data, pdf_name
    except Exception as e:
        print(f"OCR error for {jpg_path}: {e}")
        return global_idx, jpg_path.stem if hasattr(jpg_path, 'stem') else str(global_idx), None, pdf_name


def enhanced_ocr_single_card(args):
    """Worker function for enhanced OCR with early stopping."""
    jpg_path, global_idx, missing_fields = args
    try:
        img = Image.open(jpg_path)

        # Try preprocessing approaches in order, stop when all fields found
        approaches = [
            ('contrast', lambda i: ImageEnhance.Contrast(i).enhance(2.0)),
            ('grayscale_sharp', lambda i: ImageEnhance.Sharpness(i.convert('L')).enhance(2.0)),
            ('binarize', lambda i: i.convert('L').point(lambda x: 0 if x < 140 else 255, '1')),
            ('scale_2x', lambda i: i.resize((i.size[0] * 2, i.size[1] * 2), Image.LANCZOS)),
        ]

        merged = {
            'serial_no': '',
            'voter_id': '',
            'name': '',
            'relation_name': '',
            'relation_type': '',
            'house_no': '',
            'age': '',
            'gender': ''
        }

        for name, transform in approaches:
            try:
                processed_img = transform(img)
                text = pytesseract.image_to_string(processed_img, lang='tam+eng')
                result = parse_voter_card_standalone(text)

                if result:
                    # Merge non-empty values
                    for key in merged:
                        if not merged[key] and result.get(key):
                            merged[key] = result[key]

                # Check if all missing fields are now filled
                all_found = all(merged.get(f) for f in missing_fields)
                if all_found:
                    break
            except:
                continue

        return global_idx, merged
    except Exception as e:
        print(f"Enhanced OCR error for {jpg_path}: {e}")
        return global_idx, None


def parse_voter_card_standalone(text):
    """Parse OCR text from voter card to extract structured data (standalone function)."""
    data = {
        'serial_no': '',
        'voter_id': '',
        'name': '',
        'relation_name': '',
        'relation_type': '',
        'house_no': '',
        'age': '',
        'gender': ''
    }

    full_text = text

    # Extract Voter ID
    voter_id_patterns = [
        r'\b([A-Z]{2,3}\d{6,10})\b',
        r'\b([A-Z0-9]{2,3}\d{6,10})\b',
        r'\b(\d{2}[^\d\s]{1,2}\d{6,10})\b',
        r'(\d{1,3}\s+[A-Z0-9]{2,3}\d{6,10})',
    ]

    for pattern in voter_id_patterns:
        matches = re.findall(pattern, full_text)
        for match in matches:
            id_match = re.search(r'([A-Z0-9]{2,3}\d{6,10})$|(\d{2}[^\d\s]{1,2}\d{6,10})$', match)
            if id_match:
                voter_id = id_match.group(1) or id_match.group(2)
                if voter_id and len(voter_id) >= 9:
                    data['voter_id'] = voter_id
                    break
        if data['voter_id']:
            break

    lines = full_text.split('\n')

    for line in lines:
        line = line.strip()
        if not line:
            continue

        # Extract serial number
        if not data['serial_no']:
            serial_match = re.match(r'^(\d{1,4})\s*$', line)
            if serial_match:
                data['serial_no'] = serial_match.group(1)
            else:
                serial_match = re.match(r'^(\d{1,4})\s+\S', line)
                if serial_match:
                    num = serial_match.group(1)
                    if int(num) < 2000:
                        data['serial_no'] = num

        # Extract name
        if 'பெயர்' in line and ':' in line:
            if 'தந்தை' not in line and 'கணவர்' not in line:
                name_part = line.split(':', 1)[-1]
                name_part = clean_ocr_text_standalone(name_part)
                if name_part and not data['name']:
                    data['name'] = name_part

        # Extract father's name (handles both தந்தை பெயர் and தந்தையின் பெயர்)
        if ('தந்தை' in line or 'தந்தையின்' in line) and 'பெயர்' in line and ':' in line:
            rel_part = line.split(':', 1)[-1]
            rel_part = clean_ocr_text_standalone(rel_part)
            if rel_part and not data['relation_name']:
                data['relation_name'] = rel_part
                data['relation_type'] = 'Father'

        # Extract husband's name (handles both கணவர் பெயர் and கணவரின் பெயர்)
        if ('கணவர்' in line or 'கணவரின்' in line) and ':' in line:
            rel_part = line.split(':', 1)[-1]
            rel_part = clean_ocr_text_standalone(rel_part)
            if rel_part and not data['relation_name']:
                data['relation_name'] = rel_part
                data['relation_type'] = 'Husband'

        # Extract mother's name (handles both தாய் பெயர் and தாயின் பெயர்)
        if ('தாய்' in line or 'தாயின்' in line) and 'பெயர்' in line and ':' in line:
            rel_part = line.split(':', 1)[-1]
            rel_part = clean_ocr_text_standalone(rel_part)
            if rel_part and not data['relation_name']:
                data['relation_name'] = rel_part
                data['relation_type'] = 'Mother'

        # Extract other's name (இதரர் பெயர் / இதரரின் பெயர்)
        if ('இதரர்' in line or 'இதரரின்' in line) and 'பெயர்' in line and ':' in line:
            rel_part = line.split(':', 1)[-1]
            rel_part = clean_ocr_text_standalone(rel_part)
            if rel_part and not data['relation_name']:
                data['relation_name'] = rel_part
                data['relation_type'] = 'Other'

        # Extract house number
        if ('வீட்டு' in line or 'ட்டு' in line) and 'எண்' in line and ':' in line:
            house_part = line.split(':', 1)[-1]
            house_part = clean_ocr_text_standalone(house_part)
            if house_part and not data['house_no']:
                data['house_no'] = house_part

        # Extract age
        if 'வயது' in line and ':' in line:
            age_match = re.search(r'வயது\s*:\s*(\d+)', line)
            if age_match:
                data['age'] = age_match.group(1)

        # Extract gender
        if 'பாலினம்' in line:
            if 'ஆண்' in line:
                data['gender'] = 'Male'
            elif 'பெண்' in line:
                data['gender'] = 'Female'

    return data


def clean_ocr_text_standalone(text):
    """Clean common OCR artifacts from text (standalone function)."""
    if not text:
        return ''
    text = re.sub(r'\s*Photo\s*is\s*', ' ', text, flags=re.IGNORECASE)
    text = re.sub(r'\s*available\s*', ' ', text, flags=re.IGNORECASE)
    text = re.sub(r'^[\s\-–.,:]+|[\s\-–.,:]+$', '', text)
    text = re.sub(r'\s+', ' ', text)
    return text.strip()


def extract_cards_from_single_pdf(args):
    """Worker function to extract cards from a single PDF."""
    pdf_path, temp_base_dir, pdf_index = args
    try:
        pdf_path = Path(pdf_path)
        pdf_name = pdf_path.stem
        output_path = Path(temp_base_dir) / pdf_name
        output_path.mkdir(parents=True, exist_ok=True)

        doc = fitz.open(str(pdf_path))
        num_pages = len(doc)
        card_count = 0

        start_page = 3
        end_page = num_pages - 1

        for page_num in range(start_page, end_page):
            page = doc[page_num]

            zoom = 2
            mat = fitz.Matrix(zoom, zoom)
            pix = page.get_pixmap(matrix=mat)
            img_data = pix.tobytes("png")
            page_img = Image.open(io.BytesIO(img_data))

            page_width, page_height = page_img.size

            num_cols = 3
            num_rows = 10

            header_height = int(page_height * 0.035)
            footer_height = int(page_height * 0.025)
            content_height = page_height - header_height - footer_height

            card_width = page_width // num_cols
            row_height = content_height // num_rows

            for row in range(num_rows):
                for col in range(num_cols):
                    x1 = col * card_width
                    y1 = header_height + row * row_height
                    x2 = x1 + card_width
                    y2 = y1 + row_height

                    padding = 1
                    x1 = max(0, x1 + padding)
                    y1 = max(0, y1 + padding)
                    x2 = min(page_width, x2 - padding)
                    y2 = min(page_height, y2 - padding)

                    card_img = page_img.crop((x1, y1, x2, y2))

                    card_array = list(card_img.getdata())
                    if card_array:
                        avg_brightness = sum(sum(p[:3]) / 3 for p in card_array) / len(card_array)
                        if avg_brightness > 252:
                            continue

                    card_count += 1
                    card_filename = output_path / f"{card_count}.jpg"
                    card_img.save(card_filename, "JPEG", quality=95)

        doc.close()
        return pdf_index, pdf_name, card_count, str(output_path)
    except Exception as e:
        print(f"Error extracting from {pdf_path}: {e}")
        return pdf_index, Path(pdf_path).stem, 0, None


class CheckpointManager:
    """Manages saving and loading checkpoint data for resume capability."""

    def __init__(self, checkpoint_path):
        self.checkpoint_path = Path(checkpoint_path)
        self.data = {
            'phase': 0,
            'constituency_name': '',
            'folder_path': '',
            'total_pdfs': 0,
            'extracted_pdfs': {},  # pdf_name -> (card_count, output_path)
            'ocr_results': {},  # global_idx -> (s_no, data_dict, pdf_name)
            'enhanced_ocr_done': [],  # list of global_idx that have been enhanced
            'all_cards': [],  # list of (jpg_path_str, global_idx, pdf_name)
            'start_time': 0,
            'elapsed_before_resume': 0
        }

    def exists(self):
        return self.checkpoint_path.exists()

    def load(self):
        if self.exists():
            with open(self.checkpoint_path, 'r', encoding='utf-8') as f:
                self.data = json.load(f)
            print(f"Loaded checkpoint: Phase {self.data['phase']}")
            return True
        return False

    def save(self):
        with open(self.checkpoint_path, 'w', encoding='utf-8') as f:
            json.dump(self.data, f, ensure_ascii=False, indent=2)
        print(f"Checkpoint saved: Phase {self.data['phase']}")

    def delete(self):
        if self.exists():
            self.checkpoint_path.unlink()
            print("Checkpoint deleted")

    def update_phase(self, phase):
        self.data['phase'] = phase
        self.save()

    def add_extracted_pdf(self, pdf_name, card_count, output_path):
        self.data['extracted_pdfs'][pdf_name] = (card_count, output_path)

    def add_ocr_result(self, global_idx, s_no, data_dict, pdf_name):
        self.data['ocr_results'][str(global_idx)] = (s_no, data_dict, pdf_name)

    def add_enhanced_ocr(self, global_idx):
        if global_idx not in self.data['enhanced_ocr_done']:
            self.data['enhanced_ocr_done'].append(global_idx)

    def set_all_cards(self, all_cards):
        # Convert Path objects to strings for JSON serialization
        self.data['all_cards'] = [(str(p), idx, name) for p, idx, name in all_cards]

    def get_all_cards(self):
        # Convert strings back to Path objects
        return [(Path(p), idx, name) for p, idx, name in self.data['all_cards']]

    def get_ocr_results(self):
        # Convert string keys back to int
        return {int(k): tuple(v) for k, v in self.data['ocr_results'].items()}


class VoterCounterApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Electoral Roll Voter Counter (FAST - Batch)")
        self.root.geometry("750x650")
        self.root.resizable(True, True)

        # Get CPU count for parallel processing (limit to 4 to avoid memory issues)
        self.num_workers = min(4, max(1, multiprocessing.cpu_count() - 1))
        print(f"Using {self.num_workers} worker processes for parallel processing")

        self.checkpoint = None
        self.stop_requested = False

        self.style = ttk.Style()
        self.style.configure('Title.TLabel', font=('Helvetica', 16, 'bold'))
        self.style.configure('Header.TLabel', font=('Helvetica', 12, 'bold'))
        self.style.configure('Big.TLabel', font=('Helvetica', 24, 'bold'))

        self.create_widgets()

    def create_widgets(self):
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)

        title_label = ttk.Label(
            main_frame,
            text="Tamil Nadu Electoral Roll\nVoter Counter (Batch Mode)",
            style='Title.TLabel',
            justify=tk.CENTER
        )
        title_label.pack(pady=(0, 20))

        # Folder selection frame
        folder_frame = ttk.LabelFrame(main_frame, text="Select Constituency Folder", padding="10")
        folder_frame.pack(fill=tk.X, pady=(0, 10))

        self.folder_path_var = tk.StringVar()
        self.folder_entry = ttk.Entry(folder_frame, textvariable=self.folder_path_var, width=60)
        self.folder_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))

        browse_btn = ttk.Button(folder_frame, text="Browse...", command=self.browse_folder)
        browse_btn.pack(side=tk.LEFT)

        # Info frame
        info_frame = ttk.Frame(main_frame)
        info_frame.pack(fill=tk.X, pady=(0, 10))

        self.pdf_count_var = tk.StringVar(value="PDFs found: --")
        ttk.Label(info_frame, textvariable=self.pdf_count_var).pack(side=tk.LEFT)

        self.constituency_var = tk.StringVar(value="Constituency: --")
        ttk.Label(info_frame, textvariable=self.constituency_var).pack(side=tk.RIGHT)

        # Button frame
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(pady=(0, 10))

        self.extract_btn = ttk.Button(btn_frame, text="Extract All Voter Cards", command=self.extract_all_voter_cards)
        self.extract_btn.pack(side=tk.LEFT, padx=(0, 10))

        self.stop_btn = ttk.Button(btn_frame, text="Stop & Save", command=self.stop_and_save, state=tk.DISABLED)
        self.stop_btn.pack(side=tk.LEFT)

        # Progress frame
        progress_frame = ttk.LabelFrame(main_frame, text="Progress", padding="10")
        progress_frame.pack(fill=tk.X, pady=(0, 10))

        self.phase_var = tk.StringVar(value="Phase: --")
        ttk.Label(progress_frame, textvariable=self.phase_var).pack(anchor=tk.W)

        self.progress = ttk.Progressbar(progress_frame, mode='determinate')
        self.progress.pack(fill=tk.X, pady=(5, 5))

        self.detail_var = tk.StringVar(value="--")
        ttk.Label(progress_frame, textvariable=self.detail_var).pack(anchor=tk.W)

        # Stats frame
        stats_frame = ttk.LabelFrame(main_frame, text="Statistics", padding="20")
        stats_frame.pack(fill=tk.BOTH, expand=True)

        stats_inner = ttk.Frame(stats_frame)
        stats_inner.pack(expand=True)

        ttk.Label(stats_inner, text="PDFs Processed:", style='Header.TLabel').grid(row=0, column=0, sticky=tk.W, pady=5, padx=10)
        self.pdfs_done_var = tk.StringVar(value="--")
        ttk.Label(stats_inner, textvariable=self.pdfs_done_var, style='Big.TLabel', foreground='#2196F3').grid(row=0, column=1, sticky=tk.E, pady=5, padx=10)

        ttk.Label(stats_inner, text="Cards Extracted:", style='Header.TLabel').grid(row=1, column=0, sticky=tk.W, pady=5, padx=10)
        self.cards_extracted_var = tk.StringVar(value="--")
        ttk.Label(stats_inner, textvariable=self.cards_extracted_var, style='Big.TLabel', foreground='#E91E63').grid(row=1, column=1, sticky=tk.E, pady=5, padx=10)

        ttk.Label(stats_inner, text="Cards OCR'd:", style='Header.TLabel').grid(row=2, column=0, sticky=tk.W, pady=5, padx=10)
        self.cards_ocr_var = tk.StringVar(value="--")
        ttk.Label(stats_inner, textvariable=self.cards_ocr_var, style='Big.TLabel', foreground='#9C27B0').grid(row=2, column=1, sticky=tk.E, pady=5, padx=10)

        ttk.Label(stats_inner, text="Time Elapsed:", style='Header.TLabel').grid(row=3, column=0, sticky=tk.W, pady=5, padx=10)
        self.time_var = tk.StringVar(value="--")
        ttk.Label(stats_inner, textvariable=self.time_var, style='Big.TLabel', foreground='#4CAF50').grid(row=3, column=1, sticky=tk.E, pady=5, padx=10)

        # Status bar
        self.status_var = tk.StringVar(value="Ready. Please select a constituency folder.")
        status_bar = ttk.Label(main_frame, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W)
        status_bar.pack(fill=tk.X, pady=(10, 0))

    def browse_folder(self):
        folder = filedialog.askdirectory(title="Select Constituency Folder with PDFs")
        if folder:
            self.folder_path_var.set(folder)
            folder_path = Path(folder)
            constituency_name = folder_path.name
            self.constituency_var.set(f"Constituency: {constituency_name}")

            # Count PDFs
            pdf_files = list(folder_path.glob("*.pdf"))
            self.pdf_count_var.set(f"PDFs found: {len(pdf_files)}")

            # Check for existing checkpoint
            checkpoint_path = folder_path.parent / f".{constituency_name}_checkpoint.json"
            if checkpoint_path.exists():
                self.status_var.set(f"Found incomplete session! Click Extract to resume.")
            else:
                self.status_var.set(f"Selected: {constituency_name} ({len(pdf_files)} PDFs)")

    def stop_and_save(self):
        self.stop_requested = True
        self.status_var.set("Stopping... Saving checkpoint...")
        self.stop_btn.config(state=tk.DISABLED)

    def extract_all_voter_cards(self):
        folder_path = self.folder_path_var.get()
        if not folder_path:
            messagebox.showwarning("No Folder", "Please select a constituency folder first.")
            return
        if not Path(folder_path).exists():
            messagebox.showerror("Folder Not Found", "The selected folder does not exist.")
            return

        pdf_files = list(Path(folder_path).glob("*.pdf"))
        if not pdf_files:
            messagebox.showwarning("No PDFs", "No PDF files found in the selected folder.")
            return

        # Check for existing checkpoint
        folder_path_obj = Path(folder_path)
        constituency_name = folder_path_obj.name
        checkpoint_path = folder_path_obj.parent / f".{constituency_name}_checkpoint.json"

        resume_mode = False
        if checkpoint_path.exists():
            result = messagebox.askyesnocancel(
                "Resume Previous Session?",
                f"Found an incomplete session for '{constituency_name}'.\n\n"
                "Yes = Resume from where it stopped\n"
                "No = Start fresh (delete old progress)\n"
                "Cancel = Do nothing"
            )
            if result is None:  # Cancel
                return
            elif result:  # Yes - Resume
                resume_mode = True
            else:  # No - Start fresh
                checkpoint_path.unlink()

        self.stop_requested = False
        self.extract_btn.config(state=tk.DISABLED)
        self.stop_btn.config(state=tk.NORMAL)
        self.progress.config(value=0)

        thread = threading.Thread(target=self.batch_extract_thread, args=(folder_path, resume_mode))
        thread.daemon = True
        thread.start()

    def batch_extract_thread(self, folder_path, resume_mode=False):
        import shutil

        folder_path = Path(folder_path)
        constituency_name = folder_path.name
        pdf_files = sorted(folder_path.glob("*.pdf"))
        total_pdfs = len(pdf_files)

        # Setup checkpoint
        checkpoint_path = folder_path.parent / f".{constituency_name}_checkpoint.json"
        self.checkpoint = CheckpointManager(checkpoint_path)

        # Create temp directory
        temp_dir = folder_path.parent / f".{constituency_name}_temp_cards"
        temp_dir.mkdir(parents=True, exist_ok=True)

        if resume_mode and self.checkpoint.load():
            # Resume from checkpoint
            start_time = time.time()
            elapsed_before = self.checkpoint.data.get('elapsed_before_resume', 0)
            current_phase = self.checkpoint.data['phase']
            pdf_card_info = self.checkpoint.data['extracted_pdfs']
            total_cards = sum(info[0] for info in pdf_card_info.values())

            self.root.after(0, lambda: self.status_var.set(f"Resuming from Phase {current_phase}..."))
            print(f"Resuming from Phase {current_phase}")
        else:
            # Fresh start
            start_time = time.time()
            elapsed_before = 0
            current_phase = 0
            pdf_card_info = {}
            total_cards = 0

            self.checkpoint.data['constituency_name'] = constituency_name
            self.checkpoint.data['folder_path'] = str(folder_path)
            self.checkpoint.data['total_pdfs'] = total_pdfs
            self.checkpoint.data['start_time'] = start_time

        def get_elapsed():
            return time.time() - start_time + elapsed_before

        def format_time(seconds):
            return f"{int(seconds // 60)}m {int(seconds % 60)}s"

        # ===== PHASE 1: Extract cards from all PDFs =====
        if current_phase < 1:
            self.root.after(0, lambda: self.phase_var.set("Phase 1/3: Extracting cards from PDFs..."))

            # Find PDFs not yet extracted
            extracted_pdf_names = set(pdf_card_info.keys())
            pdfs_to_extract = [(str(pdf), str(temp_dir), idx) for idx, pdf in enumerate(pdf_files)
                               if pdf.stem not in extracted_pdf_names]

            completed_pdfs = len(extracted_pdf_names)

            if pdfs_to_extract:
                with ProcessPoolExecutor(max_workers=self.num_workers) as executor:
                    futures = {executor.submit(extract_cards_from_single_pdf, arg): arg[0] for arg in pdfs_to_extract}

                    for future in as_completed(futures):
                        if self.stop_requested:
                            executor.shutdown(wait=False, cancel_futures=True)
                            break

                        try:
                            pdf_idx, pdf_name, card_count, output_path = future.result()
                            pdf_card_info[pdf_name] = (card_count, output_path)
                            self.checkpoint.add_extracted_pdf(pdf_name, card_count, output_path)
                            completed_pdfs += 1
                            total_cards += card_count

                            elapsed_str = format_time(get_elapsed())
                            progress_pct = int((completed_pdfs / total_pdfs) * 100)

                            self.root.after(0, lambda c=completed_pdfs, t=total_pdfs, pct=progress_pct, tc=total_cards, e=elapsed_str: (
                                self.pdfs_done_var.set(f"{c}/{t}"),
                                self.cards_extracted_var.set(f"{tc:,}"),
                                self.progress.config(value=pct),
                                self.detail_var.set(f"Extracted {c}/{t} PDFs..."),
                                self.time_var.set(e)
                            ))

                            # Save checkpoint every 5 PDFs
                            if completed_pdfs % 5 == 0:
                                self.checkpoint.data['elapsed_before_resume'] = get_elapsed()
                                self.checkpoint.save()

                        except Exception as e:
                            print(f"Error: {e}")
                            completed_pdfs += 1

            if self.stop_requested:
                self.checkpoint.data['phase'] = 0
                self.checkpoint.data['elapsed_before_resume'] = get_elapsed()
                self.checkpoint.save()
                self.root.after(0, lambda: self.on_stopped())
                return

            print(f"\nPhase 1 complete: Extracted {total_cards} cards from {total_pdfs} PDFs")
            self.checkpoint.data['extracted_pdfs'] = pdf_card_info
            self.checkpoint.update_phase(1)
            current_phase = 1

        # ===== PHASE 2: OCR all cards =====
        if current_phase < 2:
            self.root.after(0, lambda: self.phase_var.set("Phase 2/3: OCR processing all cards..."))
            self.root.after(0, lambda: self.progress.config(value=0))

            # Collect all card images
            all_cards = []
            global_idx = 0

            for pdf_name, (card_count, output_path) in pdf_card_info.items():
                if output_path:
                    pdf_card_dir = Path(output_path)
                    jpg_files = sorted(pdf_card_dir.glob("*.jpg"), key=lambda x: int(x.stem) if x.stem.isdigit() else 0)
                    for jpg_file in jpg_files:
                        all_cards.append((jpg_file, global_idx, pdf_name))
                        global_idx += 1

            self.checkpoint.set_all_cards(all_cards)
            total_cards_to_ocr = len(all_cards)

            # Get already completed OCR results
            ocr_results = self.checkpoint.get_ocr_results()
            completed_indices = set(ocr_results.keys())
            cards_to_ocr = [(jpg, idx, name) for jpg, idx, name in all_cards if idx not in completed_indices]

            print(f"\nPhase 2: {len(completed_indices)} already done, {len(cards_to_ocr)} remaining...")

            completed_ocr = len(completed_indices)

            if cards_to_ocr:
                # Sequential processing (more reliable on Windows)
                for card in cards_to_ocr:
                    if self.stop_requested:
                        break

                    try:
                        global_idx, s_no, data, pdf_name = ocr_single_card(card)
                        ocr_results[global_idx] = (s_no, data, pdf_name)
                        self.checkpoint.add_ocr_result(global_idx, s_no, data, pdf_name)
                        completed_ocr += 1

                        if completed_ocr % 10 == 0 or completed_ocr == total_cards_to_ocr:
                            elapsed_str = format_time(get_elapsed())
                            progress_pct = int((completed_ocr / total_cards_to_ocr) * 100)

                            if completed_ocr > 0:
                                rate = (completed_ocr - len(completed_indices)) / (get_elapsed() - elapsed_before) if get_elapsed() > elapsed_before else 1
                                remaining = (total_cards_to_ocr - completed_ocr) / rate if rate > 0 else 0
                                remaining_str = f"~{format_time(remaining)} remaining"
                            else:
                                remaining_str = ""

                            self.root.after(0, lambda c=completed_ocr, t=total_cards_to_ocr, pct=progress_pct, e=elapsed_str, r=remaining_str: (
                                self.cards_ocr_var.set(f"{c:,}/{t:,}"),
                                self.progress.config(value=pct),
                                self.detail_var.set(f"OCR: {c:,}/{t:,} cards... {r}"),
                                self.time_var.set(e)
                            ))

                        # Save checkpoint every 100 cards
                        if completed_ocr % 100 == 0:
                            self.checkpoint.data['elapsed_before_resume'] = get_elapsed()
                            self.checkpoint.save()

                    except Exception as e:
                        print(f"OCR Error: {e}")
                        completed_ocr += 1

            if self.stop_requested:
                self.checkpoint.data['phase'] = 1
                self.checkpoint.data['elapsed_before_resume'] = get_elapsed()
                self.checkpoint.save()
                self.root.after(0, lambda: self.on_stopped())
                return

            print(f"\nPhase 2 complete: OCR'd {completed_ocr} cards")
            self.checkpoint.update_phase(2)
            current_phase = 2
        else:
            # Load from checkpoint
            all_cards = self.checkpoint.get_all_cards()
            ocr_results = self.checkpoint.get_ocr_results()

        # ===== PHASE 2.5: Enhanced OCR for missing data =====
        if current_phase < 3:
            self.root.after(0, lambda: self.phase_var.set("Phase 2.5/3: Fixing missing data..."))
            self.root.after(0, lambda: self.progress.config(value=0))

            field_list = ['voter_id', 'name', 'relation_type', 'relation_name', 'house_no', 'age', 'gender']
            enhanced_done = set(self.checkpoint.data.get('enhanced_ocr_done', []))

            cards_to_fix = []
            for global_idx, (s_no, data, pdf_name) in ocr_results.items():
                if global_idx in enhanced_done:
                    continue
                if data:
                    missing_fields = [f for f in field_list if not data.get(f)]
                    if missing_fields:
                        card_info = all_cards[global_idx]
                        jpg_path = card_info[0]
                        cards_to_fix.append((jpg_path, global_idx, missing_fields))

            if cards_to_fix:
                print(f"\nPhase 2.5: Fixing {len(cards_to_fix)} cards with missing data...")
                total_to_fix = len(cards_to_fix)
                fixed_count = 0

                with ProcessPoolExecutor(max_workers=self.num_workers) as executor:
                    futures = {executor.submit(enhanced_ocr_single_card, card): card[1] for card in cards_to_fix}

                    for future in as_completed(futures):
                        if self.stop_requested:
                            executor.shutdown(wait=False, cancel_futures=True)
                            break

                        try:
                            global_idx, enhanced_data = future.result()
                            self.checkpoint.add_enhanced_ocr(global_idx)

                            if enhanced_data and global_idx in ocr_results:
                                s_no, old_data, pdf_name = ocr_results[global_idx]
                                if old_data:
                                    for field in field_list:
                                        if not old_data.get(field) and enhanced_data.get(field):
                                            old_data[field] = enhanced_data[field]
                                    # Update checkpoint with merged data
                                    self.checkpoint.add_ocr_result(global_idx, s_no, old_data, pdf_name)

                            fixed_count += 1

                            if fixed_count % 20 == 0 or fixed_count == total_to_fix:
                                progress_pct = int((fixed_count / total_to_fix) * 100)
                                elapsed_str = format_time(get_elapsed())

                                self.root.after(0, lambda c=fixed_count, t=total_to_fix, pct=progress_pct, e=elapsed_str: (
                                    self.progress.config(value=pct),
                                    self.detail_var.set(f"Fixing: {c}/{t} cards..."),
                                    self.time_var.set(e)
                                ))

                            # Save checkpoint every 50 fixes
                            if fixed_count % 50 == 0:
                                self.checkpoint.data['elapsed_before_resume'] = get_elapsed()
                                self.checkpoint.save()

                        except Exception as e:
                            print(f"Enhanced OCR Error: {e}")
                            fixed_count += 1

                if self.stop_requested:
                    self.checkpoint.data['phase'] = 2
                    self.checkpoint.data['elapsed_before_resume'] = get_elapsed()
                    self.checkpoint.save()
                    self.root.after(0, lambda: self.on_stopped())
                    return

                print(f"\nPhase 2.5 complete: Fixed {fixed_count} cards")
            else:
                print("\nPhase 2.5: No cards need fixing")

            self.checkpoint.update_phase(3)

        # ===== PHASE 3: Create combined Excel =====
        self.root.after(0, lambda: self.phase_var.set("Phase 3/3: Creating Excel file..."))
        self.root.after(0, lambda: self.progress.config(value=50))

        # Reload OCR results from checkpoint (may have been updated)
        ocr_results = self.checkpoint.get_ocr_results()

        wb = Workbook()
        ws = wb.active
        ws.title = "Voter Data"

        headers = ['S.No', 'Voter ID', 'Name', 'Relation Type', 'Relation Name',
                   'House No', 'Age', 'Gender', 'Constituency']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        row_num = 2
        for global_idx in sorted(ocr_results.keys()):
            s_no, data, pdf_name = ocr_results[global_idx]
            ws.cell(row=row_num, column=1, value=s_no)
            if data:
                ws.cell(row=row_num, column=2, value=data.get('voter_id', ''))
                ws.cell(row=row_num, column=3, value=data.get('name', ''))
                ws.cell(row=row_num, column=4, value=data.get('relation_type', ''))
                ws.cell(row=row_num, column=5, value=data.get('relation_name', ''))
                ws.cell(row=row_num, column=6, value=data.get('house_no', ''))
                ws.cell(row=row_num, column=7, value=data.get('age', ''))
                ws.cell(row=row_num, column=8, value=data.get('gender', ''))
            ws.cell(row=row_num, column=9, value=constituency_name)
            row_num += 1

        column_widths = [8, 15, 25, 12, 25, 15, 8, 10, 30]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width

        excel_filename = f"{constituency_name}_excel.xlsx"
        excel_path = folder_path.parent / excel_filename
        wb.save(excel_path)
        print(f"\nExcel saved: {excel_path}")

        self.root.after(0, lambda: self.progress.config(value=75))

        # ===== CLEANUP =====
        self.root.after(0, lambda: self.detail_var.set("Cleaning up temporary files..."))
        try:
            shutil.rmtree(temp_dir)
            print(f"Deleted temp folder: {temp_dir}")
        except Exception as e:
            print(f"Warning: Could not delete temp folder: {e}")

        # Delete checkpoint file on successful completion
        self.checkpoint.delete()

        # ===== DONE =====
        elapsed_str = format_time(get_elapsed())
        total_cards_final = len(ocr_results)

        self.root.after(0, lambda: self.progress.config(value=100))
        self.root.after(0, lambda e=elapsed_str, ep=str(excel_path), tc=total_cards_final: self.batch_complete(e, ep, tc, total_pdfs))

    def on_stopped(self):
        self.phase_var.set("Stopped - Progress Saved")
        self.detail_var.set("You can resume later by selecting the same folder")
        self.extract_btn.config(state=tk.NORMAL)
        self.stop_btn.config(state=tk.DISABLED)
        self.status_var.set("Stopped. Progress saved. You can resume later.")
        messagebox.showinfo("Stopped", "Progress has been saved.\n\nYou can resume by selecting the same folder and clicking 'Extract All Voter Cards'.")

    def batch_complete(self, elapsed_str, excel_path, total_cards, total_pdfs):
        self.phase_var.set("Complete!")
        self.detail_var.set(f"Processed {total_cards:,} cards from {total_pdfs} PDFs")
        self.time_var.set(elapsed_str)
        self.extract_btn.config(state=tk.NORMAL)
        self.stop_btn.config(state=tk.DISABLED)
        self.status_var.set(f"Done! Excel saved: {Path(excel_path).name}")

        messagebox.showinfo("Batch Extraction Complete",
            f"Successfully processed {total_pdfs} PDFs!\n\n"
            f"Total voter cards: {total_cards:,}\n"
            f"Time taken: {elapsed_str}\n\n"
            f"Excel file saved to:\n{excel_path}")


def main():
    multiprocessing.freeze_support()
    root = tk.Tk()
    app = VoterCounterApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
