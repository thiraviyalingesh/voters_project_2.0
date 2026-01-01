"""
Reprocess voter cards with updated parsing logic - Full extraction
"""
import re
import io
import shutil
from pathlib import Path
from PIL import Image
import pytesseract
import fitz
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import sys

def clean_ocr_text(text):
    if not text:
        return ''
    text = re.sub(r'\s*Photo\s*is\s*', ' ', text, flags=re.IGNORECASE)
    text = re.sub(r'\s*available\s*', ' ', text, flags=re.IGNORECASE)
    text = re.sub(r'^[\s\-–.,:]+|[\s\-–.,:]+$', '', text)
    text = re.sub(r'\s+', ' ', text)
    return text.strip()


def parse_voter_card(text):
    """Parse OCR text from voter card - UPDATED with all relation types."""
    data = {
        'serial_no': '',
        'voter_id': '',
        'name': '',
        'relation_name': '',
        'relation_type': '',
        'house_no': '',
        'age': '',
        'gender': ''
    }

    full_text = text

    # Extract Voter ID
    voter_id_patterns = [
        r'\b([A-Z]{2,3}\d{6,10})\b',
        r'\b([A-Z0-9]{2,3}\d{6,10})\b',
    ]

    for pattern in voter_id_patterns:
        matches = re.findall(pattern, full_text)
        for match in matches:
            if len(match) >= 9:
                data['voter_id'] = match
                break
        if data['voter_id']:
            break

    lines = full_text.split('\n')

    for line in lines:
        line = line.strip()
        if not line:
            continue

        # Extract serial number
        if not data['serial_no']:
            serial_match = re.match(r'^(\d{1,4})\s*$', line)
            if serial_match:
                data['serial_no'] = serial_match.group(1)
            else:
                serial_match = re.match(r'^(\d{1,4})\s+\S', line)
                if serial_match:
                    num = serial_match.group(1)
                    if int(num) < 2000:
                        data['serial_no'] = num

        # Extract name
        if 'பெயர்' in line and ':' in line:
            if 'தந்தை' not in line and 'கணவர்' not in line and 'தாய்' not in line and 'இதரர்' not in line:
                name_part = line.split(':', 1)[-1]
                name_part = clean_ocr_text(name_part)
                if name_part and not data['name']:
                    data['name'] = name_part

        # Extract father's name (handles both தந்தை பெயர் and தந்தையின் பெயர்)
        if ('தந்தை' in line or 'தந்தையின்' in line) and 'பெயர்' in line and ':' in line:
            rel_part = line.split(':', 1)[-1]
            rel_part = clean_ocr_text(rel_part)
            if rel_part and not data['relation_name']:
                data['relation_name'] = rel_part
                data['relation_type'] = 'Father'

        # Extract husband's name (handles both கணவர் பெயர் and கணவரின் பெயர்)
        if ('கணவர்' in line or 'கணவரின்' in line) and ':' in line:
            rel_part = line.split(':', 1)[-1]
            rel_part = clean_ocr_text(rel_part)
            if rel_part and not data['relation_name']:
                data['relation_name'] = rel_part
                data['relation_type'] = 'Husband'

        # Extract mother's name (handles both தாய் பெயர் and தாயின் பெயர்)
        if ('தாய்' in line or 'தாயின்' in line) and 'பெயர்' in line and ':' in line:
            rel_part = line.split(':', 1)[-1]
            rel_part = clean_ocr_text(rel_part)
            if rel_part and not data['relation_name']:
                data['relation_name'] = rel_part
                data['relation_type'] = 'Mother'

        # Extract other's name (இதரர் பெயர் / இதரரின் பெயர்)
        if ('இதரர்' in line or 'இதரரின்' in line) and 'பெயர்' in line and ':' in line:
            rel_part = line.split(':', 1)[-1]
            rel_part = clean_ocr_text(rel_part)
            if rel_part and not data['relation_name']:
                data['relation_name'] = rel_part
                data['relation_type'] = 'Other'

        # Extract house number
        if ('வீட்டு' in line or 'ட்டு' in line) and 'எண்' in line and ':' in line:
            house_part = line.split(':', 1)[-1]
            house_part = clean_ocr_text(house_part)
            if house_part and not data['house_no']:
                data['house_no'] = house_part

        # Extract age
        if 'வயது' in line and ':' in line:
            age_match = re.search(r'வயது\s*:\s*(\d+)', line)
            if age_match:
                data['age'] = age_match.group(1)

        # Extract gender
        if 'பாலினம்' in line:
            if 'ஆண்' in line:
                data['gender'] = 'Male'
            elif 'பெண்' in line:
                data['gender'] = 'Female'

    return data


def extract_and_ocr_pdf(pdf_path, temp_dir):
    """Extract cards from PDF and OCR them."""
    pdf_path = Path(pdf_path)
    pdf_name = pdf_path.stem
    output_path = temp_dir / pdf_name
    output_path.mkdir(parents=True, exist_ok=True)

    results = []

    doc = fitz.open(str(pdf_path))
    num_pages = len(doc)

    start_page = 3
    end_page = num_pages - 1
    card_count = 0

    for page_num in range(start_page, end_page):
        page = doc[page_num]

        zoom = 2
        mat = fitz.Matrix(zoom, zoom)
        pix = page.get_pixmap(matrix=mat)
        img_data = pix.tobytes("png")
        page_img = Image.open(io.BytesIO(img_data))

        page_width, page_height = page_img.size

        num_cols = 3
        num_rows = 10

        header_height = int(page_height * 0.035)
        footer_height = int(page_height * 0.025)
        content_height = page_height - header_height - footer_height

        card_width = page_width // num_cols
        row_height = content_height // num_rows

        for row in range(num_rows):
            for col in range(num_cols):
                x1 = col * card_width
                y1 = header_height + row * row_height
                x2 = x1 + card_width
                y2 = y1 + row_height

                padding = 1
                x1 = max(0, x1 + padding)
                y1 = max(0, y1 + padding)
                x2 = min(page_width, x2 - padding)
                y2 = min(page_height, y2 - padding)

                card_img = page_img.crop((x1, y1, x2, y2))

                card_array = list(card_img.getdata())
                if card_array:
                    avg_brightness = sum(sum(p[:3]) / 3 for p in card_array) / len(card_array)
                    if avg_brightness > 252:
                        continue

                card_count += 1

                # OCR the card directly
                try:
                    text = pytesseract.image_to_string(card_img, lang='tam+eng')
                    data = parse_voter_card(text)
                    results.append((str(card_count), data, pdf_name))
                except Exception as e:
                    print(f"OCR error: {e}")
                    results.append((str(card_count), None, pdf_name))

    doc.close()
    return results


def create_excel(results, output_path, constituency_name):
    """Create Excel file from results."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Voter Data"

    headers = ['S.No', 'Voter ID', 'Name', 'Relation Type', 'Relation Name',
               'House No', 'Age', 'Gender', 'Constituency']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    row_num = 2
    for s_no, data, pdf_name in results:
        ws.cell(row=row_num, column=1, value=s_no)
        if data:
            ws.cell(row=row_num, column=2, value=data.get('voter_id', ''))
            ws.cell(row=row_num, column=3, value=data.get('name', ''))
            ws.cell(row=row_num, column=4, value=data.get('relation_type', ''))
            ws.cell(row=row_num, column=5, value=data.get('relation_name', ''))
            ws.cell(row=row_num, column=6, value=data.get('house_no', ''))
            ws.cell(row=row_num, column=7, value=data.get('age', ''))
            ws.cell(row=row_num, column=8, value=data.get('gender', ''))
        ws.cell(row=row_num, column=9, value=constituency_name)
        row_num += 1

    column_widths = [8, 15, 25, 12, 25, 15, 8, 10, 30]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width

    wb.save(output_path)
    print(f"Excel saved: {output_path}")


if __name__ == "__main__":
    # Process 11-Dr Radhakrishnan Nagar
    constituency = "11-Dr Radhakrishnan Nagar"
    pdf_folder = Path(constituency)

    if not pdf_folder.exists():
        print(f"PDF folder not found: {pdf_folder}")
        sys.exit(1)

    pdf_files = sorted(pdf_folder.glob("*.pdf"))
    print(f"Found {len(pdf_files)} PDFs in {constituency}")

    temp_dir = Path(f".{constituency}_temp_reprocess")
    temp_dir.mkdir(exist_ok=True)

    all_results = []
    for i, pdf_path in enumerate(pdf_files):
        print(f"Processing PDF {i+1}/{len(pdf_files)}: {pdf_path.name}...")
        results = extract_and_ocr_pdf(pdf_path, temp_dir)
        all_results.extend(results)
        print(f"  Extracted {len(results)} cards")

    output_path = Path(f"{constituency}_excel.xlsx")
    create_excel(all_results, output_path, constituency)

    # Count relation types
    relation_counts = {}
    for _, data, _ in all_results:
        if data and data.get('relation_type'):
            rt = data['relation_type']
            relation_counts[rt] = relation_counts.get(rt, 0) + 1

    print(f"\nTotal cards: {len(all_results)}")
    print(f"Relation Type counts:")
    for rt, count in sorted(relation_counts.items()):
        print(f"  {rt}: {count}")

    # Cleanup
    shutil.rmtree(temp_dir)
    print(f"\nCleaned up temp folder")
