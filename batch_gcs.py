"""
Batch Process PDFs via GCS + Document AI Batch Mode

- No page limit
- Cost: $0.01/page (10x cheaper than online)
- Uploads PDFs to GCS, processes via batch API, downloads results

Setup:
1. Create a GCS bucket
2. Add GCS_BUCKET to .env file
"""

import json
import re
import os
import time
import subprocess
from pathlib import Path
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import secrets
import webbrowser
from http.server import HTTPServer, BaseHTTPRequestHandler
from urllib.parse import urlencode, parse_qs, urlparse, quote
import urllib.request
import urllib.error

# Install packages
def install_packages():
    packages = ['openpyxl', 'python-dotenv']
    for pkg in packages:
        try:
            if pkg == 'python-dotenv':
                __import__('dotenv')
            else:
                __import__(pkg)
        except ImportError:
            print(f"Installing {pkg}...")
            subprocess.check_call(['uv', 'pip', 'install', pkg])

install_packages()

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from dotenv import load_dotenv


# ============== OAuth ==============

class OAuthManager:
    SCOPES = ['https://www.googleapis.com/auth/cloud-platform']
    AUTH_URL = 'https://accounts.google.com/o/oauth2/v2/auth'
    TOKEN_URL = 'https://oauth2.googleapis.com/token'
    REDIRECT_PORT = 8089
    REDIRECT_URI = f'http://localhost:{REDIRECT_PORT}/callback'

    def __init__(self, client_id, client_secret):
        self.client_id = client_id
        self.client_secret = client_secret
        self.access_token = None
        self.refresh_token = None
        self.token_expiry = 0
        self.token_file = Path.home() / '.google_docai_oauth_token.json'
        self.load_saved_token()

    def load_saved_token(self):
        if self.token_file.exists():
            try:
                data = json.loads(self.token_file.read_text())
                self.refresh_token = data.get('refresh_token')
                self.access_token = data.get('access_token')
                self.token_expiry = data.get('expiry', 0)
            except:
                pass

    def save_token(self):
        try:
            data = {
                'refresh_token': self.refresh_token,
                'access_token': self.access_token,
                'expiry': self.token_expiry
            }
            self.token_file.write_text(json.dumps(data))
        except:
            pass

    def get_valid_token(self):
        if self.access_token and time.time() < self.token_expiry - 60:
            return self.access_token
        if self.refresh_token:
            if self.refresh_access_token():
                return self.access_token
        if self.do_oauth_flow():
            return self.access_token
        return None

    def refresh_access_token(self):
        try:
            data = urlencode({
                'client_id': self.client_id,
                'client_secret': self.client_secret,
                'refresh_token': self.refresh_token,
                'grant_type': 'refresh_token'
            }).encode()
            req = urllib.request.Request(self.TOKEN_URL, data=data)
            with urllib.request.urlopen(req, timeout=30) as response:
                result = json.loads(response.read().decode())
                self.access_token = result['access_token']
                self.token_expiry = time.time() + result.get('expires_in', 3600)
                self.save_token()
                return True
        except:
            return False

    def do_oauth_flow(self):
        auth_code = None
        state = secrets.token_urlsafe(16)

        class CallbackHandler(BaseHTTPRequestHandler):
            def do_GET(self):
                nonlocal auth_code
                parsed = urlparse(self.path)
                if parsed.path == '/callback':
                    params = parse_qs(parsed.query)
                    if params.get('state', [None])[0] == state:
                        auth_code = params.get('code', [None])[0]
                        self.send_response(200)
                        self.send_header('Content-type', 'text/html')
                        self.end_headers()
                        self.wfile.write(b'<html><body><h1>Success!</h1><p>Close this window.</p></body></html>')
                    else:
                        self.send_response(400)
                        self.end_headers()
                else:
                    self.send_response(404)
                    self.end_headers()
            def log_message(self, format, *args):
                pass

        auth_params = {
            'client_id': self.client_id,
            'redirect_uri': self.REDIRECT_URI,
            'response_type': 'code',
            'scope': ' '.join(self.SCOPES),
            'state': state,
            'access_type': 'offline',
            'prompt': 'consent'
        }
        auth_url = f"{self.AUTH_URL}?{urlencode(auth_params)}"

        server = HTTPServer(('localhost', self.REDIRECT_PORT), CallbackHandler)
        server.timeout = 120

        webbrowser.open(auth_url)

        while auth_code is None:
            server.handle_request()

        server.server_close()

        if not auth_code:
            return False

        try:
            data = urlencode({
                'client_id': self.client_id,
                'client_secret': self.client_secret,
                'code': auth_code,
                'redirect_uri': self.REDIRECT_URI,
                'grant_type': 'authorization_code'
            }).encode()
            req = urllib.request.Request(self.TOKEN_URL, data=data)
            with urllib.request.urlopen(req, timeout=30) as response:
                result = json.loads(response.read().decode())
                self.access_token = result['access_token']
                self.refresh_token = result.get('refresh_token', self.refresh_token)
                self.token_expiry = time.time() + result.get('expires_in', 3600)
                self.save_token()
                return True
        except:
            return False


# ============== GCS Operations ==============

def upload_to_gcs(file_path, bucket_name, blob_name, token, log_func=None):
    """Upload a file to GCS using resumable upload for large files."""
    file_size = os.path.getsize(file_path)

    # For files > 5MB, use resumable upload
    if file_size > 5 * 1024 * 1024:
        return upload_resumable(file_path, bucket_name, blob_name, token, log_func)

    # Simple upload for small files
    url = f"https://storage.googleapis.com/upload/storage/v1/b/{bucket_name}/o?uploadType=media&name={quote(blob_name, safe='')}"

    with open(file_path, 'rb') as f:
        file_content = f.read()

    headers = {
        'Authorization': f'Bearer {token}',
        'Content-Type': 'application/pdf'
    }

    req = urllib.request.Request(url, data=file_content, headers=headers, method='POST')

    with urllib.request.urlopen(req, timeout=300) as response:
        return json.loads(response.read().decode())


def upload_resumable(file_path, bucket_name, blob_name, token, log_func=None):
    """Resumable upload for large files."""
    file_size = os.path.getsize(file_path)

    # Step 1: Initiate resumable upload
    init_url = f"https://storage.googleapis.com/upload/storage/v1/b/{bucket_name}/o?uploadType=resumable&name={quote(blob_name, safe='')}"

    headers = {
        'Authorization': f'Bearer {token}',
        'Content-Type': 'application/json',
        'X-Upload-Content-Type': 'application/pdf',
        'X-Upload-Content-Length': str(file_size)
    }

    req = urllib.request.Request(init_url, data=b'{}', headers=headers, method='POST')

    with urllib.request.urlopen(req, timeout=60) as response:
        upload_url = response.headers.get('Location')

    # Step 2: Upload file in chunks
    chunk_size = 5 * 1024 * 1024  # 5MB chunks

    with open(file_path, 'rb') as f:
        offset = 0
        while offset < file_size:
            chunk = f.read(chunk_size)
            chunk_len = len(chunk)
            end = offset + chunk_len - 1

            headers = {
                'Content-Length': str(chunk_len),
                'Content-Range': f'bytes {offset}-{end}/{file_size}'
            }

            req = urllib.request.Request(upload_url, data=chunk, headers=headers, method='PUT')

            try:
                with urllib.request.urlopen(req, timeout=300) as response:
                    if response.status in [200, 201]:
                        return json.loads(response.read().decode())
            except urllib.error.HTTPError as e:
                if e.code == 308:  # Resume incomplete
                    pass
                else:
                    raise

            offset += chunk_len
            if log_func:
                pct = int((offset / file_size) * 100)
                log_func(f"  Uploading: {pct}%")

    return {'name': blob_name}


def download_from_gcs(bucket_name, blob_name, token):
    """Download a file from GCS."""
    url = f"https://storage.googleapis.com/storage/v1/b/{bucket_name}/o/{quote(blob_name, safe='')}?alt=media"

    headers = {'Authorization': f'Bearer {token}'}
    req = urllib.request.Request(url, headers=headers)

    with urllib.request.urlopen(req, timeout=600) as response:
        return response.read()


def list_gcs_objects(bucket_name, prefix, token):
    """List objects in GCS bucket with prefix."""
    all_items = []
    page_token = None

    while True:
        url = f"https://storage.googleapis.com/storage/v1/b/{bucket_name}/o?prefix={quote(prefix, safe='')}"
        if page_token:
            url += f"&pageToken={page_token}"

        headers = {'Authorization': f'Bearer {token}'}
        req = urllib.request.Request(url, headers=headers)

        with urllib.request.urlopen(req, timeout=60) as response:
            result = json.loads(response.read().decode())
            items = result.get('items', [])
            all_items.extend(items)

            page_token = result.get('nextPageToken')
            if not page_token:
                break

    return all_items


def delete_gcs_object(bucket_name, blob_name, token):
    """Delete an object from GCS."""
    url = f"https://storage.googleapis.com/storage/v1/b/{bucket_name}/o/{quote(blob_name, safe='')}"

    headers = {'Authorization': f'Bearer {token}'}
    req = urllib.request.Request(url, headers=headers, method='DELETE')

    try:
        with urllib.request.urlopen(req, timeout=60):
            return True
    except:
        return False


# ============== Document AI Batch ==============

def start_batch_process(project_id, location, processor_id, gcs_input_prefix, gcs_output_uri, token):
    """Start a Document AI batch processing job for multiple PDFs."""
    url = f"https://{location}-documentai.googleapis.com/v1/projects/{project_id}/locations/{location}/processors/{processor_id}:batchProcess"

    payload = {
        "inputDocuments": {
            "gcsPrefix": {
                "gcsUriPrefix": gcs_input_prefix
            }
        },
        "documentOutputConfig": {
            "gcsOutputConfig": {
                "gcsUri": gcs_output_uri
            }
        }
    }

    headers = {
        'Authorization': f'Bearer {token}',
        'Content-Type': 'application/json'
    }

    req = urllib.request.Request(url, data=json.dumps(payload).encode(), headers=headers, method='POST')

    with urllib.request.urlopen(req, timeout=120) as response:
        result = json.loads(response.read().decode())
        return result.get('name')  # Operation name


def check_operation_status(operation_name, token):
    """Check status of a Document AI operation."""
    # Extract location from operation name
    # Format: projects/xxx/locations/us/operations/xxx
    parts = operation_name.split('/')
    location = parts[3] if len(parts) > 3 else 'us'

    url = f"https://{location}-documentai.googleapis.com/v1/{operation_name}"

    headers = {'Authorization': f'Bearer {token}'}
    req = urllib.request.Request(url, headers=headers)

    with urllib.request.urlopen(req, timeout=60) as response:
        return json.loads(response.read().decode())


def cancel_operation(operation_name, token):
    """Cancel a running operation."""
    parts = operation_name.split('/')
    location = parts[3] if len(parts) > 3 else 'us'

    url = f"https://{location}-documentai.googleapis.com/v1/{operation_name}:cancel"

    headers = {
        'Authorization': f'Bearer {token}',
        'Content-Type': 'application/json'
    }

    req = urllib.request.Request(url, data=b'{}', headers=headers, method='POST')

    try:
        with urllib.request.urlopen(req, timeout=60):
            return True
    except:
        return False


# ============== JSON to Excel ==============

def extract_part_number(filename):
    match = re.search(r'-TAM-(\d+)-WI', filename, re.IGNORECASE)
    return match.group(1) if match else ''


def get_entity_position(entity):
    page_anchor = entity.get('pageAnchor', {})
    page_refs = page_anchor.get('pageRefs', [{}])
    if not page_refs:
        return 0, 0, 0
    page_ref = page_refs[0]
    page = int(page_ref.get('page', 0))
    bounding_poly = page_ref.get('boundingPoly', {})
    vertices = bounding_poly.get('normalizedVertices', [{}])
    if not vertices:
        return page, 0, 0
    x = vertices[0].get('x', 0)
    y = vertices[0].get('y', 0)
    return page, x, y


def group_entities_to_cards(entities):
    """Group entities into voter cards using VoterID as anchor."""
    for entity in entities:
        page, x, y = get_entity_position(entity)
        entity['_page'] = page
        entity['_x'] = x
        entity['_y'] = y
        if x < 0.33:
            entity['_col'] = 0
        elif x < 0.66:
            entity['_col'] = 1
        else:
            entity['_col'] = 2

    voter_ids = [e for e in entities if e.get('type', '').lower() == 'voterid']

    cards = []
    for vid in voter_ids:
        card = {
            'serial_no': '',
            'voter_id': vid.get('mentionText', '').strip().upper(),
            'name': '',
            'relation_name': '',
            'house_no': '',
            'age': '',
            'gender': '',
            'page': vid['_page'],
            'col': vid['_col'],
            '_y_start': vid['_y'],
            '_y_end': vid['_y'] + 0.10
        }
        cards.append(card)

    cards.sort(key=lambda c: (c['page'], c['col'], c['_y_start']))

    for i, card in enumerate(cards):
        for j in range(i + 1, len(cards)):
            if cards[j]['page'] == card['page'] and cards[j]['col'] == card['col']:
                card['_y_end'] = cards[j]['_y_start']
                break

    for entity in entities:
        if entity.get('type', '').lower() == 'voterid':
            continue

        page = entity['_page']
        col = entity['_col']
        y = entity['_y']
        entity_type = entity.get('type', '').lower()
        mention_text = entity.get('mentionText', '').strip()
        mention_text = re.sub(r'\s*[-–]\s*$', '', mention_text)
        mention_text = re.sub(r'^\s*[-–]\s*', '', mention_text)

        best_card = None
        for card in cards:
            if card['page'] == page and card['col'] == col:
                if card['_y_start'] <= y < card['_y_end']:
                    best_card = card
                    break
                elif abs(y - card['_y_start']) < 0.12:
                    best_card = card

        if not best_card:
            candidates = [c for c in cards if c['page'] == page and c['col'] == col]
            if candidates:
                best_card = min(candidates, key=lambda c: abs(c['_y_start'] - y))

        if best_card:
            if entity_type == 'sno':
                best_card['serial_no'] = mention_text
            elif entity_type == 'name':
                best_card['name'] = mention_text
            elif entity_type == 'relativename':
                best_card['relation_name'] = mention_text
            elif entity_type == 'houseno':
                best_card['house_no'] = mention_text
            elif entity_type == 'age':
                age_match = re.search(r'(\d+)', mention_text)
                if age_match:
                    age = int(age_match.group(1))
                    if 18 <= age <= 120:
                        best_card['age'] = str(age)
            elif entity_type == 'sex':
                if 'பெண்' in mention_text or 'பெண' in mention_text:
                    best_card['gender'] = 'Female'
                elif 'ஆண்' in mention_text or 'ஆண' in mention_text:
                    best_card['gender'] = 'Male'

    return cards


def save_cards_to_excel(cards, output_path, source_name):
    """Save cards to Excel."""
    part_no = extract_part_number(source_name)

    wb = Workbook()
    ws = wb.active
    ws.title = "Voter Data"

    headers = ['S.No', 'Part No.', 'Voter S.No', 'Voter ID', 'Name',
               'Relation Name', 'House No', 'Age', 'Gender', 'Page']

    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    missing = {'name': 0, 'age': 0, 'gender': 0, 'voter_id': 0}

    for row_num, card in enumerate(cards, 2):
        ws.cell(row=row_num, column=1, value=row_num - 1)
        ws.cell(row=row_num, column=2, value=part_no)
        ws.cell(row=row_num, column=3, value=card.get('serial_no', ''))

        vid = card.get('voter_id', '')
        vc = ws.cell(row=row_num, column=4, value=vid)
        if not vid:
            vc.fill = yellow_fill
            missing['voter_id'] += 1

        name = card.get('name', '')
        nc = ws.cell(row=row_num, column=5, value=name)
        if not name:
            nc.fill = yellow_fill
            missing['name'] += 1

        ws.cell(row=row_num, column=6, value=card.get('relation_name', ''))
        ws.cell(row=row_num, column=7, value=card.get('house_no', ''))

        age = card.get('age', '')
        ac = ws.cell(row=row_num, column=8, value=age)
        if not age:
            ac.fill = yellow_fill
            missing['age'] += 1

        gender = card.get('gender', '')
        gc = ws.cell(row=row_num, column=9, value=gender)
        if not gender:
            gc.fill = yellow_fill
            missing['gender'] += 1

        ws.cell(row=row_num, column=10, value=card.get('page', ''))

    widths = [8, 10, 10, 15, 25, 25, 12, 8, 10, 8]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width

    wb.save(output_path)
    return missing


# ============== Main App ==============

class BatchGCSProcessor:
    def __init__(self, root):
        self.root = root
        self.root.title("Document AI Batch Processor (GCS)")
        self.root.geometry("900x750")

        self.stop_requested = False
        self.processing = False
        self.current_operation = None

        # Load .env
        env_file = Path(__file__).parent / '.env'
        if env_file.exists():
            load_dotenv(env_file)

        self.client_id = os.getenv('GOOGLE_CLIENT_ID', '')
        self.client_secret = os.getenv('GOOGLE_CLIENT_SECRET', '')
        self.project_id = os.getenv('GOOGLE_PROJECT_ID', '')
        self.processor_id = os.getenv('DOCAI_PROCESSOR_ID', '')
        self.location = os.getenv('DOCAI_LOCATION', 'us')
        self.input_bucket = os.getenv('GCS_INPUT_BUCKET', os.getenv('GCS_BUCKET', ''))
        self.output_bucket = os.getenv('GCS_OUTPUT_BUCKET', os.getenv('GCS_BUCKET', ''))

        self.oauth_manager = None

        self.create_widgets()
        self.check_credentials()

    def create_widgets(self):
        main_frame = ttk.Frame(self.root, padding="15")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Title
        ttk.Label(main_frame, text="Document AI Batch Processor (GCS)",
                  font=('Helvetica', 16, 'bold')).pack(pady=(0, 5))
        ttk.Label(main_frame, text="Uses GCS + Batch API | $0.01/page | No page limit",
                  foreground='#4CAF50', font=('Helvetica', 10, 'bold')).pack(pady=(0, 10))

        # Credentials
        cred_frame = ttk.LabelFrame(main_frame, text="Credentials", padding="10")
        cred_frame.pack(fill=tk.X, pady=(0, 10))

        self.cred_status = tk.StringVar(value="Checking...")
        ttk.Label(cred_frame, textvariable=self.cred_status).pack(side=tk.LEFT)
        ttk.Button(cred_frame, text="Authorize", command=self.authorize).pack(side=tk.RIGHT, padx=5)
        ttk.Button(cred_frame, text="Edit .env", command=self.edit_env).pack(side=tk.RIGHT, padx=5)

        # Settings
        settings_frame = ttk.LabelFrame(main_frame, text="Settings", padding="10")
        settings_frame.pack(fill=tk.X, pady=(0, 10))

        row1 = ttk.Frame(settings_frame)
        row1.pack(fill=tk.X, pady=2)
        ttk.Label(row1, text="Project ID:", width=12).pack(side=tk.LEFT)
        self.project_var = tk.StringVar(value=self.project_id)
        ttk.Entry(row1, textvariable=self.project_var, width=40).pack(side=tk.LEFT, fill=tk.X, expand=True)

        row2 = ttk.Frame(settings_frame)
        row2.pack(fill=tk.X, pady=2)
        ttk.Label(row2, text="Processor ID:", width=12).pack(side=tk.LEFT)
        self.processor_var = tk.StringVar(value=self.processor_id)
        ttk.Entry(row2, textvariable=self.processor_var, width=40).pack(side=tk.LEFT, fill=tk.X, expand=True)

        row3 = ttk.Frame(settings_frame)
        row3.pack(fill=tk.X, pady=2)
        ttk.Label(row3, text="Input Bucket:", width=12).pack(side=tk.LEFT)
        self.input_bucket_var = tk.StringVar(value=self.input_bucket)
        ttk.Entry(row3, textvariable=self.input_bucket_var, width=40).pack(side=tk.LEFT, fill=tk.X, expand=True)

        row3b = ttk.Frame(settings_frame)
        row3b.pack(fill=tk.X, pady=2)
        ttk.Label(row3b, text="Output Bucket:", width=12).pack(side=tk.LEFT)
        self.output_bucket_var = tk.StringVar(value=self.output_bucket)
        ttk.Entry(row3b, textvariable=self.output_bucket_var, width=40).pack(side=tk.LEFT, fill=tk.X, expand=True)

        row4 = ttk.Frame(settings_frame)
        row4.pack(fill=tk.X, pady=2)
        ttk.Label(row4, text="Location:", width=12).pack(side=tk.LEFT)
        self.location_var = tk.StringVar(value=self.location)
        ttk.Combobox(row4, textvariable=self.location_var, values=['us', 'eu'], width=10).pack(side=tk.LEFT)

        # Folder selection
        folder_frame = ttk.LabelFrame(main_frame, text="PDF Folder (Local)", padding="10")
        folder_frame.pack(fill=tk.X, pady=(0, 10))

        self.folder_var = tk.StringVar()
        ttk.Entry(folder_frame, textvariable=self.folder_var, width=60).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        ttk.Button(folder_frame, text="Browse...", command=self.browse_folder).pack(side=tk.LEFT)

        # Info
        info_frame = ttk.Frame(main_frame)
        info_frame.pack(fill=tk.X, pady=(0, 10))

        self.pdf_count_var = tk.StringVar(value="PDFs: --")
        ttk.Label(info_frame, textvariable=self.pdf_count_var).pack(side=tk.LEFT)

        self.cost_var = tk.StringVar(value="Est. Cost: --")
        ttk.Label(info_frame, textvariable=self.cost_var, foreground='#4CAF50',
                  font=('Helvetica', 10, 'bold')).pack(side=tk.RIGHT)

        # Buttons
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(pady=(0, 10))

        self.start_btn = ttk.Button(btn_frame, text="Start Batch Processing", command=self.start_processing)
        self.start_btn.pack(side=tk.LEFT, padx=5)

        self.stop_btn = ttk.Button(btn_frame, text="Cancel", command=self.stop_processing, state=tk.DISABLED)
        self.stop_btn.pack(side=tk.LEFT, padx=5)

        # Progress
        progress_frame = ttk.LabelFrame(main_frame, text="Progress", padding="10")
        progress_frame.pack(fill=tk.X, pady=(0, 10))

        self.phase_var = tk.StringVar(value="Ready")
        ttk.Label(progress_frame, textvariable=self.phase_var, font=('Helvetica', 10, 'bold')).pack(anchor=tk.W)

        self.progress = ttk.Progressbar(progress_frame, mode='determinate')
        self.progress.pack(fill=tk.X, pady=5)

        self.detail_var = tk.StringVar(value="")
        ttk.Label(progress_frame, textvariable=self.detail_var).pack(anchor=tk.W)

        # Stats
        stats_frame = ttk.LabelFrame(main_frame, text="Statistics", padding="10")
        stats_frame.pack(fill=tk.X, pady=(0, 10))

        stats_grid = ttk.Frame(stats_frame)
        stats_grid.pack()

        ttk.Label(stats_grid, text="PDFs:", font=('Helvetica', 10, 'bold')).grid(row=0, column=0, padx=10)
        self.pdfs_var = tk.StringVar(value="0")
        ttk.Label(stats_grid, textvariable=self.pdfs_var, font=('Helvetica', 14, 'bold'),
                  foreground='#2196F3').grid(row=0, column=1, padx=10)

        ttk.Label(stats_grid, text="Cards:", font=('Helvetica', 10, 'bold')).grid(row=0, column=2, padx=10)
        self.cards_var = tk.StringVar(value="0")
        ttk.Label(stats_grid, textvariable=self.cards_var, font=('Helvetica', 14, 'bold'),
                  foreground='#E91E63').grid(row=0, column=3, padx=10)

        ttk.Label(stats_grid, text="Time:", font=('Helvetica', 10, 'bold')).grid(row=0, column=4, padx=10)
        self.time_var = tk.StringVar(value="--")
        ttk.Label(stats_grid, textvariable=self.time_var, font=('Helvetica', 14, 'bold'),
                  foreground='#4CAF50').grid(row=0, column=5, padx=10)

        # Log
        log_frame = ttk.LabelFrame(main_frame, text="Log", padding="5")
        log_frame.pack(fill=tk.BOTH, expand=True)

        self.log_text = tk.Text(log_frame, height=12, font=('Courier', 9))
        log_scroll = ttk.Scrollbar(log_frame, orient=tk.VERTICAL, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=log_scroll.set)
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        log_scroll.pack(side=tk.RIGHT, fill=tk.Y)

    def log(self, msg):
        self.log_text.insert(tk.END, f"{time.strftime('%H:%M:%S')} - {msg}\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()

    def check_credentials(self):
        if self.client_id and self.client_secret:
            self.oauth_manager = OAuthManager(self.client_id, self.client_secret)
            if self.oauth_manager.refresh_token:
                self.cred_status.set(f"Authorized: {self.client_id[:20]}...")
            else:
                self.cred_status.set(f"Client ID set (need auth)")
        else:
            self.cred_status.set("Not configured - Edit .env file")

    def edit_env(self):
        env_file = Path(__file__).parent / '.env'
        if not env_file.exists():
            env_file.write_text(
                "GOOGLE_CLIENT_ID=your_client_id.apps.googleusercontent.com\n"
                "GOOGLE_CLIENT_SECRET=your_secret\n"
                "GOOGLE_PROJECT_ID=your_project\n"
                "DOCAI_PROCESSOR_ID=your_processor_id\n"
                "DOCAI_LOCATION=us\n"
                "GCS_INPUT_BUCKET=voter_bucket_buzz\n"
                "GCS_OUTPUT_BUCKET=output_buzz\n"
            )
        os.startfile(str(env_file)) if os.name == 'nt' else subprocess.run(['xdg-open', str(env_file)])

    def authorize(self):
        env_file = Path(__file__).parent / '.env'
        if env_file.exists():
            load_dotenv(env_file, override=True)
            self.client_id = os.getenv('GOOGLE_CLIENT_ID', '')
            self.client_secret = os.getenv('GOOGLE_CLIENT_SECRET', '')

        if not self.client_id or not self.client_secret:
            messagebox.showerror("Error", "Configure credentials in .env first")
            return

        self.oauth_manager = OAuthManager(self.client_id, self.client_secret)
        self.log("Starting authorization...")

        def do_auth():
            token = self.oauth_manager.get_valid_token()
            if token:
                self.root.after(0, lambda: self.cred_status.set(f"Authorized: {self.client_id[:20]}..."))
                self.root.after(0, lambda: self.log("Authorization successful!"))
            else:
                self.root.after(0, lambda: self.log("Authorization failed"))

        threading.Thread(target=do_auth, daemon=True).start()

    def browse_folder(self):
        folder = filedialog.askdirectory(title="Select Folder with PDFs")
        if folder:
            self.folder_var.set(folder)
            self.count_pdfs()

    def count_pdfs(self):
        folder = Path(self.folder_var.get())
        if not folder.exists():
            return

        pdfs = list(folder.glob('*.pdf')) + list(folder.glob('*.PDF'))
        count = len(pdfs)

        # Estimate pages
        est_pages = count * 25
        cost = est_pages * 0.01  # Batch mode pricing

        self.pdf_count_var.set(f"PDFs: {count}")
        self.cost_var.set(f"Est. Cost: ${cost:.2f} (batch mode)")
        self.log(f"Found {count} PDF files")

    def stop_processing(self):
        self.stop_requested = True
        self.stop_btn.config(state=tk.DISABLED)
        self.phase_var.set("Cancelling...")
        self.log("Cancellation requested...")

        # Cancel the operation if running
        if self.current_operation:
            try:
                token = self.oauth_manager.get_valid_token()
                if token:
                    cancel_operation(self.current_operation, token)
                    self.log("Batch operation cancelled")
            except Exception as e:
                self.log(f"Cancel error: {e}")

    def start_processing(self):
        folder = self.folder_var.get()
        if not folder or not Path(folder).exists():
            messagebox.showerror("Error", "Select a valid folder")
            return

        if not self.oauth_manager:
            messagebox.showerror("Error", "Please authorize first")
            return

        input_bucket = self.input_bucket_var.get()
        output_bucket = self.output_bucket_var.get()
        if not input_bucket or not output_bucket:
            messagebox.showerror("Error", "Enter both input and output bucket names")
            return

        token = self.oauth_manager.get_valid_token()
        if not token:
            messagebox.showerror("Error", "Failed to get token - re-authorize")
            return

        project_id = self.project_var.get()
        processor_id = self.processor_var.get()

        if not project_id or not processor_id:
            messagebox.showerror("Error", "Enter Project ID and Processor ID")
            return

        self.stop_requested = False
        self.processing = True
        self.start_btn.config(state=tk.DISABLED)
        self.stop_btn.config(state=tk.NORMAL)

        threading.Thread(target=self.process_batch, args=(folder, input_bucket, output_bucket, project_id, processor_id), daemon=True).start()

    def process_batch(self, folder, input_bucket, output_bucket, project_id, processor_id):
        """Process all PDFs via GCS batch mode."""
        folder = Path(folder)
        location = self.location_var.get()

        # Get unique PDFs (avoid duplicates from case-insensitive glob)
        pdf_set = {}
        for p in folder.glob('*.pdf'):
            pdf_set[p.name.lower()] = p
        for p in folder.glob('*.PDF'):
            pdf_set[p.name.lower()] = p
        pdfs = sorted(pdf_set.values(), key=lambda x: x.name)
        total = len(pdfs)

        if total == 0:
            self.root.after(0, lambda: messagebox.showerror("Error", "No PDF files found"))
            self.root.after(0, self.reset_ui)
            return

        start_time = time.time()
        job_id = f"batch_{int(time.time())}"
        input_prefix = f"{job_id}/"
        output_prefix = f"{job_id}/"

        try:
            # Phase 1: Upload PDFs to GCS
            self.root.after(0, lambda: self.phase_var.set("Phase 1: Uploading PDFs to GCS..."))
            self.root.after(0, lambda: self.log(f"Uploading {total} PDFs to gs://{input_bucket}/{input_prefix}"))

            for i, pdf_path in enumerate(pdfs):
                if self.stop_requested:
                    raise Exception("Cancelled by user")

                token = self.oauth_manager.get_valid_token()
                blob_name = f"{input_prefix}{pdf_path.name}"

                self.root.after(0, lambda n=pdf_path.name, idx=i+1: (
                    self.detail_var.set(f"Uploading {idx}/{total}: {n}"),
                    self.progress.config(value=(idx/total)*30)
                ))

                upload_to_gcs(pdf_path, input_bucket, blob_name, token)
                self.root.after(0, lambda n=pdf_path.name: self.log(f"  Uploaded: {n}"))

            self.root.after(0, lambda: self.log(f"All {total} PDFs uploaded"))

            # Phase 2: Start batch processing
            self.root.after(0, lambda: self.phase_var.set("Phase 2: Starting batch processing..."))
            self.root.after(0, lambda: self.progress.config(value=35))

            token = self.oauth_manager.get_valid_token()
            gcs_input = f"gs://{input_bucket}/{input_prefix}"
            gcs_output = f"gs://{output_bucket}/{output_prefix}"

            self.root.after(0, lambda: self.log(f"Starting batch job..."))
            self.root.after(0, lambda: self.log(f"  Input: {gcs_input}"))
            self.root.after(0, lambda: self.log(f"  Output: {gcs_output}"))

            operation_name = start_batch_process(project_id, location, processor_id, gcs_input, gcs_output, token)
            self.current_operation = operation_name

            self.root.after(0, lambda: self.log(f"Batch job started: {operation_name.split('/')[-1]}"))

            # Phase 3: Poll for completion
            self.root.after(0, lambda: self.phase_var.set("Phase 3: Processing (this may take a while)..."))

            poll_count = 0
            while True:
                if self.stop_requested:
                    raise Exception("Cancelled by user")

                time.sleep(30)  # Poll every 30 seconds
                poll_count += 1

                token = self.oauth_manager.get_valid_token()
                status = check_operation_status(operation_name, token)

                done = status.get('done', False)
                metadata = status.get('metadata', {})
                state = metadata.get('state', 'RUNNING')

                # Update progress
                individual_statuses = metadata.get('individualProcessStatuses', [])
                completed = sum(1 for s in individual_statuses if s.get('status', {}).get('code', 1) == 0 or 'outputGcsDestination' in s)

                elapsed = time.time() - start_time
                elapsed_str = f"{int(elapsed//60)}m {int(elapsed%60)}s"

                self.root.after(0, lambda c=completed, t=total, e=elapsed_str: (
                    self.detail_var.set(f"Processed: {c}/{t} PDFs | Time: {e}"),
                    self.progress.config(value=35 + (c/t)*50),
                    self.pdfs_var.set(f"{c}/{t}"),
                    self.time_var.set(e)
                ))

                if done:
                    error = status.get('error')
                    if error:
                        raise Exception(f"Batch failed: {error.get('message', 'Unknown error')}")
                    break

                if poll_count % 4 == 0:  # Log every 2 minutes
                    self.root.after(0, lambda s=state, c=completed: self.log(f"  Status: {s}, Completed: {c}/{total}"))

            self.root.after(0, lambda: self.log("Batch processing complete!"))

            # Phase 4: Download and process results
            self.root.after(0, lambda: self.phase_var.set("Phase 4: Downloading results..."))
            self.root.after(0, lambda: self.progress.config(value=85))

            # Create output folder
            output_folder = folder.parent / f"{folder.name}_excel"
            output_folder.mkdir(exist_ok=True)

            # List output files
            token = self.oauth_manager.get_valid_token()
            output_objects = list_gcs_objects(output_bucket, output_prefix, token)
            json_files = [o for o in output_objects if o['name'].endswith('.json')]

            self.root.after(0, lambda c=len(json_files): self.log(f"Found {c} result files"))

            total_cards = 0
            processed_files = 0

            for obj in json_files:
                if self.stop_requested:
                    break

                blob_name = obj['name']
                self.root.after(0, lambda b=blob_name: self.log(f"  Output path: {b}"))

                # Extract original PDF name from path
                # Format: batch_xxx/operation_id/original_filename.pdf/0/document-0.json
                # Or: batch_xxx/original_filename.pdf/0/document-0.json
                parts = blob_name.split('/')
                pdf_name = "unknown"

                # Find the part that contains .pdf
                for part in parts:
                    if '.pdf' in part.lower():
                        # Remove .pdf extension and anything after
                        pdf_name = re.sub(r'\.pdf.*$', '', part, flags=re.IGNORECASE)
                        break

                # If not found, check for -output.json pattern
                if pdf_name == "unknown":
                    for part in parts:
                        if '-output.json' in part.lower():
                            pdf_name = part.replace('-output.json', '').replace('-output.JSON', '')
                            break

                # Still not found, try to find a long alphanumeric part that looks like a filename
                if pdf_name == "unknown":
                    for part in parts:
                        if len(part) > 20 and '-' in part and not part.isdigit():
                            pdf_name = part
                            break

                # Last resort: use third part if exists
                if pdf_name == "unknown" and len(parts) >= 3:
                    pdf_name = parts[2]

                self.root.after(0, lambda n=pdf_name: self.detail_var.set(f"Processing: {n}"))

                try:
                    token = self.oauth_manager.get_valid_token()
                    json_content = download_from_gcs(output_bucket, blob_name, token)
                    document = json.loads(json_content.decode('utf-8'))

                    entities = document.get('entities', [])

                    # Debug: count entity types
                    voter_id_count = sum(1 for e in entities if e.get('type', '').lower() == 'voterid')

                    if entities:
                        cards = group_entities_to_cards(entities)

                        excel_path = output_folder / f"{pdf_name}_excel.xlsx"
                        save_cards_to_excel(cards, excel_path, pdf_name)

                        total_cards += len(cards)
                        processed_files += 1

                        # Capture for lambda
                        _pdf_name = pdf_name
                        _cards = len(cards)
                        _voter_ids = voter_id_count
                        self.root.after(0, lambda n=_pdf_name, c=_cards, v=_voter_ids, tc=total_cards: (
                            self.cards_var.set(f"{tc:,}"),
                            self.log(f"  {n}: {c} cards (VoterIDs detected: {v})")
                        ))
                except Exception as e:
                    self.root.after(0, lambda e=str(e): self.log(f"  Error: {e}"))

            # Complete
            elapsed = time.time() - start_time
            elapsed_str = f"{int(elapsed//60)}m {int(elapsed%60)}s"

            self.root.after(0, lambda: self.phase_var.set("Complete!"))
            self.root.after(0, lambda: self.progress.config(value=100))
            self.root.after(0, lambda e=elapsed_str: self.time_var.set(e))

            self.root.after(0, lambda: self.log(f"\n{'='*40}"))
            self.root.after(0, lambda: self.log("COMPLETE!"))
            self.root.after(0, lambda p=processed_files, t=total: self.log(f"Processed: {p}/{t} PDFs"))
            self.root.after(0, lambda c=total_cards: self.log(f"Total cards: {c:,}"))
            self.root.after(0, lambda e=elapsed_str: self.log(f"Time: {e}"))
            self.root.after(0, lambda o=str(output_folder): self.log(f"Output: {o}"))

            self.root.after(0, lambda: messagebox.showinfo("Complete",
                f"Batch processing complete!\n\n"
                f"PDFs: {processed_files}/{total}\n"
                f"Cards: {total_cards:,}\n"
                f"Time: {elapsed_str}\n\n"
                f"Output: {output_folder}"))

        except Exception as e:
            self.root.after(0, lambda err=str(e): self.log(f"ERROR: {err}"))
            self.root.after(0, lambda err=str(e): messagebox.showerror("Error", err))

        finally:
            self.current_operation = None
            self.root.after(0, self.reset_ui)

    def reset_ui(self):
        self.start_btn.config(state=tk.NORMAL)
        self.stop_btn.config(state=tk.DISABLED)
        self.processing = False


def main():
    root = tk.Tk()
    app = BatchGCSProcessor(root)
    root.mainloop()


if __name__ == "__main__":
    main()
