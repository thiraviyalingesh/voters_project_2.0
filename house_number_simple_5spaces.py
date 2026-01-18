"""
House Number Extraction - SIMPLE 5-SPACES RULE + MULTI-PREPROCESSING
NO cropping! Try ALL preprocessing approaches and show which one works best!

Algorithm:
1. OCR the FULL image with MULTIPLE preprocessing approaches
2. Find line containing "வீட்டு எண் :"
3. Extract text from ":" (colon) until 5 consecutive spaces or "Photo"
4. Show results from ALL approaches so you can see which gets the dash!

Example:
Card shows: வீட்டு எண் : 1-39/10

OCR Results:
- Original:     "139/10"   ❌ (missed dash)
- Contrast 2x:  "1-39/10"  ✅ (got dash!)
- Binarize:     "1-39/10"  ✅ (got dash!)
- Scale 2x:     "1-39/10"  ✅ (got dash!)
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from pathlib import Path
import subprocess
import re

# Install dependencies
def install_packages():
    packages = ['openpyxl', 'pillow', 'pytesseract']
    for pkg in packages:
        try:
            __import__(pkg.replace('-', '_'))
        except ImportError:
            print(f"Installing {pkg}...")
            subprocess.check_call(['uv', 'pip', 'install', pkg])

install_packages()

from openpyxl import load_workbook
from PIL import Image, ImageTk, ImageEnhance
import pytesseract


def extract_house_from_text(text):
    """Extract house number from OCR text using 5-spaces rule."""
    for line in text.split('\n'):
        if ('வீட்டு' in line or 'ட்டு' in line) and 'எண்' in line and ':' in line:
            after_colon = line.split(':', 1)[-1]

            # Method 1: Find where 5+ consecutive spaces occur
            match = re.search(r'\s{5,}', after_colon)
            if match:
                house_part = after_colon[:match.start()].strip()
            else:
                # Method 2: No 5 spaces, look for "Photo" keyword
                if 'Photo' in after_colon or 'photo' in after_colon:
                    house_part = re.split(r'Photo|photo|PHOTO', after_colon)[0].strip()
                else:
                    # Method 3: Take first word only
                    house_part = after_colon.split()[0] if after_colon.split() else after_colon.strip()

            # Clean common OCR mistakes
            house_part = house_part.replace('°', '-')
            house_part = house_part.replace('o', '0')
            house_part = house_part.replace('O', '0')
            house_part = house_part.replace('|', '1')
            house_part = house_part.replace('I', '1')
            house_part = house_part.replace('%', 'A')
            house_part = house_part.replace('$', 'S')

            # Remove Tamil characters at end
            house_part = re.sub(r'[அ-ஔக-னா-ௌ].*', '', house_part).strip()

            # Final cleanup
            house_part = house_part.rstrip('.,;:!?')
            house_part = house_part.lstrip('.,;:!?')

            if house_part and re.search(r'\d', house_part):
                return (house_part, line.strip())

    return ('', '')


def ocr_with_all_approaches(img):
    """
    OCR the image with ALL preprocessing approaches.
    Returns dict with results from each approach.
    """
    approaches = [
        ('Original', lambda i: i),
        ('Contrast 2x', lambda i: ImageEnhance.Contrast(i).enhance(2.0)),
        ('Contrast 3x', lambda i: ImageEnhance.Contrast(i).enhance(3.0)),
        ('Grayscale', lambda i: i.convert('L')),
        ('Gray+Sharp', lambda i: ImageEnhance.Sharpness(i.convert('L')).enhance(2.0)),
        ('Binarize', lambda i: i.convert('L').point(lambda x: 0 if x < 140 else 255, '1')),
        ('Scale 2x', lambda i: i.resize((i.size[0] * 2, i.size[1] * 2), Image.LANCZOS)),
        ('Scale 3x', lambda i: i.resize((i.size[0] * 3, i.size[1] * 3), Image.LANCZOS)),
    ]

    results = {}

    for name, transform in approaches:
        try:
            processed = transform(img)
            text = pytesseract.image_to_string(processed, lang='tam+eng')
            house_no, raw_line = extract_house_from_text(text)
            results[name] = {
                'house_no': house_no,
                'raw_line': raw_line,
                'full_text': text[:500]  # First 500 chars for debugging
            }
        except Exception as e:
            results[name] = {
                'house_no': f'ERROR: {str(e)[:30]}',
                'raw_line': '',
                'full_text': ''
            }

    return results


class HouseNumber5SpacesViewer:
    def __init__(self, root):
        self.root = root
        self.root.title("House Number Viewer - Multi-Preprocessing Comparison")
        self.root.geometry("1600x950")
        self.root.resizable(True, True)

        # Data
        self.excel_path = None
        self.workbook = None
        self.worksheet = None
        self.image_folder = None
        self.all_rows = []
        self.current_index = 0

        self.create_widgets()

    def create_widgets(self):
        # Main container
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Top section - File selection
        file_frame = ttk.LabelFrame(main_frame, text="Load Files", padding="10")
        file_frame.pack(fill=tk.X, pady=(0, 10))

        # Excel file row
        excel_row = ttk.Frame(file_frame)
        excel_row.pack(fill=tk.X, pady=2)
        ttk.Label(excel_row, text="Excel File:", width=12).pack(side=tk.LEFT)
        self.excel_path_var = tk.StringVar()
        ttk.Entry(excel_row, textvariable=self.excel_path_var, width=80).pack(side=tk.LEFT, padx=5)
        ttk.Button(excel_row, text="Browse", command=self.browse_excel).pack(side=tk.LEFT, padx=5)
        ttk.Button(excel_row, text="Load", command=self.load_excel).pack(side=tk.LEFT, padx=5)

        # Image folder row
        img_row = ttk.Frame(file_frame)
        img_row.pack(fill=tk.X, pady=2)
        ttk.Label(img_row, text="Image Folder:", width=12).pack(side=tk.LEFT)
        self.image_folder_var = tk.StringVar()
        ttk.Entry(img_row, textvariable=self.image_folder_var, width=80).pack(side=tk.LEFT, padx=5)
        ttk.Button(img_row, text="Browse", command=self.browse_image_folder).pack(side=tk.LEFT, padx=5)

        # Info row with navigation
        info_row = ttk.Frame(main_frame)
        info_row.pack(fill=tk.X, pady=5)

        self.info_var = tk.StringVar(value="No file loaded")
        ttk.Label(info_row, textvariable=self.info_var, font=('Helvetica', 11, 'bold'),
                 foreground='#2196F3').pack(side=tk.LEFT, padx=10)

        # Navigation buttons
        nav_frame = ttk.Frame(info_row)
        nav_frame.pack(side=tk.RIGHT)
        ttk.Button(nav_frame, text="<<<< Previous", command=self.prev_card,
                  width=15).pack(side=tk.LEFT, padx=5)
        ttk.Button(nav_frame, text="Next >>>>", command=self.next_card,
                  width=15).pack(side=tk.LEFT, padx=5)

        # Excel house number
        excel_frame = ttk.Frame(main_frame)
        excel_frame.pack(fill=tk.X, pady=5)
        ttk.Label(excel_frame, text="House No in Excel:", font=('Helvetica', 11, 'bold')).pack(side=tk.LEFT)
        self.excel_house_var = tk.StringVar(value="--")
        ttk.Label(excel_frame, textvariable=self.excel_house_var,
                 font=('Courier', 16, 'bold'), foreground='#9C27B0').pack(side=tk.LEFT, padx=10)

        # ===== MULTI-PREPROCESSING RESULTS PANEL =====
        results_frame = ttk.LabelFrame(main_frame, text="OCR Results - ALL Preprocessing Approaches (Compare which gets the dash!)", padding="10")
        results_frame.pack(fill=tk.X, pady=(0, 10))

        # Create labels for each approach
        self.approach_labels = {}
        approaches = ['Original', 'Contrast 2x', 'Contrast 3x', 'Grayscale', 'Gray+Sharp', 'Binarize', 'Scale 2x', 'Scale 3x']

        for i, approach in enumerate(approaches):
            row_frame = ttk.Frame(results_frame)
            row_frame.pack(fill=tk.X, pady=2)

            # Approach name
            ttk.Label(row_frame, text=f"{approach}:", width=12,
                     font=('Helvetica', 10, 'bold')).pack(side=tk.LEFT)

            # House number result
            house_var = tk.StringVar(value="--")
            house_label = ttk.Label(row_frame, textvariable=house_var,
                                   font=('Courier', 14, 'bold'), width=20)
            house_label.pack(side=tk.LEFT, padx=5)

            # Raw line
            line_var = tk.StringVar(value="--")
            line_label = ttk.Label(row_frame, textvariable=line_var,
                                  font=('Courier', 9), foreground='#666')
            line_label.pack(side=tk.LEFT, padx=5)

            self.approach_labels[approach] = {
                'house_var': house_var,
                'house_label': house_label,
                'line_var': line_var
            }

        # Image display
        view_frame = ttk.LabelFrame(main_frame, text="FULL VOTER CARD", padding="10")
        view_frame.pack(fill=tk.BOTH, expand=True, pady=10)

        self.full_image_label = ttk.Label(view_frame, text="No image loaded",
                                          anchor=tk.CENTER, background='#f0f0f0')
        self.full_image_label.pack(fill=tk.BOTH, expand=True)

        # Keyboard shortcuts
        self.root.bind('<Left>', lambda e: self.prev_card())
        self.root.bind('<Right>', lambda e: self.next_card())

    def browse_excel(self):
        file_path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if file_path:
            self.excel_path_var.set(file_path)

            excel_path = Path(file_path)
            parent = excel_path.parent
            constituency_name = excel_path.stem.replace('_excel', '').replace('_gpu_excel', '')

            possible_folders = [
                parent / f".{constituency_name}_temp_cards",
                parent / ".test_temp_cards",
            ]

            for folder in possible_folders:
                if folder.exists():
                    self.image_folder_var.set(str(folder))
                    break

    def browse_image_folder(self):
        folder = filedialog.askdirectory(title="Select Image Folder")
        if folder:
            self.image_folder_var.set(folder)

    def load_excel(self):
        excel_path = self.excel_path_var.get()
        if not excel_path or not Path(excel_path).exists():
            messagebox.showerror("Error", "Please select a valid Excel file.")
            return

        image_folder = self.image_folder_var.get()
        if not image_folder or not Path(image_folder).exists():
            messagebox.showerror("Error", "Please select a valid image folder.")
            return

        try:
            self.excel_path = Path(excel_path)
            self.image_folder = Path(image_folder)
            self.workbook = load_workbook(self.excel_path)
            self.worksheet = self.workbook.active

            self.load_all_rows()

            if len(self.all_rows) > 0:
                self.current_index = 0
                self.display_current_card()
            else:
                messagebox.showwarning("Warning", "No data found in Excel")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to load: {e}")
            import traceback
            traceback.print_exc()

    def load_all_rows(self):
        self.all_rows = []

        sno_col = 1
        part_col = 2
        voter_id_col = 3
        name_col = 4
        house_col = 7
        source_folder_col = 11
        card_file_col = 12

        for row_num in range(2, self.worksheet.max_row + 1):
            sno_val = self.worksheet.cell(row=row_num, column=sno_col).value
            part_no = self.worksheet.cell(row=row_num, column=part_col).value
            voter_id = self.worksheet.cell(row=row_num, column=voter_id_col).value
            name = self.worksheet.cell(row=row_num, column=name_col).value
            house_no = self.worksheet.cell(row=row_num, column=house_col).value
            source_folder = self.worksheet.cell(row=row_num, column=source_folder_col).value
            card_file = self.worksheet.cell(row=row_num, column=card_file_col).value

            self.all_rows.append({
                'row_num': row_num,
                'sno': sno_val,
                'part_no': str(part_no) if part_no else '',
                'source_folder': source_folder or '',
                'card_file': card_file or '',
                'house_no': str(house_no).strip() if house_no else '',
                'voter_id': voter_id or '',
                'name': name or ''
            })

    def display_current_card(self):
        if not self.all_rows or self.current_index >= len(self.all_rows):
            return

        row_data = self.all_rows[self.current_index]

        self.info_var.set(
            f"Card {self.current_index + 1} / {len(self.all_rows)} | "
            f"Row: {row_data['row_num']} | "
            f"Voter ID: {row_data['voter_id']} | "
            f"Name: {row_data['name'][:40]}"
        )

        self.excel_house_var.set(row_data['house_no'] if row_data['house_no'] else '???')

        # Reset all approach results
        for approach, labels in self.approach_labels.items():
            labels['house_var'].set("--")
            labels['line_var'].set("--")
            labels['house_label'].configure(foreground='black')

        source_folder = row_data['source_folder']
        card_file = row_data['card_file']

        if not source_folder or not card_file:
            self.full_image_label.configure(image='', text="No image path in Excel")
            return

        image_path = self.image_folder / source_folder / card_file

        if not image_path.exists():
            card_stem = Path(card_file).stem
            for ext in ['.png', '.jpg', '.jpeg']:
                alt_path = self.image_folder / source_folder / f"{card_stem}{ext}"
                if alt_path.exists():
                    image_path = alt_path
                    break

        if not image_path.exists():
            self.full_image_label.configure(image='', text=f"Image not found:\n{image_path}")
            return

        try:
            img = Image.open(image_path)

            # OCR with ALL preprocessing approaches
            results = ocr_with_all_approaches(img)

            # Display results for each approach
            for approach, data in results.items():
                if approach in self.approach_labels:
                    labels = self.approach_labels[approach]
                    house_no = data['house_no']
                    raw_line = data['raw_line']

                    labels['house_var'].set(house_no if house_no else 'NOT FOUND')
                    labels['line_var'].set(f'"{raw_line[:80]}"' if raw_line else '(no line found)')

                    # Color code: Green if has dash, Red if missing dash but has digits
                    if house_no:
                        if '-' in house_no or '/' in house_no:
                            labels['house_label'].configure(foreground='#4CAF50')  # Green - has separator
                        else:
                            labels['house_label'].configure(foreground='#FF9800')  # Orange - no separator
                    else:
                        labels['house_label'].configure(foreground='#F44336')  # Red - not found

            # Display full image
            width, height = img.size
            max_width = 1200
            ratio = max_width / width
            new_size = (int(width * ratio), int(height * ratio))
            img_display = img.resize(new_size, Image.LANCZOS)

            self.full_image_photo = ImageTk.PhotoImage(img_display)
            self.full_image_label.configure(image=self.full_image_photo, text='')

        except Exception as e:
            self.full_image_label.configure(image='', text=f"Error:\n{e}")
            import traceback
            traceback.print_exc()

    def prev_card(self):
        if self.current_index > 0:
            self.current_index -= 1
            self.display_current_card()

    def next_card(self):
        if self.current_index < len(self.all_rows) - 1:
            self.current_index += 1
            self.display_current_card()


def main():
    root = tk.Tk()
    app = HouseNumber5SpacesViewer(root)
    root.mainloop()


if __name__ == "__main__":
    main()
