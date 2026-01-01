"""
Electoral Roll Voter Counter - GUI Application v5.0 (FAST VERSION)
Extracts voter counts from Tamil Nadu Electoral Roll PDFs
Optimized with parallel processing for faster OCR
Supports batch processing of entire constituency folders
Features checkpoint/resume capability for interrupted sessions

v5.0 Features:
- All v4.0 features
- ~3x faster OCR with tessdata_fast models
- Optimized Tesseract flags (--psm 6 --oem 1)
- Reduced image resolution (zoom 1.5 instead of 2)
- Streamlined preprocessing (2 approaches instead of 5)
"""

import re
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from pathlib import Path
import threading
import subprocess
import multiprocessing
from concurrent.futures import ProcessPoolExecutor, as_completed
import time
import json

# Only install packages in main process (not in worker processes)
def install_packages():
    packages = ['pymupdf', 'pytesseract', 'pillow', 'openpyxl']
    for pkg in packages:
        try:
            __import__(pkg.replace('-', '_'))
        except ImportError:
            print(f"Installing {pkg}...")
            subprocess.check_call(['uv', 'pip', 'install', pkg])

# Check if this is the main process (not a worker)
if multiprocessing.current_process().name == 'MainProcess':
    install_packages()

import fitz
from PIL import Image, ImageEnhance
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

try:
    import pytesseract
except ImportError:
    subprocess.check_call(['uv', 'pip', 'install', 'pytesseract'])
    import pytesseract

# Tesseract speed config: --psm 6 (single block), --oem 1 (LSTM only)
TESS_CONFIG = '--psm 6 --oem 1'


def extract_part_number(pdf_name):
    """
    Extract part number from PDF filename.

    Examples:
    - '2026-EROLLGEN-S22-11-SIR-DraftRoll-Revision1-TAM-1-WI' → 1
    - '2026-EROLLGEN-S22-11-SIR-DraftRoll-Revision1-TAM-2-WI' → 2
    - '2026-EROLLGEN-S22-11-SIR-DraftRoll-Revision1-TAM-15-WI' → 15

    Pattern: Look for number between '-TAM-' and '-WI' or at the end before -WI
    """
    if not pdf_name:
        return ''

    # Try pattern: -TAM-{number}-WI
    match = re.search(r'-TAM-(\d+)-WI', pdf_name, re.IGNORECASE)
    if match:
        return match.group(1)

    # Try pattern: -{number}-WI at the end
    match = re.search(r'-(\d+)-WI$', pdf_name, re.IGNORECASE)
    if match:
        return match.group(1)

    # Try pattern: just get the last number before -WI
    match = re.search(r'(\d+)[^0-9]*WI', pdf_name, re.IGNORECASE)
    if match:
        return match.group(1)

    # Fallback: try to get any trailing number
    match = re.search(r'-(\d+)$', pdf_name)
    if match:
        return match.group(1)

    return ''


# Worker functions for parallel processing (must be at module level for pickling)
def ocr_single_card(args):
    """Worker function to OCR a single voter card image."""
    jpg_path, global_idx, pdf_name = args
    try:
        img = Image.open(jpg_path)
        text = pytesseract.image_to_string(img, lang='tam+eng', config=TESS_CONFIG)
        data = parse_voter_card_standalone(text)
        # Handle both Path objects and strings
        if hasattr(jpg_path, 'stem'):
            stem = jpg_path.stem
        else:
            stem = Path(jpg_path).stem
        return global_idx, stem, data, pdf_name
    except Exception as e:
        print(f"OCR error for {jpg_path}: {e}")
        if hasattr(jpg_path, 'stem'):
            stem = jpg_path.stem
        else:
            try:
                stem = Path(jpg_path).stem
            except:
                stem = str(global_idx)
        return global_idx, stem, None, pdf_name


def enhanced_ocr_age_gender(args):
    """Worker function for enhanced OCR focused on Age and Gender only - FAST version."""
    jpg_path, global_idx = args
    try:
        # Handle both Path objects and strings
        img = Image.open(str(jpg_path) if hasattr(jpg_path, 'stem') else jpg_path)

        # Crop bottom portion where Age/Gender appears (tighter crop = faster)
        width, height = img.size
        bottom_crop = img.crop((0, int(height * 0.70), width, height))

        result = {'age': '', 'gender': ''}

        # Single approach: high contrast on bottom crop only (1 OCR call instead of 4)
        try:
            processed_img = ImageEnhance.Contrast(bottom_crop).enhance(2.5)
            text = pytesseract.image_to_string(processed_img, lang='tam+eng', config=TESS_CONFIG)

            # Extract age
            age_match = re.search(r'வயது\s*:\s*(\d+)', text)
            if age_match:
                result['age'] = age_match.group(1)

            # Extract gender
            if 'பாலினம்' in text:
                if 'ஆண்' in text:
                    result['gender'] = 'Male'
                elif 'பெண்' in text:
                    result['gender'] = 'Female'
        except:
            pass

        return global_idx, result
    except Exception as e:
        print(f"Enhanced OCR error for {jpg_path}: {e}")
        return global_idx, None


def enhanced_ocr_single_card(args):
    """Worker function for enhanced OCR with early stopping."""
    jpg_path, global_idx, missing_fields = args
    try:
        img = Image.open(jpg_path)

        # Reduced to 2 most effective preprocessing approaches (was 4)
        approaches = [
            ('contrast', lambda i: ImageEnhance.Contrast(i).enhance(2.0)),
            ('binarize', lambda i: i.convert('L').point(lambda x: 0 if x < 140 else 255, '1')),
        ]

        merged = {
            'serial_no': '',
            'voter_id': '',
            'name': '',
            'relation_name': '',
            'relation_type': '',
            'house_no': '',
            'age': '',
            'gender': ''
        }

        for name, transform in approaches:
            try:
                processed_img = transform(img)
                text = pytesseract.image_to_string(processed_img, lang='tam+eng', config=TESS_CONFIG)
                result = parse_voter_card_standalone(text)

                if result:
                    # Merge non-empty values
                    for key in merged:
                        if not merged[key] and result.get(key):
                            merged[key] = result[key]

                # Check if all missing fields are now filled
                all_found = all(merged.get(f) for f in missing_fields)
                if all_found:
                    break
            except:
                continue

        return global_idx, merged
    except Exception as e:
        print(f"Enhanced OCR error for {jpg_path}: {e}")
        return global_idx, None


def parse_voter_card_standalone(text):
    """Parse OCR text from voter card to extract structured data (standalone function)."""
    data = {
        'serial_no': '',
        'voter_id': '',
        'name': '',
        'relation_name': '',
        'relation_type': '',
        'house_no': '',
        'age': '',
        'gender': ''
    }

    full_text = text

    # Extract Voter ID
    voter_id_patterns = [
        r'\b([A-Z]{2,3}\d{6,10})\b',
        r'\b([A-Z0-9]{2,3}\d{6,10})\b',
        r'\b(\d{2}[^\d\s]{1,2}\d{6,10})\b',
        r'(\d{1,3}\s+[A-Z0-9]{2,3}\d{6,10})',
    ]

    for pattern in voter_id_patterns:
        matches = re.findall(pattern, full_text)
        for match in matches:
            id_match = re.search(r'([A-Z0-9]{2,3}\d{6,10})$|(\d{2}[^\d\s]{1,2}\d{6,10})$', match)
            if id_match:
                voter_id = id_match.group(1) or id_match.group(2)
                if voter_id and len(voter_id) >= 9:
                    data['voter_id'] = voter_id
                    break
        if data['voter_id']:
            break

    lines = full_text.split('\n')

    for line in lines:
        line = line.strip()
        if not line:
            continue

        # Extract serial number
        if not data['serial_no']:
            serial_match = re.match(r'^(\d{1,4})\s*$', line)
            if serial_match:
                data['serial_no'] = serial_match.group(1)
            else:
                serial_match = re.match(r'^(\d{1,4})\s+\S', line)
                if serial_match:
                    num = serial_match.group(1)
                    if int(num) < 2000:
                        data['serial_no'] = num

        # Extract name
        if 'பெயர்' in line and ':' in line:
            if 'தந்தை' not in line and 'கணவர்' not in line:
                name_part = line.split(':', 1)[-1]
                name_part = clean_ocr_text_standalone(name_part)
                if name_part and not data['name']:
                    data['name'] = name_part

        # Extract father's name
        if ('தந்தை' in line or 'தந்தையின்' in line) and 'பெயர்' in line and ':' in line:
            rel_part = line.split(':', 1)[-1]
            rel_part = clean_ocr_text_standalone(rel_part)
            if rel_part and not data['relation_name']:
                data['relation_name'] = rel_part
                data['relation_type'] = 'Father'

        # Extract husband's name
        if ('கணவர்' in line or 'கணவரின்' in line) and ':' in line:
            rel_part = line.split(':', 1)[-1]
            rel_part = clean_ocr_text_standalone(rel_part)
            if rel_part and not data['relation_name']:
                data['relation_name'] = rel_part
                data['relation_type'] = 'Husband'

        # Extract mother's name
        if ('தாய்' in line or 'தாயின்' in line) and 'பெயர்' in line and ':' in line:
            rel_part = line.split(':', 1)[-1]
            rel_part = clean_ocr_text_standalone(rel_part)
            if rel_part and not data['relation_name']:
                data['relation_name'] = rel_part
                data['relation_type'] = 'Mother'

        # Extract other's name
        if ('இதரர்' in line or 'இதரரின்' in line) and 'பெயர்' in line and ':' in line:
            rel_part = line.split(':', 1)[-1]
            rel_part = clean_ocr_text_standalone(rel_part)
            if rel_part and not data['relation_name']:
                data['relation_name'] = rel_part
                data['relation_type'] = 'Other'

        # Extract house number
        if ('வீட்டு' in line or 'ட்டு' in line) and 'எண்' in line and ':' in line:
            house_part = line.split(':', 1)[-1]
            house_part = clean_ocr_text_standalone(house_part)
            if house_part and not data['house_no']:
                data['house_no'] = house_part

        # Extract age
        if 'வயது' in line and ':' in line:
            age_match = re.search(r'வயது\s*:\s*(\d+)', line)
            if age_match:
                data['age'] = age_match.group(1)

        # Extract gender
        if 'பாலினம்' in line:
            if 'ஆண்' in line:
                data['gender'] = 'Male'
            elif 'பெண்' in line:
                data['gender'] = 'Female'

    return data


def clean_ocr_text_standalone(text):
    """Clean common OCR artifacts from text (standalone function)."""
    if not text:
        return ''
    text = re.sub(r'\s*Photo\s*is\s*', ' ', text, flags=re.IGNORECASE)
    text = re.sub(r'\s*available\s*', ' ', text, flags=re.IGNORECASE)
    text = re.sub(r'^[\s\-–.,:]+|[\s\-–.,:]+$', '', text)
    text = re.sub(r'\s+', ' ', text)
    return text.strip()


def extract_cards_from_single_pdf(args):
    """Worker function to extract cards from a single PDF."""
    pdf_path, temp_base_dir, pdf_index = args
    try:
        pdf_path = Path(pdf_path)
        pdf_name = pdf_path.stem
        output_path = Path(temp_base_dir) / pdf_name
        output_path.mkdir(parents=True, exist_ok=True)

        doc = fitz.open(str(pdf_path))
        num_pages = len(doc)
        card_count = 0

        start_page = 3
        end_page = num_pages - 1

        for page_num in range(start_page, end_page):
            page = doc[page_num]

            # Reduced zoom from 2 to 1.5 for faster processing
            zoom = 1.5
            mat = fitz.Matrix(zoom, zoom)
            pix = page.get_pixmap(matrix=mat)
            img_data = pix.tobytes("png")
            page_img = Image.open(io.BytesIO(img_data))

            page_width, page_height = page_img.size

            num_cols = 3
            num_rows = 10

            header_height = int(page_height * 0.035)
            footer_height = int(page_height * 0.025)
            content_height = page_height - header_height - footer_height

            card_width = page_width // num_cols
            row_height = content_height // num_rows

            for row in range(num_rows):
                for col in range(num_cols):
                    x1 = col * card_width
                    y1 = header_height + row * row_height
                    x2 = x1 + card_width
                    y2 = y1 + row_height

                    padding = 1
                    x1 = max(0, x1 + padding)
                    y1 = max(0, y1 + padding)
                    x2 = min(page_width, x2 - padding)
                    y2 = min(page_height, y2 - padding)

                    card_img = page_img.crop((x1, y1, x2, y2))

                    card_array = list(card_img.getdata())
                    if card_array:
                        avg_brightness = sum(sum(p[:3]) / 3 for p in card_array) / len(card_array)
                        if avg_brightness > 252:
                            continue

                    card_count += 1
                    card_filename = output_path / f"{card_count}.jpg"
                    card_img.save(card_filename, "JPEG", quality=95)

        doc.close()
        return pdf_index, pdf_name, card_count, str(output_path)
    except Exception as e:
        print(f"Error extracting from {pdf_path}: {e}")
        return pdf_index, Path(pdf_path).stem, 0, None


class CheckpointManager:
    """Manages saving and loading checkpoint data for resume capability."""

    def __init__(self, checkpoint_path):
        self.checkpoint_path = Path(checkpoint_path)
        self.data = {
            'phase': 0,
            'constituency_name': '',
            'folder_path': '',
            'total_pdfs': 0,
            'extracted_pdfs': {},
            'ocr_results': {},
            'enhanced_ocr_done': [],
            'all_cards': [],
            'start_time': 0,
            'elapsed_before_resume': 0
        }

    def exists(self):
        return self.checkpoint_path.exists()

    def load(self):
        if self.exists():
            try:
                # Get file size to estimate load time
                file_size = self.checkpoint_path.stat().st_size
                size_mb = file_size / (1024 * 1024)
                print(f"Loading checkpoint ({size_mb:.1f} MB)...")

                with open(self.checkpoint_path, 'r', encoding='utf-8') as f:
                    self.data = json.load(f)

                cards_count = len(self.data.get('all_cards', []))
                ocr_count = len(self.data.get('ocr_results', {}))
                print(f"Loaded checkpoint: Phase {self.data['phase']}, {cards_count:,} cards, {ocr_count:,} OCR results")
                return True
            except json.JSONDecodeError as e:
                print(f"Checkpoint file corrupted: {e}")
                print("Deleting corrupted checkpoint and starting fresh...")
                self.delete()
                return False
        return False

    def save(self):
        # Write to temp file first, then rename (atomic operation to prevent corruption)
        temp_path = self.checkpoint_path.with_suffix('.tmp')
        try:
            with open(temp_path, 'w', encoding='utf-8') as f:
                json.dump(self.data, f, ensure_ascii=False, indent=2)
            # Replace old checkpoint with new one
            temp_path.replace(self.checkpoint_path)
            print(f"Checkpoint saved: Phase {self.data['phase']}")
        except Exception as e:
            print(f"Error saving checkpoint: {e}")
            if temp_path.exists():
                temp_path.unlink()

    def delete(self):
        if self.exists():
            self.checkpoint_path.unlink()
            print("Checkpoint deleted")

    def update_phase(self, phase):
        self.data['phase'] = phase
        self.save()

    def add_extracted_pdf(self, pdf_name, card_count, output_path):
        self.data['extracted_pdfs'][pdf_name] = (card_count, output_path)

    def add_ocr_result(self, global_idx, s_no, data_dict, pdf_name):
        self.data['ocr_results'][str(global_idx)] = (s_no, data_dict, pdf_name)

    def add_enhanced_ocr(self, global_idx):
        if global_idx not in self.data['enhanced_ocr_done']:
            self.data['enhanced_ocr_done'].append(global_idx)

    def set_all_cards(self, all_cards):
        self.data['all_cards'] = [(str(p), idx, name) for p, idx, name in all_cards]

    def get_all_cards(self):
        # Keep paths as strings for speed, convert to Path only when needed
        return self.data.get('all_cards', [])

    def get_ocr_results(self):
        return {int(k): tuple(v) for k, v in self.data['ocr_results'].items()}


class VoterCounterApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Electoral Roll Voter Counter v5.0 (FAST - Batch)")
        self.root.geometry("850x800")
        self.root.resizable(True, True)

        # Use 10 parallel workers
        self.num_workers = 10
        print(f"Using {self.num_workers} worker processes for parallel processing")

        self.checkpoint = None
        self.stop_requested = False

        self.style = ttk.Style()
        self.style.configure('Title.TLabel', font=('Helvetica', 16, 'bold'))
        self.style.configure('Header.TLabel', font=('Helvetica', 12, 'bold'))
        self.style.configure('Big.TLabel', font=('Helvetica', 24, 'bold'))

        self.create_widgets()

    def create_widgets(self):
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)

        title_label = ttk.Label(
            main_frame,
            text="Tamil Nadu Electoral Roll\nVoter Counter v5.0 (Batch Mode)",
            style='Title.TLabel',
            justify=tk.CENTER
        )
        title_label.pack(pady=(0, 15))

        # Folder selection frame
        folder_frame = ttk.LabelFrame(main_frame, text="Select Constituency Folder", padding="10")
        folder_frame.pack(fill=tk.X, pady=(0, 10))

        self.folder_path_var = tk.StringVar()
        self.folder_entry = ttk.Entry(folder_frame, textvariable=self.folder_path_var, width=60)
        self.folder_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))

        browse_btn = ttk.Button(folder_frame, text="Browse...", command=self.browse_folder)
        browse_btn.pack(side=tk.LEFT)

        # Info frame
        info_frame = ttk.Frame(main_frame)
        info_frame.pack(fill=tk.X, pady=(0, 10))

        self.pdf_count_var = tk.StringVar(value="PDFs found: --")
        ttk.Label(info_frame, textvariable=self.pdf_count_var).pack(side=tk.LEFT)

        self.constituency_var = tk.StringVar(value="Constituency: --")
        ttk.Label(info_frame, textvariable=self.constituency_var).pack(side=tk.RIGHT)

        # Button frame
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(pady=(0, 10))

        self.extract_btn = ttk.Button(btn_frame, text="Extract All Voter Cards", command=self.extract_all_voter_cards)
        self.extract_btn.pack(side=tk.LEFT, padx=5)

        self.stop_btn = ttk.Button(btn_frame, text="Stop & Save", command=self.stop_and_save, state=tk.DISABLED)
        self.stop_btn.pack(side=tk.LEFT, padx=5)

        # Progress frame
        progress_frame = ttk.LabelFrame(main_frame, text="Progress", padding="10")
        progress_frame.pack(fill=tk.X, pady=(0, 10))

        self.phase_var = tk.StringVar(value="Phase: --")
        ttk.Label(progress_frame, textvariable=self.phase_var).pack(anchor=tk.W)

        self.progress = ttk.Progressbar(progress_frame, mode='determinate')
        self.progress.pack(fill=tk.X, pady=(5, 5))

        self.detail_var = tk.StringVar(value="--")
        ttk.Label(progress_frame, textvariable=self.detail_var).pack(anchor=tk.W)

        # Stats frame
        stats_frame = ttk.LabelFrame(main_frame, text="Statistics", padding="10")
        stats_frame.pack(fill=tk.X, pady=(0, 10))

        stats_inner = ttk.Frame(stats_frame)
        stats_inner.pack(expand=True)

        ttk.Label(stats_inner, text="PDFs Processed:", style='Header.TLabel').grid(row=0, column=0, sticky=tk.W, pady=3, padx=10)
        self.pdfs_done_var = tk.StringVar(value="--")
        ttk.Label(stats_inner, textvariable=self.pdfs_done_var, style='Big.TLabel', foreground='#2196F3').grid(row=0, column=1, sticky=tk.E, pady=3, padx=10)

        ttk.Label(stats_inner, text="Cards Extracted:", style='Header.TLabel').grid(row=1, column=0, sticky=tk.W, pady=3, padx=10)
        self.cards_extracted_var = tk.StringVar(value="--")
        ttk.Label(stats_inner, textvariable=self.cards_extracted_var, style='Big.TLabel', foreground='#E91E63').grid(row=1, column=1, sticky=tk.E, pady=3, padx=10)

        ttk.Label(stats_inner, text="Cards OCR'd:", style='Header.TLabel').grid(row=2, column=0, sticky=tk.W, pady=3, padx=10)
        self.cards_ocr_var = tk.StringVar(value="--")
        ttk.Label(stats_inner, textvariable=self.cards_ocr_var, style='Big.TLabel', foreground='#9C27B0').grid(row=2, column=1, sticky=tk.E, pady=3, padx=10)

        ttk.Label(stats_inner, text="Time Elapsed:", style='Header.TLabel').grid(row=3, column=0, sticky=tk.W, pady=3, padx=10)
        self.time_var = tk.StringVar(value="--")
        ttk.Label(stats_inner, textvariable=self.time_var, style='Big.TLabel', foreground='#4CAF50').grid(row=3, column=1, sticky=tk.E, pady=3, padx=10)

        # Data Quality frame
        quality_frame = ttk.LabelFrame(main_frame, text="Data Quality (Age & Gender)", padding="10")
        quality_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))

        quality_grid = ttk.Frame(quality_frame)
        quality_grid.pack(fill=tk.X, pady=(0, 5))

        ttk.Label(quality_grid, text="Complete:", style='Header.TLabel').grid(row=0, column=0, sticky=tk.W, padx=5)
        self.complete_rows_var = tk.StringVar(value="--")
        ttk.Label(quality_grid, textvariable=self.complete_rows_var, foreground='#4CAF50', font=('Helvetica', 12, 'bold')).grid(row=0, column=1, sticky=tk.W, padx=5)

        ttk.Label(quality_grid, text="Missing:", style='Header.TLabel').grid(row=0, column=2, sticky=tk.W, padx=5)
        self.missing_rows_var = tk.StringVar(value="--")
        ttk.Label(quality_grid, textvariable=self.missing_rows_var, foreground='#F44336', font=('Helvetica', 12, 'bold')).grid(row=0, column=3, sticky=tk.W, padx=5)

        ttk.Label(quality_grid, text="Completeness:", style='Header.TLabel').grid(row=0, column=4, sticky=tk.W, padx=5)
        self.completeness_var = tk.StringVar(value="--")
        ttk.Label(quality_grid, textvariable=self.completeness_var, foreground='#2196F3', font=('Helvetica', 12, 'bold')).grid(row=0, column=5, sticky=tk.W, padx=5)

        self.quality_text = tk.Text(quality_frame, height=6, width=80, font=('Courier', 9))
        self.quality_text.pack(fill=tk.BOTH, expand=True)
        self.quality_text.config(state=tk.DISABLED)

        # Status bar
        self.status_var = tk.StringVar(value="Ready. Please select a constituency folder.")
        status_bar = ttk.Label(main_frame, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W)
        status_bar.pack(fill=tk.X, pady=(10, 0))

    def browse_folder(self):
        folder = filedialog.askdirectory(title="Select Constituency Folder with PDFs")
        if folder:
            self.folder_path_var.set(folder)
            folder_path = Path(folder)
            constituency_name = folder_path.name
            self.constituency_var.set(f"Constituency: {constituency_name}")

            pdf_files = list(folder_path.glob("*.pdf"))
            self.pdf_count_var.set(f"PDFs found: {len(pdf_files)}")

            checkpoint_path = folder_path.parent / f".{constituency_name}_checkpoint.json"
            if checkpoint_path.exists():
                self.status_var.set(f"Found incomplete session! Click Extract to resume.")
            else:
                self.status_var.set(f"Selected: {constituency_name} ({len(pdf_files)} PDFs)")

    def stop_and_save(self):
        self.stop_requested = True
        self.status_var.set("Stopping... Saving checkpoint...")
        self.stop_btn.config(state=tk.DISABLED)

    def update_quality_text(self, text):
        self.quality_text.config(state=tk.NORMAL)
        self.quality_text.delete(1.0, tk.END)
        self.quality_text.insert(tk.END, text)
        self.quality_text.config(state=tk.DISABLED)

    def extract_all_voter_cards(self):
        folder_path = self.folder_path_var.get()
        if not folder_path:
            messagebox.showwarning("No Folder", "Please select a constituency folder first.")
            return
        if not Path(folder_path).exists():
            messagebox.showerror("Folder Not Found", "The selected folder does not exist.")
            return

        pdf_files = list(Path(folder_path).glob("*.pdf"))
        if not pdf_files:
            messagebox.showwarning("No PDFs", "No PDF files found in the selected folder.")
            return

        folder_path_obj = Path(folder_path)
        constituency_name = folder_path_obj.name
        checkpoint_path = folder_path_obj.parent / f".{constituency_name}_checkpoint.json"

        resume_mode = False
        if checkpoint_path.exists():
            result = messagebox.askyesnocancel(
                "Resume Previous Session?",
                f"Found an incomplete session for '{constituency_name}'.\n\n"
                "Yes = Resume from where it stopped\n"
                "No = Start fresh (delete old progress)\n"
                "Cancel = Do nothing"
            )
            if result is None:
                return
            elif result:
                resume_mode = True
            else:
                checkpoint_path.unlink()

        self.stop_requested = False
        self.extract_btn.config(state=tk.DISABLED)
        self.stop_btn.config(state=tk.NORMAL)
        self.progress.config(value=0)

        thread = threading.Thread(target=self.batch_extract_thread, args=(folder_path, resume_mode))
        thread.daemon = True
        thread.start()

    def batch_extract_thread(self, folder_path, resume_mode=False):
        folder_path = Path(folder_path)
        constituency_name = folder_path.name
        pdf_files = sorted(folder_path.glob("*.pdf"))
        total_pdfs = len(pdf_files)

        checkpoint_path = folder_path.parent / f".{constituency_name}_checkpoint.json"
        self.checkpoint = CheckpointManager(checkpoint_path)

        # Image folder - NOT deleted after processing
        image_dir = folder_path.parent / f".{constituency_name}_temp_cards"
        image_dir.mkdir(parents=True, exist_ok=True)

        if resume_mode and self.checkpoint.load():
            start_time = time.time()
            elapsed_before = self.checkpoint.data.get('elapsed_before_resume', 0)
            current_phase = self.checkpoint.data['phase']
            pdf_card_info = self.checkpoint.data['extracted_pdfs']
            total_cards = sum(info[0] for info in pdf_card_info.values())

            self.root.after(0, lambda: self.status_var.set(f"Resuming from Phase {current_phase}..."))
        else:
            start_time = time.time()
            elapsed_before = 0
            current_phase = 0
            pdf_card_info = {}
            total_cards = 0

            self.checkpoint.data['constituency_name'] = constituency_name
            self.checkpoint.data['folder_path'] = str(folder_path)
            self.checkpoint.data['total_pdfs'] = total_pdfs
            self.checkpoint.data['start_time'] = start_time

        def get_elapsed():
            return time.time() - start_time + elapsed_before

        def format_time(seconds):
            return f"{int(seconds // 60)}m {int(seconds % 60)}s"

        # ===== PHASE 1: Extract cards from PDFs =====
        if current_phase < 1:
            self.root.after(0, lambda: self.phase_var.set("Phase 1/4: Extracting cards from PDFs..."))

            extracted_pdf_names = set(pdf_card_info.keys())
            pdfs_to_extract = [(str(pdf), str(image_dir), idx) for idx, pdf in enumerate(pdf_files)
                               if pdf.stem not in extracted_pdf_names]

            completed_pdfs = len(extracted_pdf_names)

            if pdfs_to_extract:
                with ProcessPoolExecutor(max_workers=self.num_workers) as executor:
                    futures = {executor.submit(extract_cards_from_single_pdf, arg): arg[0] for arg in pdfs_to_extract}

                    for future in as_completed(futures):
                        if self.stop_requested:
                            executor.shutdown(wait=False, cancel_futures=True)
                            break

                        try:
                            pdf_idx, pdf_name, card_count, output_path = future.result()
                            pdf_card_info[pdf_name] = (card_count, output_path)
                            self.checkpoint.add_extracted_pdf(pdf_name, card_count, output_path)
                            completed_pdfs += 1
                            total_cards += card_count

                            elapsed_str = format_time(get_elapsed())
                            progress_pct = int((completed_pdfs / total_pdfs) * 100)

                            self.root.after(0, lambda c=completed_pdfs, t=total_pdfs, pct=progress_pct, tc=total_cards, e=elapsed_str: (
                                self.pdfs_done_var.set(f"{c}/{t}"),
                                self.cards_extracted_var.set(f"{tc:,}"),
                                self.progress.config(value=pct),
                                self.detail_var.set(f"Extracted {c}/{t} PDFs..."),
                                self.time_var.set(e)
                            ))

                            if completed_pdfs % 5 == 0:
                                self.checkpoint.data['elapsed_before_resume'] = get_elapsed()
                                self.checkpoint.save()

                        except Exception as e:
                            print(f"Error: {e}")
                            completed_pdfs += 1

            if self.stop_requested:
                self.checkpoint.data['phase'] = 0
                self.checkpoint.data['elapsed_before_resume'] = get_elapsed()
                self.checkpoint.save()
                self.root.after(0, lambda: self.on_stopped())
                return

            self.checkpoint.data['extracted_pdfs'] = pdf_card_info

            # Pre-build card list at end of Phase 1 for faster resume
            self.root.after(0, lambda: self.detail_var.set("Building card index for checkpoint..."))
            all_cards_phase1 = []
            global_idx = 0
            for pdf_name, (card_count, output_path) in pdf_card_info.items():
                if output_path:
                    pdf_card_dir = Path(output_path)
                    try:
                        jpg_files = [f for f in pdf_card_dir.iterdir() if f.suffix.lower() == '.jpg']
                        jpg_files.sort(key=lambda x: int(x.stem) if x.stem.isdigit() else 0)
                        for jpg_file in jpg_files:
                            all_cards_phase1.append((jpg_file, global_idx, pdf_name))
                            global_idx += 1
                    except:
                        pass
            self.checkpoint.set_all_cards(all_cards_phase1)

            self.checkpoint.update_phase(1)
            current_phase = 1

        # ===== PHASE 2: OCR all cards =====
        if current_phase < 2:
            # Check if card list is already in checkpoint (skip scanning)
            saved_cards = self.checkpoint.get_all_cards()
            if saved_cards and len(saved_cards) > 0:
                self.root.after(0, lambda: self.detail_var.set(f"Loaded {len(saved_cards):,} cards from checkpoint"))
                all_cards = saved_cards
            else:
                self.root.after(0, lambda: self.phase_var.set("Phase 2/4: Scanning card images..."))
                self.root.after(0, lambda: self.progress.config(value=0))
                self.root.after(0, lambda: self.detail_var.set("Building card list from extracted images..."))

                all_cards = []
                global_idx = 0
                total_dirs = len(pdf_card_info)
                scanned_dirs = 0

                for pdf_name, (card_count, output_path) in pdf_card_info.items():
                    if output_path:
                        pdf_card_dir = Path(output_path)
                        # Use os.scandir for faster file listing
                        try:
                            jpg_files = []
                            for entry in pdf_card_dir.iterdir():
                                if entry.suffix.lower() == '.jpg':
                                    jpg_files.append(entry)
                            # Sort by numeric filename
                            jpg_files.sort(key=lambda x: int(x.stem) if x.stem.isdigit() else 0)
                            for jpg_file in jpg_files:
                                all_cards.append((jpg_file, global_idx, pdf_name))
                                global_idx += 1
                        except Exception as e:
                            print(f"Error scanning {output_path}: {e}")

                    scanned_dirs += 1
                    if scanned_dirs % 20 == 0 or scanned_dirs == total_dirs:
                        progress_pct = int((scanned_dirs / total_dirs) * 100)
                        self.root.after(0, lambda s=scanned_dirs, t=total_dirs, p=progress_pct, c=len(all_cards): (
                            self.progress.config(value=p),
                            self.detail_var.set(f"Scanning folders: {s}/{t} ({c:,} cards found)")
                        ))

                self.checkpoint.set_all_cards(all_cards)

            self.root.after(0, lambda: self.phase_var.set("Phase 2/4: OCR processing all cards..."))
            total_cards_to_ocr = len(all_cards)

            # Load OCR results with progress feedback
            self.root.after(0, lambda: self.detail_var.set("Loading previous OCR results..."))
            print("Loading OCR results from checkpoint...")

            # Faster: use pre-built set from checkpoint data directly
            ocr_data = self.checkpoint.data.get('ocr_results', {})
            completed_indices = set(int(k) for k in ocr_data.keys())
            print(f"Loaded {len(completed_indices):,} completed OCR results")

            self.root.after(0, lambda c=len(completed_indices): self.detail_var.set(f"Building work queue ({c:,} already done)..."))
            print("Building cards to OCR list...")

            # Build list of cards that still need OCR (keep as strings - PIL accepts string paths)
            cards_to_ocr = [(p, idx, name) for p, idx, name in all_cards if idx not in completed_indices]
            print(f"Cards to OCR: {len(cards_to_ocr):,}")

            # Build ocr_results dict only for cards we need to update
            ocr_results = {int(k): (v[0], v[1], v[2]) for k, v in ocr_data.items()}

            completed_ocr = len(completed_indices)

            if cards_to_ocr:
                # Parallel OCR processing - submit in batches to avoid memory issues
                print(f"Starting OCR with {self.num_workers} workers...")
                self.root.after(0, lambda: self.detail_var.set("Starting OCR workers..."))

                batch_size = 500  # Submit 500 tasks at a time
                card_iter = iter(cards_to_ocr)

                with ProcessPoolExecutor(max_workers=self.num_workers) as executor:
                    # Start with initial batch
                    futures = {}
                    for _ in range(min(batch_size, len(cards_to_ocr))):
                        try:
                            card = next(card_iter)
                            futures[executor.submit(ocr_single_card, card)] = card[1]
                        except StopIteration:
                            break

                    while futures:
                        if self.stop_requested:
                            executor.shutdown(wait=False, cancel_futures=True)
                            break

                        # Wait for any future to complete
                        done_futures = []
                        for future in list(futures.keys()):
                            if future.done():
                                done_futures.append(future)

                        if not done_futures:
                            # No futures done yet, wait a bit
                            time.sleep(0.01)
                            continue

                        for future in done_futures:
                            try:
                                global_idx, s_no, data, pdf_name = future.result()
                                ocr_results[global_idx] = (s_no, data, pdf_name)
                                self.checkpoint.add_ocr_result(global_idx, s_no, data, pdf_name)
                                completed_ocr += 1

                                if completed_ocr % 50 == 0 or completed_ocr == total_cards_to_ocr:
                                    elapsed_str = format_time(get_elapsed())
                                    progress_pct = int((completed_ocr / total_cards_to_ocr) * 100)

                                    if completed_ocr > len(completed_indices):
                                        rate = (completed_ocr - len(completed_indices)) / max(1, get_elapsed() - elapsed_before)
                                        remaining = (total_cards_to_ocr - completed_ocr) / max(1, rate)
                                        remaining_str = f"~{format_time(remaining)} left"
                                    else:
                                        remaining_str = ""

                                    self.root.after(0, lambda c=completed_ocr, t=total_cards_to_ocr, pct=progress_pct, e=elapsed_str, r=remaining_str: (
                                        self.cards_ocr_var.set(f"{c:,}/{t:,}"),
                                        self.progress.config(value=pct),
                                        self.detail_var.set(f"OCR: {c:,}/{t:,} cards... {r}"),
                                        self.time_var.set(e)
                                    ))

                                if completed_ocr % 200 == 0:
                                    self.checkpoint.data['elapsed_before_resume'] = get_elapsed()
                                    self.checkpoint.save()

                            except Exception as e:
                                print(f"OCR Error: {e}")
                                completed_ocr += 1

                            # Remove completed future and add new one
                            del futures[future]
                            try:
                                card = next(card_iter)
                                futures[executor.submit(ocr_single_card, card)] = card[1]
                            except StopIteration:
                                pass

            if self.stop_requested:
                self.checkpoint.data['phase'] = 1
                self.checkpoint.data['elapsed_before_resume'] = get_elapsed()
                self.checkpoint.save()
                self.root.after(0, lambda: self.on_stopped())
                return

            self.checkpoint.update_phase(2)
            current_phase = 2
        else:
            all_cards = self.checkpoint.get_all_cards()
            ocr_results = self.checkpoint.get_ocr_results()

        # ===== PHASE 3: Enhanced OCR for missing Age/Gender ONLY =====
        if current_phase < 3:
            self.root.after(0, lambda: self.phase_var.set("Phase 3/4: Fixing missing Age/Gender..."))
            self.root.after(0, lambda: self.progress.config(value=0))

            enhanced_done = set(self.checkpoint.data.get('enhanced_ocr_done', []))

            cards_to_fix = []
            for global_idx, (s_no, data, pdf_name) in ocr_results.items():
                if global_idx in enhanced_done:
                    continue
                if data:
                    missing_age = not data.get('age')
                    missing_gender = not data.get('gender')
                    if missing_age or missing_gender:
                        card_info = all_cards[global_idx]
                        jpg_path = card_info[0]
                        cards_to_fix.append((jpg_path, global_idx))

            if cards_to_fix:
                total_to_fix = len(cards_to_fix)
                fixed_count = 0

                with ProcessPoolExecutor(max_workers=self.num_workers) as executor:
                    futures = {executor.submit(enhanced_ocr_age_gender, card): card[1] for card in cards_to_fix}

                    for future in as_completed(futures):
                        if self.stop_requested:
                            executor.shutdown(wait=False, cancel_futures=True)
                            break

                        try:
                            global_idx, enhanced_data = future.result()
                            self.checkpoint.add_enhanced_ocr(global_idx)

                            if enhanced_data and global_idx in ocr_results:
                                s_no, old_data, pdf_name = ocr_results[global_idx]
                                if old_data:
                                    if not old_data.get('age') and enhanced_data.get('age'):
                                        old_data['age'] = enhanced_data['age']
                                    if not old_data.get('gender') and enhanced_data.get('gender'):
                                        old_data['gender'] = enhanced_data['gender']
                                    self.checkpoint.add_ocr_result(global_idx, s_no, old_data, pdf_name)

                            fixed_count += 1

                            if fixed_count % 20 == 0 or fixed_count == total_to_fix:
                                progress_pct = int((fixed_count / total_to_fix) * 100)
                                elapsed_str = format_time(get_elapsed())

                                self.root.after(0, lambda c=fixed_count, t=total_to_fix, pct=progress_pct, e=elapsed_str: (
                                    self.progress.config(value=pct),
                                    self.detail_var.set(f"Fixing: {c}/{t} cards..."),
                                    self.time_var.set(e)
                                ))

                            if fixed_count % 50 == 0:
                                self.checkpoint.data['elapsed_before_resume'] = get_elapsed()
                                self.checkpoint.save()

                        except Exception as e:
                            print(f"Enhanced OCR Error: {e}")
                            fixed_count += 1

                if self.stop_requested:
                    self.checkpoint.data['phase'] = 2
                    self.checkpoint.data['elapsed_before_resume'] = get_elapsed()
                    self.checkpoint.save()
                    self.root.after(0, lambda: self.on_stopped())
                    return

            self.checkpoint.update_phase(3)

        # ===== PHASE 4: Create Excel =====
        self.root.after(0, lambda: self.phase_var.set("Phase 4/4: Creating Excel file..."))
        self.root.after(0, lambda: self.progress.config(value=50))

        ocr_results = self.checkpoint.get_ocr_results()

        wb = Workbook()
        ws = wb.active
        ws.title = "Voter Data"

        # Updated headers with Part No. and source tracking columns
        headers = ['S.No', 'Part No.', 'Voter ID', 'Name', 'Relation Type', 'Relation Name',
                   'House No', 'Age', 'Gender', 'Constituency', 'Source Folder', 'Card File']

        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        row_num = 2
        missing_age_count = 0
        missing_gender_count = 0

        for global_idx in sorted(ocr_results.keys()):
            s_no, data, pdf_name = ocr_results[global_idx]

            # Column 1: S.No (image filename)
            ws.cell(row=row_num, column=1, value=s_no)

            # Column 2: Part No. (extracted from PDF filename)
            part_no = extract_part_number(pdf_name)
            ws.cell(row=row_num, column=2, value=part_no)

            if data:
                ws.cell(row=row_num, column=3, value=data.get('voter_id', ''))
                ws.cell(row=row_num, column=4, value=data.get('name', ''))
                ws.cell(row=row_num, column=5, value=data.get('relation_type', ''))
                ws.cell(row=row_num, column=6, value=data.get('relation_name', ''))
                ws.cell(row=row_num, column=7, value=data.get('house_no', ''))

                age_val = data.get('age', '')
                gender_val = data.get('gender', '')

                age_cell = ws.cell(row=row_num, column=8, value=age_val)
                gender_cell = ws.cell(row=row_num, column=9, value=gender_val)

                if not age_val:
                    age_cell.fill = yellow_fill
                    missing_age_count += 1
                if not gender_val:
                    gender_cell.fill = yellow_fill
                    missing_gender_count += 1
            else:
                for col in range(3, 10):
                    cell = ws.cell(row=row_num, column=col, value='')
                    if col in [8, 9]:
                        cell.fill = yellow_fill
                missing_age_count += 1
                missing_gender_count += 1

            ws.cell(row=row_num, column=10, value=constituency_name)

            # Source tracking columns
            ws.cell(row=row_num, column=11, value=pdf_name)  # Source Folder
            ws.cell(row=row_num, column=12, value=f"{s_no}.jpg")  # Card File

            row_num += 1

        # Updated column widths for new layout
        column_widths = [8, 10, 15, 25, 12, 25, 15, 8, 10, 30, 50, 10]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width

        excel_filename = f"{constituency_name}_excel.xlsx"
        excel_path = folder_path.parent / excel_filename
        wb.save(excel_path)

        self.root.after(0, lambda: self.progress.config(value=90))

        # Update quality display
        total_cards_final = len(ocr_results)
        complete = total_cards_final - max(missing_age_count, missing_gender_count)
        missing_total = max(missing_age_count, missing_gender_count)
        completeness = (complete / total_cards_final * 100) if total_cards_final > 0 else 0

        self.root.after(0, lambda: self.complete_rows_var.set(f"{complete:,}"))
        self.root.after(0, lambda: self.missing_rows_var.set(f"{missing_total:,}"))
        self.root.after(0, lambda: self.completeness_var.set(f"{completeness:.1f}%"))

        report = (
            f"Total: {total_cards_final:,}\n"
            f"Missing Age: {missing_age_count:,}\n"
            f"Missing Gender: {missing_gender_count:,}\n"
            f"Completeness: {completeness:.1f}%"
        )
        self.root.after(0, lambda: self.update_quality_text(report))

        # Delete checkpoint (but NOT images!)
        self.checkpoint.delete()

        elapsed_str = format_time(get_elapsed())

        self.root.after(0, lambda: self.progress.config(value=100))
        self.root.after(0, lambda e=elapsed_str, ep=str(excel_path), tc=total_cards_final:
                        self.batch_complete(e, ep, tc, total_pdfs, missing_age_count, missing_gender_count))

    def on_stopped(self):
        self.phase_var.set("Stopped - Progress Saved")
        self.detail_var.set("You can resume later")
        self.extract_btn.config(state=tk.NORMAL)
        self.stop_btn.config(state=tk.DISABLED)
        self.status_var.set("Stopped. Progress saved.")
        messagebox.showinfo("Stopped", "Progress saved. Resume by selecting same folder.")

    def batch_complete(self, elapsed_str, excel_path, total_cards, total_pdfs, missing_age, missing_gender):
        self.phase_var.set("Complete!")
        self.detail_var.set(f"Processed {total_cards:,} cards from {total_pdfs} PDFs")
        self.time_var.set(elapsed_str)
        self.extract_btn.config(state=tk.NORMAL)
        self.stop_btn.config(state=tk.DISABLED)
        self.status_var.set(f"Done! Excel: {Path(excel_path).name} | Images kept!")

        messagebox.showinfo("Extraction Complete",
            f"Processed {total_pdfs} PDFs!\n\n"
            f"Total cards: {total_cards:,}\n"
            f"Missing Age: {missing_age:,}\n"
            f"Missing Gender: {missing_gender:,}\n"
            f"Time: {elapsed_str}\n\n"
            f"Excel: {excel_path}\n\n"
            f"Images are KEPT for future fixes!")


def main():
    multiprocessing.freeze_support()
    root = tk.Tk()
    app = VoterCounterApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
